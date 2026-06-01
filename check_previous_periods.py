import openpyxl

target_file = r"D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx"

print("=" * 80)
print("查看目标文件中之前几期的数据格式")
print("=" * 80)

wb = openpyxl.load_workbook(target_file)
ws = wb['跟随号码统计']

# 查看2026123期或之前的数据
print(f"\n总行数: {ws.max_row}")

# 查找最后几期有数据的行
print("\n=== 查找最后几期的数据 ===")
period_data = {}

for row_idx in range(ws.max_row, 1, -1):
    period = ws.cell(row=row_idx, column=9).value
    if period and period != 'None':
        if period not in period_data:
            period_data[period] = []
        
        # 读取这一行的前4列
        row_nums = []
        for col_idx in range(1, 5):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and val != 'None':
                row_nums.append(str(val))
        
        if row_nums:
            period_data[period].extend(row_nums)
        
        if len(period_data) >= 3:
            break

for period in sorted(period_data.keys(), reverse=True):
    nums = period_data[period]
    print(f"\n期数 {period}: 共 {len(nums)} 个号码")
    print(f"  号码: {nums[:20]}...")  # 只显示前20个

# 查看具体的几行数据
print("\n=== 查看具体行数据（最后20行） ===")
for row_idx in range(max(1, ws.max_row - 19), ws.max_row + 1):
    row_data = []
    for col_idx in range(1, 11):
        val = ws.cell(row=row_idx, column=col_idx).value
        row_data.append(str(val))
    print(f"行{row_idx}: {row_data}")

print("\n" + "=" * 80)
