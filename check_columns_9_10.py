import openpyxl

target_file = r"D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据_backup_correct.xlsx"

print("=" * 80)
print("查看备份文件中第9列和第10列的数据")
print("=" * 80)

wb = openpyxl.load_workbook(target_file)
ws = wb['跟随号码统计']

print(f"\n总行数: {ws.max_row}")

# 查看第9列和第10列的数据分布
print("\n=== 第9列数据（前50行） ===")
col9_values = {}
for row_idx in range(1, min(51, ws.max_row + 1)):
    val = ws.cell(row=row_idx, column=9).value
    if val:
        if str(val) not in col9_values:
            col9_values[str(val)] = 0
        col9_values[str(val)] += 1

for k, v in sorted(col9_values.items())[:20]:
    print(f"  {k}: {v} 行")

print("\n=== 第10列数据（前50行） ===")
col10_values = {}
for row_idx in range(1, min(51, ws.max_row + 1)):
    val = ws.cell(row=row_idx, column=10).value
    if val:
        if str(val) not in col10_values:
            col10_values[str(val)] = 0
        col10_values[str(val)] += 1

for k, v in sorted(col10_values.items())[:20]:
    print(f"  {k}: {v} 行")

# 查看中间部分的数据
print("\n=== 查看第500-520行的第9列和第10列 ===")
for row_idx in range(500, min(521, ws.max_row + 1)):
    col9 = ws.cell(row=row_idx, column=9).value
    col10 = ws.cell(row=row_idx, column=10).value
    print(f"行{row_idx}: 第9列={col9}, 第10列={col10}")

print("\n" + "=" * 80)
