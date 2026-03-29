import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl 库才能生成 Excel。请在终端执行: pip install openpyxl")
    sys.exit(1)

def generate_excel():
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")) 
    data_sum_dir = os.path.join(base_dir, "src", "data-sum")
    history_file = os.path.join(base_dir, "src", "data", "kl8_history_final.txt")

    # 1. 加载历史开奖数据
    actual_draws = {}
    if os.path.exists(history_file):
        with open(history_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line: continue
                m = re.match(r"date:([^,]+),period:([^,]+),numbers:([\d\-]+)", line)
                if m:
                    date_str = m.group(1).replace('-', '')
                    nums = [int(n) for n in m.group(3).split('-')]
                    actual_draws[date_str] = nums

    # 2. 读取手工汇总数据
    expert_dates = []
    file_reading_order = {} # {date: [num1, num2, ...]}
    
    if os.path.exists(data_sum_dir):
        for item in os.listdir(data_sum_dir):
            date_dir = os.path.join(data_sum_dir, item)
            if os.path.isdir(date_dir) and re.match(r"^\d{8}$", item):
                txt_file = os.path.join(date_dir, f"{item}-data.txt")
                if os.path.exists(txt_file):
                    expert_dates.append(item)
                    ordered_nums = []
                    with open(txt_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if not line: continue
                            
                            # 如果有箭头，取箭头后的内容；否则取整行内容
                            targets = line.split('→')[-1]
                            
                            for num_str in targets.strip().split():
                                try:
                                    ordered_nums.append(int(num_str))
                                except: pass
                    file_reading_order[item] = ordered_nums

    if not expert_dates:
        print("未找到任何专家数据。")
        return

    expert_dates.sort(reverse=True) # 最新日期排在上面

    # 3. 生成 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "专家追踪对比"

    # 冻结首行和首列
    ws.freeze_panes = 'B2'

    # 样式定义
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), 
                         right=Side(style='thin', color='DDDDDD'), 
                         top=Side(style='thin', color='DDDDDD'), 
                         bottom=Side(style='thin', color='DDDDDD'))

    hit_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")     # 命中 (绿色)
    miss_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")    # 未中 (浅灰)

    hit_font = Font(color="006400", bold=True)
    miss_font = Font(color="333333")

    # 表头
    ws.cell(row=1, column=1, value="日期 / 序号").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = center_align
    ws.column_dimensions['A'].width = 15

    # 填充号码列的序号 (横轴 1..N)
    max_nums = max([len(v) for v in file_reading_order.values()]) if file_reading_order else 80
    for i in range(1, max_nums + 1):
        cell = ws.cell(row=1, column=i + 1, value=f"#{i}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i + 1)].width = 5

    # 4. 逐行填充数据 (每一行代表一天)
    for row_idx, date in enumerate(expert_dates, start=2):
        actual = set(actual_draws.get(date, []))
        expert_nums = file_reading_order.get(date, [])

        # A 列：日期
        date_cell = ws.cell(row=row_idx, column=1, value=f"{date[:4]}-{date[4:6]}-{date[6:]}")
        date_cell.alignment = center_align
        date_cell.border = thin_border
        date_cell.font = Font(bold=True)

        # 后面跟随：当天的排号
        for col_offset, num in enumerate(expert_nums, start=2):
            cell = ws.cell(row=row_idx, column=col_offset, value=num)
            cell.alignment = center_align
            cell.border = thin_border
            
            if num in actual:
                cell.fill = hit_fill
                cell.font = hit_font
            else:
                cell.fill = miss_fill
                cell.font = miss_font

        ws.row_dimensions[row_idx].height = 25

    # 保存文件
    output_file = os.path.join(data_sum_dir, "每期专家关注号命中追踪.xlsx")
    wb.save(output_file)
    print(f"✅ 命中追踪 Excel 生成成功！文件路径: {output_file}")

if __name__ == "__main__":
    generate_excel()
