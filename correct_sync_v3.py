import openpyxl
from datetime import datetime

print("=" * 80)
print("正确处理2026124期热码数据")
print("=" * 80)

# 文件路径
source_file = r"D:\Dpanqianyi\Python-Project\data\热码统计\20260514-2026124期-热码统计.xlsx"
target_file = r"D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx"

# 1. 读取源文件
print("\n1. 读取热码统计源文件...")
wb_source = openpyxl.load_workbook(source_file)
ws_source = wb_source['Sheet1']

# 提取所有带*的号码（从4个区域）
star_numbers = []
for row_idx in range(3, ws_source.max_row + 1):
    for col in [1, 5, 9, 13]:  # 4个区域的号码列
        cell_value = ws_source.cell(row=row_idx, column=col).value
        if cell_value and '*' in str(cell_value):
            number = str(cell_value).replace('*', '').strip()
            star_numbers.append(number)

# 去重并排序
star_numbers_unique = sorted(list(set(star_numbers)), key=lambda x: int(x))
print(f"   提取到 {len(star_numbers_unique)} 个唯一带*号码")
print(f"   号码列表: {star_numbers_unique}")

# 2. 组织数据为跟随号码统计格式
# 格式：每行4个号码，每5行为一组
print("\n2. 组织数据格式...")
star_rows = []
for i in range(0, len(star_numbers_unique), 4):
    group = star_numbers_unique[i:i+4]
    while len(group) < 4:
        group.append(None)
    star_rows.append(group)

print(f"   共组织成 {len(star_rows)} 行数据")
for idx, row in enumerate(star_rows, 1):
    print(f"   第{idx}行: {row}")

# 3. 读取目标文件
print("\n3. 读取目标文件...")
wb_target = openpyxl.load_workbook(target_file)
ws_target = wb_target['跟随号码统计']
print(f"   当前行数: {ws_target.max_row}")

# 4. 追加数据
# 第1行：期数标识（如 "2026124期-数据1"）
# 第2-最后一行：数据行
# 每行格式：[号码1, 号码2, 号码3, 号码4, None, None, None, None, 序号, None]

print("\n4. 追加数据到目标文件...")
start_row = ws_target.max_row + 1

# 添加期数标识行
ws_target.cell(row=start_row, column=1, value=f"2026124期-数据1")
print(f"   行{start_row}: 期数标识")

# 添加数据行
for idx, row_data in enumerate(star_rows, 1):
    target_row = start_row + idx
    for col_idx, value in enumerate(row_data, 1):
        ws_target.cell(row=target_row, column=col_idx, value=value)
    # 第9列是序号
    ws_target.cell(row=target_row, column=9, value=idx)
    print(f"   行{target_row}: 序号{idx}")

# 5. 保存文件
print("\n5. 保存文件...")
wb_target.save(target_file)
print(f"   已保存至: {target_file}")

# 6. 验证
print("\n6. 验证结果...")
wb_verify = openpyxl.load_workbook(target_file)
ws_verify = wb_verify['跟随号码统计']
print(f"   现在总行数: {ws_verify.max_row}")
print(f"   新增行数: {ws_verify.max_row - start_row + 1}")

# 显示新增的数据
print("\n   新增数据预览:")
for row_idx in range(start_row, min(start_row + 10, ws_verify.max_row + 1)):
    row_data = []
    for col_idx in range(1, 11):
        val = ws_verify.cell(row=row_idx, column=col_idx).value
        row_data.append(str(val))
    print(f"   行{row_idx}: {row_data}")

print("\n" + "=" * 80)
print("✅ 处理完成！")
print("=" * 80)
