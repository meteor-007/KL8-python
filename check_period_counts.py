import openpyxl

target_file = r"D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据_backup_correct.xlsx"

print("=" * 80)
print("查看备份文件中2026123期及之前的数据量")
print("=" * 80)

wb = openpyxl.load_workbook(target_file)
ws = wb['跟随号码统计']

# 查找2026123期的所有数据
print("\n=== 查找2026123期的数据 ===")
period_123_count = 0
period_123_rows = []

for row_idx in range(1, ws.max_row + 1):
    period = ws.cell(row=row_idx, column=9).value
    if str(period) == '2026123':
        period_123_count += 1
        if len(period_123_rows) < 10:  # 只显示前10行
            row_data = []
            for col_idx in range(1, 11):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val))
            period_123_rows.append((row_idx, row_data))

print(f"2026123期共有 {period_123_count} 行数据")
for row_idx, row_data in period_123_rows:
    print(f"行{row_idx}: {row_data}")

# 查看2026122期
print("\n=== 查找2026122期的数据 ===")
period_122_count = 0
for row_idx in range(1, ws.max_row + 1):
    period = ws.cell(row=row_idx, column=9).value
    if str(period) == '2026122':
        period_122_count += 1

print(f"2026122期共有 {period_122_count} 行数据")

# 查看2026120期
print("\n=== 查找2026120期的数据 ===")
period_120_count = 0
for row_idx in range(1, ws.max_row + 1):
    period = ws.cell(row=row_idx, column=9).value
    if str(period) == '2026120':
        period_120_count += 1

print(f"2026120期共有 {period_120_count} 行数据")

print("\n" + "=" * 80)
print("结论：每期应该有多少行数据？")
print("=" * 80)
