import openpyxl

target_file = r"D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据_backup_correct.xlsx"

wb = openpyxl.load_workbook(target_file)
ws = wb['跟随号码统计']

print("=" * 80)
print("查看跟随号码统计sheet的表头和第1行数据")
print("=" * 80)

print(f"\n总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

# 查看第1行（应该是标题或期数标识）
print("\n=== 第1行 ===")
row1_data = []
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col_idx).value
    row1_data.append(str(val))
print(row1_data)

# 查看第2行
print("\n=== 第2行 ===")
row2_data = []
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=2, column=col_idx).value
    row2_data.append(str(val))
print(row2_data)

# 查看第6行（可能是期数标识）
print("\n=== 第6行 ===")
row6_data = []
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=6, column=col_idx).value
    row6_data.append(str(val))
print(row6_data)

# 查看第7行
print("\n=== 第7行 ===")
row7_data = []
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=7, column=col_idx).value
    row7_data.append(str(val))
print(row7_data)

print("\n" + "=" * 80)
