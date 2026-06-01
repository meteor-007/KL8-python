import openpyxl

# 查看2026124期热码统计文件的完整结构
hot_code_file = r"D:\Dpanqianyi\Python-Project\data\热码统计\20260514-2026124期-热码统计.xlsx"

print("=" * 80)
print("重新分析2026124期热码统计文件的真实结构")
print("=" * 80)

wb = openpyxl.load_workbook(hot_code_file)
ws = wb['Sheet1']

print(f"\n总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

# 打印所有行的数据，看看真实结构
print("\n=== 完整数据结构（前30行） ===")
for row_idx in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        row_data.append(str(cell.value))
    print(f"行{row_idx}: {row_data}")

# 查看是否有其他sheet包含数据1、数据2
print("\n=== 所有Sheet名称 ===")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws_temp = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    print(f"行数: {ws_temp.max_row}, 列数: {ws_temp.max_column}")
    if ws_temp.max_row > 0 and ws_temp.max_row <= 50:
        for row_idx in range(1, min(6, ws_temp.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(15, ws_temp.max_column + 1)):
                cell = ws_temp.cell(row=row_idx, column=col_idx)
                row_data.append(str(cell.value))
            print(f"  行{row_idx}: {row_data}")

print("\n" + "=" * 80)
