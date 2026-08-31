# -*- coding: utf-8 -*-
"""
纯净池评分器 (Pure Pool Scorer)
================================
数据源1 B区域 → 去点位 → 去重 = 纯净池
再通过 遗漏回补 + 连续出现 + 双源验证 + 近期命中 评分排序

每期输出：纯净池号码 + 评分 + 推荐星级
"""
import openpyxl
import re
import os
import sys
import collections
import logging
import numpy as np
from typing import Dict, Set, List, Tuple, Optional

logger = logging.getLogger(__name__)

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()


EXCEL_FILE = os.path.join(_PROJ, '跟随+点位+开奖数据.xlsx')


def _parse_b_zone_b_nums(ws, start_row: int) -> Set[int]:
    """解析数据1或数据2的B区域(列6-9)号码"""
    b_nums = set()
    curr = start_row + 1
    for cycle in range(4):
        for _ in range(4):
            if curr > ws.max_row:
                break
            for col in range(6, 10):
                val = str(ws.cell(row=curr, column=col).value or '').strip().replace('*', '')
                if val.isdigit() and 1 <= int(val) <= 80:
                    b_nums.add(int(val))
            curr += 1
        curr += 1
    return b_nums


def load_b_zone_data() -> Dict[str, Dict[str, Set[int]]]:
    """
    加载Excel中数据1和数据2的B区域号码
    
    Returns:
        {period: {'data1': set, 'data2': set}}
    """
    if not os.path.exists(EXCEL_FILE):
        return {}

    result = collections.defaultdict(lambda: {'data1': set(), 'data2': set()})

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        try:
            ws = wb['跟随号码统计']

            markers = []
            for r in range(1, ws.max_row + 1):
                v = str(ws.cell(row=r, column=1).value or '')
                m = re.search(r'(\d+)期数据(1|2)', v)
                if m:
                    markers.append((r, m.group(1), int(m.group(2))))

            for start_row, period, dtype in markers:
                b_nums = _parse_b_zone_b_nums(ws, start_row)
                key = 'data1' if dtype == 1 else 'data2'
                result[period][key] = b_nums
        finally:
            wb.close()
    except Exception as e:
        print(f"[PurePool] 加载Excel失败: {e}")

    return dict(result)


def compute_pure_pool(
    period: str,
    b_zone_data: Dict[str, Dict[str, Set[int]]],
    points: Set[int],
) -> Set[int]:
    """
    计算指定期的纯净池 = 数据1 B区域 - 点位号码
    """
    d1_b = b_zone_data.get(period, {}).get('data1', set())
    return d1_b - points


def compute_number_omission(num: int, history) -> int:
    """计算号码遗漏期数 (history降序，最新在前)

    统一 1 基定义 (修复 off-by-one): 返回"距离上次出现的期数 + 1",
    即号码在最新期出现 → 遗漏=1, 上一期出现 → 遗漏=2, 依此类推,
    与 gap 计数语义一致 (之前返回 0 基 idx, 与遗漏语义差 1)。
    调用方 (score_pure_pool / pure_pool_lr_trainer 特征) 均按相对大小使用, 不受影响。
    """
    for idx, h in enumerate(history):
        if num in h['numbers']:
            return idx + 1
    return len(history)


def compute_consecutive_in_pool(
    num: int,
    period: str,
    b_zone_data: Dict[str, Dict[str, Set[int]]],
    points_map: Dict[str, Set[int]],
    recent_periods: List[str],
) -> int:
    """计算号码在纯净池中连续出现期数"""
    consec = 0
    found_period = False
    for p in recent_periods:
        if p == period:
            found_period = True
        if not found_period:
            continue
        pts = points_map.get(p, set())
        pure = b_zone_data.get(p, {}).get('data1', set()) - pts
        if num in pure:
            consec += 1
        else:
            break
    return consec


def compute_recent_hit_count(num: int, history, n: int = 10) -> int:
    """号码近n期命中次数"""
    count = 0
    for i, h in enumerate(history[:n]):
        if num in h['numbers']:
            count += 1
    return count


def score_pure_pool(
    period: str,
    b_zone_data: Dict[str, Dict[str, Set[int]]],
    points: Set[int],
    points_map: Dict[str, Set[int]],
    history,
    recent_periods: List[str],
) -> List[Dict]:
    """
    对纯净池号码进行评分排序
    
    评分维度:
      - 遗漏回补: 遗漏>=3 +1分, >=6 +2分, >=10 +3分
      - 连续关注: 在纯净池连续出现>=2 +1分, >=4 +2分
      - 双源验证: 同时在数据2 B区域(非点位)出现 +2分
      - 近期活跃: 近10期命中>=3次 +1分
    
    Returns:
        按评分降序排列的号码列表, 每个元素:
        {
            'number': int,
            'score': int,
            'omission': int,
            'consecutive': int,
            'dual_source': bool,
            'recent_hits': int,
            'stars': str,  # *** >=4, ** >=3, * >=2
        }
    """
    d1_b = b_zone_data.get(period, {}).get('data1', set())
    d2_b = b_zone_data.get(period, {}).get('data2', set())
    pure = d1_b - points
    d2_pure = d2_b - points  # 数据2 B区域也去点位

    if not pure:
        return []

    results = []
    for num in pure:
        # 遗漏
        omission = compute_number_omission(num, history)

        # 连续出现
        consec = compute_consecutive_in_pool(
            num, period, b_zone_data, points_map, recent_periods
        )

        # 双源验证
        dual = num in d2_pure

        # 近期命中
        recent_hits = compute_recent_hit_count(num, history, 10)

        # 评分
        score = 0

        # 遗漏回补 (冷号回补力)
        if omission >= 10:
            score += 3
        elif omission >= 6:
            score += 2
        elif omission >= 3:
            score += 1

        # 连续关注 (持续出现=强信号)
        if consec >= 4:
            score += 2
        elif consec >= 2:
            score += 1

        # 双源验证 (两个数据源都指向)
        if dual:
            score += 2

        # 近期活跃
        if recent_hits >= 4:
            score += 2
        elif recent_hits >= 3:
            score += 1

        # 星级
        if score >= 5:
            stars = "***"
        elif score >= 3:
            stars = "**"
        elif score >= 2:
            stars = "*"
        else:
            stars = ""

        results.append({
            'number': num,
            'score': score,
            'omission': omission,
            'consecutive': consec,
            'dual_source': dual,
            'recent_hits': recent_hits,
            'stars': stars,
        })

    # 按评分降序，同分按遗漏升序(冷号优先)
    results.sort(key=lambda x: (-x['score'], x['omission']))

    return results


def _apply_lr_probs(scored: List[Dict], weights: Dict) -> List[Dict]:
    """为池内号码附加 lr_prob；失败则原样返回。"""
    if not scored or not weights:
        return scored
    try:
        from core.pure_pool_lr_trainer import featurize_row, predict_proba
        import numpy as np
        names = weights.get('feature_names') or []
        w = np.asarray(weights.get('weights') or [], dtype=float)
        if len(names) != len(w) or len(w) == 0:
            return scored
        b = float(weights.get('bias', 0.0))
        X = np.asarray([featurize_row(s) for s in scored], dtype=float)
        probs = predict_proba(X, w, b)
        out = []
        for s, p in zip(scored, probs):
            row = dict(s)
            row['lr_prob'] = float(p)
            out.append(row)
        out.sort(key=lambda r: (-r.get('lr_prob', 0), r.get('omission', 0)))
        return out
    except Exception as e:
        logger.warning(f"[PurePool] LR 概率应用失败, 回退规则评分: {e}")
        return scored


def select_high_confidence(
    scored: List[Dict],
    weights: Optional[Dict] = None,
) -> Tuple[List[int], str]:
    """
    选取高置信定胆号码。
    Returns: (号码列表, 来源标签 rule|lr|lr_soft)
    """
    if not scored:
        return [], 'rule'

    use_lr = bool(weights and weights.get('active'))
    if use_lr:
        try:
            from core.pure_pool_lr_trainer import select_picks
            import numpy as np
            w = np.asarray(weights['weights'], dtype=float)
            b = float(weights.get('bias', 0.0))
            delta = float(weights.get('delta', 0.02))
            top_k = int(weights.get('top_k', 3))
            strict = select_picks(scored, w, b, delta, top_k, soft_fallback=False)
            if strict:
                return [p['number'] for p in strict], 'lr'
            soft = select_picks(scored, w, b, delta, top_k, soft_fallback=True)
            if soft:
                return [p['number'] for p in soft], 'lr_soft'
        except Exception as e:
            logger.warning(f"[PurePool] LR 定胆失败, 回退规则: {e}")

    top = [s['number'] for s in scored if s.get('score', 0) >= 3]
    if top:
        return top, 'rule'
    starred = [s['number'] for s in scored if s.get('stars')]
    if starred:
        return starred[:3], 'rule'
    return [s['number'] for s in scored[:3]], 'rule'


def format_pure_pool_report(
    period: str,
    scored: List[Dict],
    b_zone_data: Dict[str, Dict[str, Set[int]]],
    points: Set[int],
    weights: Optional[Dict] = None,
    high_conf: Optional[List[int]] = None,
    high_conf_source: str = 'rule',
) -> str:
    """
    格式化纯净池报告 (Markdown)
    """
    d1_b = b_zone_data.get(period, {}).get('data1', set())
    pure = sorted(d1_b - points)

    scored_view = _apply_lr_probs(scored, weights) if weights else list(scored)

    lines = []
    lines.append(f"### 8. 纯净池定胆 (Pure Pool Scorer)")
    lines.append(f"")
    lines.append(
        f"- **策略逻辑**：数据源1 B区域 → 去点位 → 纯净池 → "
        f"阶跃规则评分 + L2逻辑回归概率（方案1数据驱动权重）"
    )
    lines.append(f"- **纯净池号码**：`{pure}` (共{len(pure)}个)")

    if weights:
        wf = weights.get('wf') or {}
        active = weights.get('active', False)
        lines.append(
            f"- **LR权重**：active=`{active}` | delta=`{weights.get('delta')}` | "
            f"top_k=`{weights.get('top_k')}` | 训练截止=`{weights.get('train_end_issue', '')}`"
        )
        if wf.get('ok'):
            lines.append(
                f"- **WF回测**：Lift=`{wf.get('lr_lift', 0):.2f}x` "
                f"(旧规则 `{wf.get('old_lift', 0):.2f}x`, Δ=`{wf.get('lift_gain_vs_old', 0):+.2f}`) "
                f"| 命中率=`{wf.get('lr_hit_rate', 0):.1%}` | 均码=`{wf.get('lr_avg_size', 0):.1f}` "
                f"| folds=`{wf.get('folds')}`"
            )

    lines.append(f"")
    has_prob = any('lr_prob' in s for s in scored_view)
    if has_prob:
        lines.append(f"| 号码 | 规则分 | LR概率 | 遗漏 | 连续 | 双源 | 近10期 | 推荐 |")
        lines.append(f"|:----:|:------:|:------:|:----:|:----:|:----:|:-----:|:----:|")
        for s in scored_view:
            dual_str = "Y" if s['dual_source'] else "-"
            prob = s.get('lr_prob')
            prob_s = f"{prob:.3f}" if isinstance(prob, float) else "-"
            lines.append(
                f"| {s['number']:02d} | {s['score']} | {prob_s} | {s['omission']} | "
                f"{s['consecutive']} | {dual_str} | {s['recent_hits']} | {s['stars']} |"
            )
    else:
        lines.append(f"| 号码 | 评分 | 遗漏 | 连续 | 双源 | 近10期命中 | 推荐 |")
        lines.append(f"|:----:|:----:|:----:|:----:|:----:|:---------:|:----:|")
        for s in scored_view:
            dual_str = "Y" if s['dual_source'] else "-"
            lines.append(
                f"| {s['number']:02d} | {s['score']} | {s['omission']} | "
                f"{s['consecutive']} | {dual_str} | {s['recent_hits']} | {s['stars']} |"
            )

    # 旧规则对照
    rule_top = [s['number'] for s in scored if s.get('score', 0) >= 3]
    lines.append(f"")
    lines.append(f"- **旧规则高置信 (评分>=3)**：`{rule_top}`")

    # LR 影子 / 主推
    lr_shadow: List[int] = []
    if weights and scored:
        try:
            from core.pure_pool_lr_trainer import select_picks
            import numpy as np
            w = np.asarray(weights['weights'], dtype=float)
            b = float(weights.get('bias', 0.0))
            delta = float(weights.get('delta', 0.02))
            top_k = int(weights.get('top_k', 3))
            lr_shadow = [
                p['number']
                for p in select_picks(scored, w, b, delta, top_k, soft_fallback=True)
            ]
        except Exception:
            lr_shadow = []
    lines.append(f"- **LR定胆 (影子/候选)**：`{lr_shadow}`")

    if high_conf is None:
        high_conf, high_conf_source = select_high_confidence(scored, weights)
    src_map = {'lr': 'LR主推', 'lr_soft': 'LR软回退(P≥25%)', 'rule': '规则主推'}
    src_label = src_map.get(high_conf_source, high_conf_source)
    lines.append(f"- **高置信定胆 ({src_label})**：`{high_conf}`")

    lines.append(f"")
    return "\n".join(lines)


def run_pure_pool_analysis(
    target_period: str,
    history,
    points_map: Dict[str, Set[int]],
) -> Tuple[str, List[Dict]]:
    """完整纯净池分析。返回 (markdown_report, scored_list)。"""
    b_zone_data = load_b_zone_data()

    pts = points_map.get(target_period, set())
    if not pts and history:
        for h in history:
            p = points_map.get(h['issue'], set())
            if p:
                pts = p
                break

    recent_periods = [h['issue'] for h in history[:50]]
    if target_period not in recent_periods:
        recent_periods.insert(0, target_period)

    scored = score_pure_pool(
        target_period, b_zone_data, pts, points_map, history, recent_periods
    )

    weights = None
    try:
        from core.pure_pool_lr_trainer import load_weights
        weights = load_weights()
    except Exception:
        weights = None

    high_conf, high_src = select_high_confidence(scored, weights)
    scored = _apply_lr_probs(scored, weights) if weights else scored
    for s in scored:
        s['high_conf_pick'] = s['number'] in set(high_conf)
        s['high_conf_source'] = high_src

    report = format_pure_pool_report(
        target_period, scored, b_zone_data, pts,
        weights=weights, high_conf=high_conf, high_conf_source=high_src,
    )
    return report, scored


# ── 独立运行测试 ──
if __name__ == '__main__':
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    # 加载历史
    HISTORY_FILE = os.path.join(_PROJ, 'kl8_history_final.txt')
    POINTS_FILE = os.path.join(_PROJ, 'daily_points.txt')

    history = []
    with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            m = re.search(r'period:(\d+),numbers:([\d-]+)', line)
            if m:
                issue = m.group(1)
                nums = sorted([int(x) for x in m.group(2).split('-') if x.isdigit()])
                if len(nums) >= 15:
                    history.append({'issue': issue, 'numbers': nums})
    history.sort(key=lambda h: h['issue'], reverse=True)

    points_map = {}
    with open(POINTS_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            mp = re.search(r'period:(\d+)', line)
            mt = re.search(r'points:([\d\s]+)', line)
            if mp and mt:
                points_map[mp.group(1)] = set(int(x) for x in mt.group(1).split())

    target = str(int(history[0]['issue']) + 1)
    report, scored = run_pure_pool_analysis(target, history, points_map)

    print(report)
    print(f"\n目标期号: {target}")
    print(f"纯净池号码数: {len(scored)}")
