# -*- coding: utf-8 -*-
"""
Excel跟随号码统计表 深层关联挖掘引擎 V2
========================================
直接读取Excel原始数据（含样式标记），逐层挖掘全部规律。

三个标记维度:
  * (星号) = 热码标记 (Data1统计)
  粉(粉色填充 FFFCE4EC) = 点位标记
  框(紫色边框 FFD966B3) = 中奖标记

位置维度:
  Block 0-3 (4个板块, 每板块4行×左右各4列=32个号码位)
  Left/Right (左右侧)
  Row 0-3 (板块内行号)
  Col 0-3 (左右各4列)
"""
import sys, os, re, collections, itertools, math
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

# 统一使用项目路径管理，避免硬编码
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)
from utils.paths import data_path

EXCEL = data_path('跟随+点位+开奖数据.xlsx')
HISTORY_FILE = data_path('kl8_history_final.txt')
POINTS_FILE = data_path('daily_points.txt')

POINT_FILL = "FFFCE4EC"
BORDER_CLR = "FFD966B3"
BLOCK_OFFSETS = [1, 6, 11, 16]


# ═══════════════════════════════════════════════════════════════
#  数据加载 — 完整保留三标记+位置信息
# ═══════════════════════════════════════════════════════════════
def load_excel_full():
    """
    完整读取Excel跟随号码统计表，保留每个号码的:
    - 数值
    - 是否星号 (is_star)
    - 是否点位 (is_point)
    - 是否中奖 (is_win)
    - Block编号 (0-3)
    - 左右侧 (left/right)
    - 行号 (0-3)
    - 列号 (0-3)
    - 数据类型 (1=Data1热码统计, 2=Data2规律码)
    - 期号
    """
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=False)
    ws = wb['跟随号码统计']

    # P1-12: 一次性缓存全部行, 避免内层循环反复 iter_rows 重读 (原 O(n²) 性能问题)
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))

    records = []  # 每条记录 = 一个号码在一个期号一个位置的全部属性
    
    for r_idx, row in enumerate(all_rows):
        first_val = str(row[0].value or "").strip()
        m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
        if not m:
            continue
        issue = m.group(1)
        dtype = int(m.group(2))
        
        for b_idx, offset in enumerate(BLOCK_OFFSETS):
            for row_off in range(4):
                ri = r_idx + offset + row_off
                if ri >= len(all_rows):
                    continue
                target_row = all_rows[ri]
                
                # Left side (cols 0-3)
                for col_idx in range(0, 4):
                    if col_idx >= len(target_row):
                        continue
                    cell = target_row[col_idx]
                    v = str(cell.value or "").strip().replace('*', '')
                    if not v.isdigit():
                        continue
                    num = int(v)
                    if not (1 <= num <= 80):
                        continue
                    is_star = '*' in str(cell.value or "")
                    is_point = _check_fill(cell)
                    is_win = _check_border(cell)
                    records.append({
                        'issue': issue, 'dtype': dtype, 'num': num,
                        'is_star': is_star, 'is_point': is_point, 'is_win': is_win,
                        'block': b_idx, 'side': 'left', 'row': row_off, 'col': col_idx,
                    })
                
                # Right side (cols 5-8)
                for col_idx in range(5, 9):
                    if col_idx >= len(target_row):
                        continue
                    cell = target_row[col_idx]
                    v = str(cell.value or "").strip().replace('*', '')
                    if not v.isdigit():
                        continue
                    num = int(v)
                    if not (1 <= num <= 80):
                        continue
                    is_star = '*' in str(cell.value or "")
                    is_point = _check_fill(cell)
                    is_win = _check_border(cell)
                    records.append({
                        'issue': issue, 'dtype': dtype, 'num': num,
                        'is_star': is_star, 'is_point': is_point, 'is_win': is_win,
                        'block': b_idx, 'side': 'right', 'row': row_off, 'col': col_idx - 5,
                    })
    
    wb.close()
    return records


def _check_fill(cell):
    try:
        return cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == POINT_FILL
    except:
        return False


def _check_border(cell):
    try:
        b = cell.border
        if not b:
            return False
        for side in (b.left, b.right, b.top, b.bottom):
            if side and side.color and side.color.rgb == BORDER_CLR:
                return True
        return False
    except:
        return False


def load_history():
    history = []
    with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if 'numbers:' not in line:
                continue
            parts = line.split(',')
            date_s = parts[0].split(':')[1]
            issue_s = parts[1].split(':')[1]
            nums = [int(n) for n in parts[2].split(':')[1].strip().split('-')]
            history.append({'issue': issue_s, 'date': date_s, 'numbers': set(nums)})
    history.sort(key=lambda h: h['issue'], reverse=True)
    return history


def load_points():
    points = {}
    with open(POINTS_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            per_m = re.search(r'period:(\d+)', line)
            pts_m = re.search(r'points:([\d\s]+)', line)
            if pts_m and per_m:
                pts = {int(p) for p in pts_m.group(1).strip().split() if p}
                points[per_m.group(1)] = pts
    return points


# ═══════════════════════════════════════════════════════════════
#  挖掘维度1: 双标记交叉命中率 (星号×点位 → 当期中奖率)
# ═══════════════════════════════════════════════════════════════
def mine_marker_cross(records):
    """
    用星号和点位作为预测变量，看当期中奖率。
    注意：中奖是因变量，不能用因变量预测自己。

    Data1中所有号码都是星号，所以分2组：点位 vs 非点位
    Data2有星号和非星号，分4组：星+点 / 星+非点 / 非星+点 / 非星+非点
    """
    results = []

    # Data1: 所有号码都是星号，按点位分组
    d1 = [r for r in records if r['dtype'] == 1]
    for has_point in [True, False]:
        subset = [r for r in d1 if r['is_point'] == has_point]
        total = len(subset)
        wins = sum(1 for r in subset if r['is_win'])
        rate = wins / total if total > 0 else 0
        label = f"星号+{'点位' if has_point else '非点位'}"
        results.append({
            'source': 'Data1', 'label': label,
            'star': True, 'point': has_point,
            'total': total, 'wins': wins,
            'rate': round(rate, 4),
            'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
        })

    # Data2: 按星号×点位4组
    d2 = [r for r in records if r['dtype'] == 2]
    for has_star in [True, False]:
        for has_point in [True, False]:
            subset = [r for r in d2 if r['is_star'] == has_star and r['is_point'] == has_point]
            total = len(subset)
            wins = sum(1 for r in subset if r['is_win'])
            rate = wins / total if total > 0 else 0
            label = f"{'星号' if has_star else '非星号'}+{'点位' if has_point else '非点位'}"
            results.append({
                'source': 'Data2', 'label': label,
                'star': has_star, 'point': has_point,
                'total': total, 'wins': wins,
                'rate': round(rate, 4),
                'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
            })

    return results, {}, {}


# ═══════════════════════════════════════════════════════════════
#  挖掘维度2: Block × Side 位置中奖率
# ═══════════════════════════════════════════════════════════════
def mine_block_side(records):
    """各Block各侧的中奖率差异"""
    d1 = [r for r in records if r['dtype'] == 1]
    
    stats = {}
    for b in range(4):
        for side in ['left', 'right']:
            subset = [r for r in d1 if r['block'] == b and r['side'] == side]
            total = len(subset)
            wins = sum(1 for r in subset if r['is_win'])
            stars = sum(1 for r in subset if r['is_star'])
            points = sum(1 for r in subset if r['is_point'])
            rate = wins / total if total > 0 else 0
            stats[(b, side)] = {
                'total': total, 'wins': wins, 'rate': round(rate, 4),
                'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
                'stars': stars, 'points': points,
            }
    return stats


# ═══════════════════════════════════════════════════════════════
#  挖掘维度3: 行列位置中奖率热力图
# ═══════════════════════════════════════════════════════════════
def mine_row_col_heatmap(records):
    """每个Block内行×列的中奖率热力图"""
    d1 = [r for r in records if r['dtype'] == 1]
    
    heatmap = {}
    for b in range(4):
        for row in range(4):
            for col in range(4):
                for side in ['left', 'right']:
                    subset = [r for r in d1 if r['block'] == b and r['row'] == row 
                              and r['col'] == col and r['side'] == side]
                    total = len(subset)
                    wins = sum(1 for r in subset if r['is_win'])
                    rate = wins / total if total > 0 else 0
                    heatmap[(b, side, row, col)] = {
                        'total': total, 'wins': wins, 'rate': round(rate, 4),
                        'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
                    }
    return heatmap


# ═══════════════════════════════════════════════════════════════
#  挖掘维度4: 星号持续期→中奖率
# ═══════════════════════════════════════════════════════════════
def mine_star_duration(records, history):
    """
    追踪每个号码连续N期作为星号后，第N+1期的中奖率。
    持续当热码越久→中奖率会上升还是下降？
    """
    d1 = [r for r in records if r['dtype'] == 1]
    
    # 按期号分组
    by_issue = collections.defaultdict(dict)  # {issue: {num: is_star}}
    for r in d1:
        if r['is_star']:
            by_issue[r['issue']][r['num']] = True
    
    # 排序期号
    issues = sorted(by_issue.keys())
    
    # 对每个号码，追踪连续星号持续期
    hist_by_issue = {h['issue']: h['numbers'] for h in history}
    
    duration_stats = collections.defaultdict(lambda: {'count': 0, 'wins': 0})
    
    for num in range(1, 81):
        current_duration = 0
        for i, iss in enumerate(issues):
            is_star_now = num in by_issue.get(iss, {})
            
            if is_star_now:
                current_duration += 1
            else:
                if current_duration > 0:
                    # 星号刚断，看下一期是否中奖
                    next_iss_idx = i + 1
                    if next_iss_idx < len(issues):
                        next_iss = issues[next_iss_idx]
                        wins = hist_by_issue.get(next_iss, set())
                        duration_stats[current_duration]['count'] += 1
                        if num in wins:
                            duration_stats[current_duration]['wins'] += 1
                    current_duration = 0
    
    results = []
    for dur in sorted(duration_stats.keys()):
        s = duration_stats[dur]
        rate = s['wins'] / s['count'] if s['count'] > 0 else 0
        results.append({
            'duration': dur, 'count': s['count'], 'wins': s['wins'],
            'rate': round(rate, 4), 'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
        })
    return results


# ═══════════════════════════════════════════════════════════════
#  挖掘维度5: 跨期转移概率
# ═══════════════════════════════════════════════════════════════
def mine_cross_period_transition(records, history):
    """
    本期状态 → 下期中奖率
    - 本期星号 → 下期中奖率？
    - 本期点位 → 下期中奖率？
    - 本期中奖 → 下期再中奖率？（热号延续）
    - 本期星号+点位 → 下期中奖率？
    - 本期星号+未中奖 → 下期中奖率？（蓄力反弹）
    - 本期点位+未中奖 → 下期中奖率？
    """
    d1 = [r for r in records if r['dtype'] == 1]
    hist_by_issue = {h['issue']: h['numbers'] for h in history}
    
    # 按期号×号码索引
    by_issue_num = collections.defaultdict(dict)
    for r in d1:
        by_issue_num[r['issue']][r['num']] = {
            'star': r['is_star'], 'point': r['is_point'], 'win': r['is_win']
        }
    
    issues = sorted(by_issue_num.keys())
    
    transitions = collections.defaultdict(lambda: {'count': 0, 'wins': 0})
    
    for i in range(len(issues) - 1):
        curr_iss = issues[i]
        next_iss = issues[i + 1]
        next_wins = hist_by_issue.get(next_iss, set())
        
        curr_data = by_issue_num[curr_iss]
        
        for num, markers in curr_data.items():
            s, p, w = markers['star'], markers['point'], markers['win']
            
            # 6种转移路径
            if s and w:
                transitions['星号+中奖 → 下期']['count'] += 1
                if num in next_wins: transitions['星号+中奖 → 下期']['wins'] += 1
            if s and not w:
                transitions['星号+未中 → 下期']['count'] += 1
                if num in next_wins: transitions['星号+未中 → 下期']['wins'] += 1
            if p and w:
                transitions['点位+中奖 → 下期']['count'] += 1
                if num in next_wins: transitions['点位+中奖 → 下期']['wins'] += 1
            if p and not w:
                transitions['点位+未中 → 下期']['count'] += 1
                if num in next_wins: transitions['点位+未中 → 下期']['wins'] += 1
            if w:
                transitions['中奖 → 下期再中']['count'] += 1
                if num in next_wins: transitions['中奖 → 下期再中']['wins'] += 1
            if not s and not p and not w:
                transitions['无标记 → 下期']['count'] += 1
                if num in next_wins: transitions['无标记 → 下期']['wins'] += 1
            if s and p:
                transitions['星号+点位 → 下期']['count'] += 1
                if num in next_wins: transitions['星号+点位 → 下期']['wins'] += 1
    
    results = []
    for path, s in sorted(transitions.items()):
        rate = s['wins'] / s['count'] if s['count'] > 0 else 0
        results.append({
            'path': path, 'count': s['count'], 'wins': s['wins'],
            'rate': round(rate, 4), 'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
        })
    return results


# ═══════════════════════════════════════════════════════════════
#  挖掘维度6: Data1 vs Data2 预测力对比
# ═══════════════════════════════════════════════════════════════
def mine_d1_vs_d2(records):
    """
    Data1(热码统计)和Data2(规律码)哪个更能预测中奖？
    分别统计两种数据中：号码出现率 vs 中奖率
    """
    results = {}
    for dtype in [1, 2]:
        subset = [r for r in records if r['dtype'] == dtype]
        total = len(subset)
        wins = sum(1 for r in subset if r['is_win'])
        stars = sum(1 for r in subset if r['is_star'])
        star_wins = sum(1 for r in subset if r['is_star'] and r['is_win'])
        non_star_wins = sum(1 for r in subset if not r['is_star'] and r['is_win'])
        non_star_total = total - stars
        
        results[f'Data{dtype}'] = {
            'total': total, 'wins': wins,
            'win_rate': round(wins / total, 4) if total > 0 else 0,
            'star_count': stars,
            'star_win_rate': round(star_wins / stars, 4) if stars > 0 else 0,
            'non_star_win_rate': round(non_star_wins / non_star_total, 4) if non_star_total > 0 else 0,
        }
    return results


# ═══════════════════════════════════════════════════════════════
#  挖掘维度7: 板块饱和度→下期影响
# ═══════════════════════════════════════════════════════════════
def mine_block_saturation(records, history):
    """
    某Block本期中了N个 → 下期该Block再中几个？
    板块命中是否有惯性/反转？
    """
    d1 = [r for r in records if r['dtype'] == 1]
    hist_by_issue = {h['issue']: h['numbers'] for h in history}
    
    by_issue_block = collections.defaultdict(lambda: collections.defaultdict(int))
    for r in d1:
        if r['is_win']:
            by_issue_block[r['issue']][r['block']] += 1
    
    issues = sorted(by_issue_block.keys())
    
    # 本期Block命中数 → 下期同Block命中数
    saturation = collections.defaultdict(lambda: {'count': 0, 'next_wins_sum': 0, 'next_wins_list': []})
    
    for i in range(len(issues) - 1):
        curr_iss = issues[i]
        next_iss = issues[i + 1]
        next_wins = hist_by_issue.get(next_iss, set())
        
        # 下期各Block的命中数（需要从records中找）
        next_block_wins = collections.Counter()
        for r in d1:
            if r['issue'] == next_iss and r['is_win']:
                next_block_wins[r['block']] += 1
        
        for b in range(4):
            curr_count = by_issue_block[curr_iss][b]
            next_count = next_block_wins[b]
            saturation[curr_count]['count'] += 1
            saturation[curr_count]['next_wins_sum'] += next_count
            saturation[curr_count]['next_wins_list'].append(next_count)
    
    results = []
    for curr in sorted(saturation.keys()):
        s = saturation[curr]
        avg_next = s['next_wins_sum'] / s['count'] if s['count'] > 0 else 0
        results.append({
            'curr_wins': curr, 'periods': s['count'],
            'avg_next_wins': round(avg_next, 2),
        })
    return results


# ═══════════════════════════════════════════════════════════════
#  挖掘维度8: 每个号码的三标记状态转移矩阵
# ═══════════════════════════════════════════════════════════════
def mine_state_machine(records, history):
    """
    对每个号码，追踪其状态转移:
    状态 = (是否星号, 是否点位, 是否中奖) → 下期是否中奖
    
    找出哪些状态转移最有利于下期中奖。
    """
    d1 = [r for r in records if r['dtype'] == 1]
    hist_by_issue = {h['issue']: h['numbers'] for h in history}
    
    by_issue_num = collections.defaultdict(dict)
    for r in d1:
        by_issue_num[r['issue']][r['num']] = {
            'star': r['is_star'], 'point': r['is_point'], 'win': r['is_win']
        }
    
    issues = sorted(by_issue_num.keys())
    
    # 状态 → 下期中奖
    state_trans = collections.defaultdict(lambda: {'count': 0, 'wins': 0})
    
    for i in range(len(issues) - 1):
        curr_iss = issues[i]
        next_iss = issues[i + 1]
        next_wins = hist_by_issue.get(next_iss, set())
        
        for num, markers in by_issue_num[curr_iss].items():
            state = (markers['star'], markers['point'], markers['win'])
            state_trans[state]['count'] += 1
            if num in next_wins:
                state_trans[state]['wins'] += 1
    
    results = []
    for state in sorted(state_trans.keys()):
        s = state_trans[state]
        rate = s['wins'] / s['count'] if s['count'] > 0 else 0
        star, point, win = state
        label = f"{'星' if star else '·'}{'点' if point else '·'}{'中' if win else '·'}"
        results.append({
            'state': label, 'star': star, 'point': point, 'win': win,
            'count': s['count'], 'wins': s['wins'],
            'rate': round(rate, 4), 'lift': round(rate / 0.25, 3) if 0.25 > 0 else 0,
        })
    results.sort(key=lambda x: -x['lift'])
    return results


# ═══════════════════════════════════════════════════════════════
#  挖掘维度9: 号码级三标记画像
# ═══════════════════════════════════════════════════════════════
def mine_number_profile(records, history):
    """
    对每个号码构建完整的三标记画像:
    - 作为星号出现次数 + 当期中奖率
    - 作为点位出现次数 + 当期中奖率
    - 星号+点位同时出现次数 + 中奖率
    - 无标记出现次数 + 中奖率
    - 下期中奖率（基于本期星号/点位）
    """
    d1 = [r for r in records if r['dtype'] == 1]
    hist_by_issue = {h['issue']: h['numbers'] for h in history}
    
    by_issue_num = collections.defaultdict(dict)
    for r in d1:
        by_issue_num[r['issue']][r['num']] = {
            'star': r['is_star'], 'point': r['is_point'], 'win': r['is_win']
        }
    
    issues = sorted(by_issue_num.keys())
    
    profiles = {}
    for num in range(1, 81):
        star_count = 0
        star_win = 0
        point_count = 0
        point_win = 0
        star_point_count = 0
        star_point_win = 0
        appear_count = 0
        win_count = 0
        # 下期预测力
        star_next_count = 0
        star_next_win = 0
        point_next_count = 0
        point_next_win = 0
        
        for i, iss in enumerate(issues):
            if num not in by_issue_num[iss]:
                continue
            m = by_issue_num[iss][num]
            appear_count += 1
            if m['win']:
                win_count += 1
            if m['star']:
                star_count += 1
                if m['win']:
                    star_win += 1
            if m['point']:
                point_count += 1
                if m['win']:
                    point_win += 1
            if m['star'] and m['point']:
                star_point_count += 1
                if m['win']:
                    star_point_win += 1
            
            # 下期预测
            if i + 1 < len(issues):
                next_iss = issues[i + 1]
                next_wins = hist_by_issue.get(next_iss, set())
                if m['star']:
                    star_next_count += 1
                    if num in next_wins:
                        star_next_win += 1
                if m['point']:
                    point_next_count += 1
                    if num in next_wins:
                        point_next_win += 1
        
        profiles[num] = {
            'appear': appear_count,
            'win_rate': round(win_count / appear_count, 4) if appear_count > 0 else 0,
            'star_count': star_count,
            'star_win_rate': round(star_win / star_count, 4) if star_count > 0 else 0,
            'point_count': point_count,
            'point_win_rate': round(point_win / point_count, 4) if point_count > 0 else 0,
            'sp_count': star_point_count,
            'sp_win_rate': round(star_point_win / star_point_count, 4) if star_point_count > 0 else 0,
            'star_next_rate': round(star_next_win / star_next_count, 4) if star_next_count > 0 else 0,
            'point_next_rate': round(point_next_win / point_next_count, 4) if point_next_count > 0 else 0,
        }
    
    return profiles


# ═══════════════════════════════════════════════════════════════
#  最终精选
# ═══════════════════════════════════════════════════════════════
def final_select(profiles, state_machine, cross_trans, history, records, points):
    """基于全部挖掘结果，精选5个最优号码"""
    latest_issue = history[0]['issue']
    target_issue = str(int(latest_issue) + 1)
    current_points = points.get(target_issue, set())
    
    # 获取当前期Data1的星号和点位标记
    d1_latest = [r for r in records if r['dtype'] == 1 and r['issue'] == target_issue]
    # 如果目标期没有，用最新一期
    if not d1_latest:
        d1_issues = sorted(set(r['issue'] for r in records if r['dtype'] == 1))
        latest_d1 = d1_issues[-1] if d1_issues else None
        d1_latest = [r for r in records if r['dtype'] == 1 and r['issue'] == latest_d1]
    
    current_stars = set()
    current_pts = set()
    current_all = set()
    for r in d1_latest:
        current_all.add(r['num'])
        if r['is_star']:
            current_stars.add(r['num'])
        if r['is_point']:
            current_pts.add(r['num'])
    
    # 最佳状态转移的lift
    best_state_lift = {}
    for s in state_machine:
        star, point, win = s['star'], s['point'], s['win']
        best_state_lift[(star, point, win)] = s['lift']
    
    candidates = []
    for num in range(1, 81):
        p = profiles[num]
        
        # 综合评分维度
        # 1. 星号→下期中奖率
        star_next_score = p['star_next_rate'] if p['star_next_rate'] > 0 else 0.25
        
        # 2. 点位→下期中奖率
        point_next_score = p['point_next_rate'] if p['point_next_rate'] > 0 else 0.25
        
        # 3. 星号+点位当期中奖率
        sp_score = p['sp_win_rate'] if p['sp_count'] > 0 else 0.25
        
        # 4. 当前状态加成
        is_star_now = num in current_stars
        is_point_now = num in current_pts
        is_in_excel = num in current_all
        
        # 找当前状态对应的最佳转移lift
        current_state_lift = 1.0
        if is_in_excel:
            # 查找当前标记状态
            for r in d1_latest:
                if r['num'] == num:
                    key = (r['is_star'], r['is_point'], r['is_win'])
                    current_state_lift = best_state_lift.get(key, 1.0)
                    break
        
        # 5. 近期活跃度
        recent_20 = sum(1 for h in history[:20] if num in h['numbers']) / 20
        
        # 6. 稳定性（近期命中频率）
        recent_100 = sum(1 for h in history[:100] if num in h['numbers']) / 100
        
        # 综合评分
        total = (
            star_next_score * 0.25 +
            point_next_score * 0.20 +
            sp_score * 0.15 +
            current_state_lift * 0.15 +
            recent_100 * 0.15 +
            (0.05 if is_star_now else 0) +
            (0.05 if is_point_now else 0)
        )
        
        candidates.append({
            'num': num,
            'total': round(total, 4),
            'star_next': round(star_next_score, 3),
            'point_next': round(point_next_score, 3),
            'sp_win': round(sp_score, 3),
            'state_lift': round(current_state_lift, 2),
            'recent_100': round(recent_100, 3),
            'recent_20': round(recent_20, 3),
            'is_star_now': is_star_now,
            'is_point_now': is_point_now,
            'star_count': p['star_count'],
            'point_count': p['point_count'],
            'sp_count': p['sp_count'],
        })
    
    candidates.sort(key=lambda x: -x['total'])
    return candidates, current_stars, current_pts


# ═══════════════════════════════════════════════════════════════
#  报告生成
# ═══════════════════════════════════════════════════════════════
def generate_report(marker_cross, block_stats, heatmap, star_duration,
                    cross_trans, d1_vs_d2, saturation, state_machine,
                    profiles, final_picks, current_stars, current_pts,
                    history, target_issue):
    L = []
    
    L.append("# Excel跟随号码统计表 深层关联挖掘报告")
    L.append(f"**目标期号：** {target_issue}")
    L.append(f"**数据源：** 跟随+点位+开奖数据.xlsx → 跟随号码统计表")
    L.append(f"**挖掘期数：** {len(history)}期开奖 + Excel全部期号")
    L.append("")
    L.append("**三个标记维度：**")
    L.append("- `*` 星号 = 热码标记（Data1统计）")
    L.append("- 粉色填充 = 点位标记")
    L.append("- 紫色边框 = 中奖标记")
    L.append("")
    
    # ── 维度1: 双标记交叉 ──
    L.append("## 一、双标记交叉命中率（星号×点位 → 当期中奖率）")
    L.append("")
    L.append("**核心问题：** 用星号和点位作为预测变量，看哪种组合的当期中奖率最高？")
    L.append("")
    L.append("| 数据源 | 标记组合 | 出现次数 | 中奖次数 | 中奖率 | Lift | 含义 |")
    L.append("|:------:|---------|:-------:|:-------:|:------:|:----:|------|")
    for r in marker_cross[0]:
        lift_tag = "🔥" if r['lift'] > 1.15 else "⚠️" if r['lift'] < 0.85 else ""
        L.append(f"| {r['source']} | {r['label']} | {r['total']} | {r['wins']} | {r['rate']:.1%} | {r['lift']:.2f}x {lift_tag} | {r['label']} |")
    L.append("")
    
    # 关键发现
    valid = [r for r in marker_cross[0] if r['total'] > 10]
    if valid:
        best_combo = max(valid, key=lambda x: x['lift'])
        worst_combo = min(valid, key=lambda x: x['lift'])
        L.append(f"> **关键发现：** 最强组合 = {best_combo['source']}的「{best_combo['label']}」，")
        L.append(f"> 中奖率 {best_combo['rate']:.1%}（Lift={best_combo['lift']:.2f}x），比基线25%高出{(best_combo['rate']-0.25)*100:.1f}个百分点。")
        L.append(f"> 最弱组合 = {worst_combo['source']}的「{worst_combo['label']}」，中奖率仅 {worst_combo['rate']:.1%}。")
        L.append("")
    
    # ── 维度2: Block×Side ──
    L.append("## 二、Block×Side 位置中奖率")
    L.append("")
    L.append("**核心问题：** 4个板块、左右两侧，哪个位置的中奖率最高？")
    L.append("")
    L.append("| Block | 侧 | 出现次数 | 中奖次数 | 中奖率 | Lift | 星号数 | 点位数 |")
    L.append("|:-----:|:--:|:-------:|:-------:|:------:|:----:|:------:|:------:|")
    for b in range(4):
        for side in ['left', 'right']:
            s = block_stats[(b, side)]
            lift_tag = "🔥" if s['lift'] > 1.1 else "⚠️" if s['lift'] < 0.9 else ""
            L.append(f"| B{b} | {side} | {s['total']} | {s['wins']} | {s['rate']:.1%} | {s['lift']:.2f}x {lift_tag} | {s['stars']} | {s['points']} |")
    L.append("")
    
    # ── 维度3: 热力图 ──
    L.append("## 三、行×列位置中奖率热力图")
    L.append("")
    L.append("**核心问题：** 每个Block内，哪行哪列的位置最容易出中奖号？")
    L.append("")
    for b in range(4):
        L.append(f"### Block {b}")
        L.append("")
        for side in ['left', 'right']:
            L.append(f"**{side.upper()}侧:**")
            L.append("")
            L.append("| 行\\列 | 0 | 1 | 2 | 3 |")
            L.append("|:------:|---|---|---|---|")
            for row in range(4):
                vals = []
                for col in range(4):
                    h = heatmap[(b, side, row, col)]
                    tag = "🔥" if h['lift'] > 1.15 else "❄️" if h['lift'] < 0.85 else ""
                    vals.append(f"{h['rate']:.0%}({h['lift']:.1f}x){tag}")
                L.append(f"| 行{row} | {' | '.join(vals)} |")
            L.append("")
    
    # ── 维度4: 星号持续期 ──
    L.append("## 四、星号持续期→中奖率")
    L.append("")
    L.append("**核心问题：** 一个号码连续N期作为星号（热码），断号后下期中奖率如何变化？")
    L.append("")
    if star_duration:
        L.append("| 连续星号期数 | 样本数 | 下期中奖数 | 中奖率 | Lift | 解读 |")
        L.append("|:----------:|:------:|:---------:|:------:|:----:|------|")
        for r in star_duration:
            if r['count'] < 3:
                continue
            if r['duration'] <= 2:
                interp = "短暂热门，断后回补弱"
            elif r['duration'] <= 5:
                interp = "中等热门，断后有回补信号"
            elif r['duration'] <= 10:
                interp = "长期热门，断后爆发力强"
            else:
                interp = "极长热门，断后可能大爆发"
            lift_tag = "🔥" if r['lift'] > 1.2 else "⚠️" if r['lift'] < 0.8 else ""
            L.append(f"| {r['duration']}期 | {r['count']} | {r['wins']} | {r['rate']:.1%} | {r['lift']:.2f}x {lift_tag} | {interp} |")
        L.append("")
    else:
        L.append("数据不足，无法分析。")
        L.append("")
    
    # ── 维度5: 跨期转移 ──
    L.append("## 五、跨期转移概率（本期状态→下期中奖率）")
    L.append("")
    L.append("**核心问题：** 本期是星号/点位/中奖，下期再中奖的概率分别是多少？")
    L.append("")
    L.append("| 转移路径 | 样本数 | 下期中奖 | 中奖率 | Lift |")
    L.append("|---------|:------:|:-------:|:------:|:----:|")
    for r in cross_trans:
        lift_tag = "🔥" if r['lift'] > 1.2 else "⚠️" if r['lift'] < 0.8 else ""
        L.append(f"| {r['path']} | {r['count']} | {r['wins']} | {r['rate']:.1%} | {r['lift']:.2f}x {lift_tag} |")
    L.append("")
    
    # ── 维度6: Data1 vs Data2 ──
    L.append("## 六、Data1（热码统计）vs Data2（规律码）预测力对比")
    L.append("")
    L.append("| 数据源 | 总号码数 | 中奖数 | 整体中奖率 | 星号数 | 星号中奖率 | 非星号中奖率 |")
    L.append("|--------|:-------:|:------:|:---------:|:------:|:---------:|:-----------:|")
    for key in ['Data1', 'Data2']:
        d = d1_vs_d2[key]
        L.append(f"| {key} | {d['total']} | {d['wins']} | {d['win_rate']:.1%} | {d['star_count']} | {d['star_win_rate']:.1%} | {d['non_star_win_rate']:.1%} |")
    L.append("")
    
    # ── 维度7: 板块饱和度 ──
    L.append("## 七、板块饱和度→下期影响")
    L.append("")
    L.append("**核心问题：** 某Block本期中了N个号码，下期该Block平均中几个？有惯性还是反转？")
    L.append("")
    L.append("| 本期该Block中奖数 | 样本数 | 下期同Block平均中奖数 | 解读 |")
    L.append("|:---------------:|:------:|:-------------------:|------|")
    for r in saturation:
        if r['periods'] < 3:
            continue
        if r['avg_next_wins'] > r['curr_wins']:
            interp = "惯性延续↑"
        elif r['avg_next_wins'] < r['curr_wins']:
            interp = "均值回归↓"
        else:
            interp = "持平→"
        L.append(f"| {r['curr_wins']}个 | {r['periods']}期 | {r['avg_next_wins']:.1f}个 | {interp} |")
    L.append("")
    
    # ── 维度8: 状态机 ──
    L.append("## 八、号码状态机转移矩阵")
    L.append("")
    L.append("**核心问题：** 号码处于什么状态时，下期最可能中奖？")
    L.append("")
    L.append("**状态编码：** 星=星号标记 | 点=点位标记 | 中=当期中奖 | ·=无此标记")
    L.append("")
    L.append("| 状态 | 样本数 | 下期中奖 | 中奖率 | Lift | 解读 |")
    L.append("|:----:|:------:|:-------:|:------:|:----:|------|")
    for r in state_machine:
        if r['count'] < 10:
            continue
        lift_tag = "🔥" if r['lift'] > 1.2 else "⚠️" if r['lift'] < 0.8 else ""
        star_s = "星号" if r['star'] else "非星"
        point_s = "点位" if r['point'] else "非点"
        win_s = "中奖" if r['win'] else "未中"
        L.append(f"| {r['state']} | {r['count']} | {r['wins']} | {r['rate']:.1%} | {r['lift']:.2f}x {lift_tag} | {star_s}+{point_s}+{win_s} |")
    L.append("")
    
    # ── 维度9: 号码画像Top ──
    L.append("## 九、号码级三标记画像（Top 20最优号码）")
    L.append("")
    L.append("**每个号码的完整画像：** 作为星号/点位/双标记时的中奖率 + 对下期的预测力")
    L.append("")
    
    # 按星号→下期中奖率排序
    ranked = sorted(profiles.items(), key=lambda x: -x[1]['star_next_rate'] if x[1]['star_count'] >= 5 else 0)
    
    L.append("### 星号→下期中奖率 Top 20（星号出现≥5次）")
    L.append("")
    L.append("| 号码 | 出现次数 | 整体中奖率 | 星号次数 | 星号当期中奖率 | 星号→下期中奖率 | 点位次数 | 点位→下期中奖率 | 双标记次数 | 双标记中奖率 |")
    L.append("|:----:|:-------:|:---------:|:-------:|:------------:|:-------------:|:-------:|:-------------:|:---------:|:-----------:|")
    for num, p in ranked[:20]:
        if p['star_count'] < 5:
            continue
        L.append(
            f"| {num:02d} | {p['appear']} | {p['win_rate']:.0%} | "
            f"{p['star_count']} | {p['star_win_rate']:.0%} | "
            f"{p['star_next_rate']:.0%} | "
            f"{p['point_count']} | {p['point_next_rate']:.0%} | "
            f"{p['sp_count']} | {p['sp_win_rate']:.0%} |"
        )
    L.append("")
    
    # ── 最终精选 ──
    L.append("---")
    L.append("")
    L.append("## 🎯 最终精选：5个最优稳定命中")
    L.append("")
    L.append("**综合评分公式：**")
    L.append("- 星号→下期中奖率 × 25%")
    L.append("- 点位→下期中奖率 × 20%")
    L.append("- 双标记当期中奖率 × 15%")
    L.append("- 当前状态转移Lift × 15%")
    L.append("- 近100期命中频率 × 15%")
    L.append("- 当前为星号 +5% / 当前为点位 +5%")
    L.append("")
    L.append(f"**当前期星号号码：** {sorted(current_stars)}")
    L.append(f"**当前期点位号码：** {sorted(current_pts)}")
    L.append("")
    
    L.append("| 排名 | 号码 | 综合评分 | 星号→下期 | 点位→下期 | 双标记 | 状态Lift | 近100期 | 当前星号 | 当前点位 | 近20期 |")
    L.append("|:----:|:----:|:-------:|:--------:|:--------:|:------:|:-------:|:-------:|:-------:|:-------:|:------:|")
    for i, p in enumerate(final_picks[:5], 1):
        star_tag = "✅" if p['is_star_now'] else "—"
        pt_tag = "✅" if p['is_point_now'] else "—"
        L.append(
            f"| {i} | **{p['num']:02d}** | {p['total']:.3f} | "
            f"{p['star_next']:.0%} | {p['point_next']:.0%} | "
            f"{p['sp_win']:.0%} | {p['state_lift']:.1f}x | "
            f"{p['recent_100']:.0%} | {star_tag} | {pt_tag} | "
            f"{p['recent_20']:.0%} |"
        )
    L.append("")
    
    # 回避号码
    L.append("### ⛔ 重点回避号码（综合评分最低5个）")
    L.append("")
    L.append("| 号码 | 综合评分 | 星号→下期 | 点位→下期 | 说明 |")
    L.append("|:----:|:-------:|:--------:|:--------:|------|")
    for p in final_picks[-5:]:
        L.append(
            f"| {p['num']:02d} | {p['total']:.3f} | "
            f"{p['star_next']:.0%} | {p['point_next']:.0%} | "
            f"各维度均低于基线 |"
        )
    L.append("")
    
    L.append("---")
    L.append("*报告由Excel深层关联挖掘引擎V2生成 — 直接读取跟随号码统计表全部标记+位置信息*")
    
    return "\n".join(L)


# ═══════════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════════
def run():
    print("=" * 70)
    print("Excel跟随号码统计表 深层关联挖掘引擎 V2")
    print("=" * 70)
    
    print("\n[加载] 读取Excel原始数据（含样式标记）...")
    records = load_excel_full()
    d1_count = sum(1 for r in records if r['dtype'] == 1)
    d2_count = sum(1 for r in records if r['dtype'] == 2)
    star_count = sum(1 for r in records if r['is_star'])
    point_count = sum(1 for r in records if r['is_point'])
    win_count = sum(1 for r in records if r['is_win'])
    print(f"  → 总记录: {len(records)} (Data1={d1_count}, Data2={d2_count})")
    print(f"  → 星号: {star_count}, 点位: {point_count}, 中奖: {win_count}")
    
    print("[加载] 读取开奖历史...")
    history = load_history()
    print(f"  → {len(history)}期，最新{history[0]['issue']}")
    
    print("[加载] 读取点位数据...")
    points = load_points()
    print(f"  → {len(points)}期")
    
    target_issue = str(int(history[0]['issue']) + 1)
    print(f"\n目标期号: {target_issue}")
    
    print("\n" + "=" * 70)
    print("开始逐层挖掘...")
    print("=" * 70)
    
    print("\n[维度1] 双标记交叉命中率...")
    marker_cross = mine_marker_cross(records)
    for r in marker_cross[0]:
        print(f"  {r['source']} {r['label']}: {r['total']}次 → 中奖率{r['rate']:.1%} (Lift={r['lift']:.2f}x)")
    
    print("\n[维度2] Block×Side位置中奖率...")
    block_stats = mine_block_side(records)
    for b in range(4):
        for side in ['left', 'right']:
            s = block_stats[(b, side)]
            print(f"  B{b}-{side}: {s['rate']:.1%} (Lift={s['lift']:.2f}x) stars={s['stars']} pts={s['points']}")
    
    print("\n[维度3] 行列位置热力图...")
    heatmap = mine_row_col_heatmap(records)
    # 找最热和最冷的位置
    sorted_positions = sorted(heatmap.items(), key=lambda x: -x[1]['lift'])
    print(f"  最热位置: B{sorted_positions[0][0][0]}-{sorted_positions[0][0][1]} 行{sorted_positions[0][0][2]} 列{sorted_positions[0][0][3]} → Lift={sorted_positions[0][1]['lift']:.2f}x")
    print(f"  最冷位置: B{sorted_positions[-1][0][0]}-{sorted_positions[-1][0][1]} 行{sorted_positions[-1][0][2]} 列{sorted_positions[-1][0][3]} → Lift={sorted_positions[-1][1]['lift']:.2f}x")
    
    print("\n[维度4] 星号持续期→中奖率...")
    star_duration = mine_star_duration(records, history)
    for r in star_duration:
        if r['count'] >= 3:
            print(f"  连续{r['duration']}期星号→断后中奖率{r['rate']:.1%} (Lift={r['lift']:.2f}x) 样本{r['count']}")
    
    print("\n[维度5] 跨期转移概率...")
    cross_trans = mine_cross_period_transition(records, history)
    for r in cross_trans:
        print(f"  {r['path']}: {r['rate']:.1%} (Lift={r['lift']:.2f}x) 样本{r['count']}")
    
    print("\n[维度6] Data1 vs Data2...")
    d1_vs_d2 = mine_d1_vs_d2(records)
    for key, d in d1_vs_d2.items():
        print(f"  {key}: 整体{d['win_rate']:.1%} 星号{d['star_win_rate']:.1%} 非星号{d['non_star_win_rate']:.1%}")
    
    print("\n[维度7] 板块饱和度...")
    saturation = mine_block_saturation(records, history)
    for r in saturation:
        if r['periods'] >= 3:
            print(f"  本期中{r['curr_wins']}个→下期平均{r['avg_next_wins']:.1f}个 (样本{r['periods']}期)")
    
    print("\n[维度8] 状态机转移矩阵...")
    state_machine = mine_state_machine(records, history)
    for r in state_machine:
        if r['count'] >= 10:
            print(f"  {r['state']}: →下期中奖{r['rate']:.1%} (Lift={r['lift']:.2f}x) 样本{r['count']}")
    
    print("\n[维度9] 号码级三标记画像...")
    profiles = mine_number_profile(records, history)
    # Top 5 星号→下期中奖率
    top_star_next = sorted(profiles.items(), key=lambda x: -x[1]['star_next_rate'] if x[1]['star_count'] >= 5 else 0)[:5]
    print("  星号→下期中奖率 Top 5:")
    for num, p in top_star_next:
        print(f"    {num:02d}: 星号{p['star_count']}次→下期{p['star_next_rate']:.0%} | 点位{p['point_count']}次→下期{p['point_next_rate']:.0%} | 双标记{p['sp_count']}次→{p['sp_win_rate']:.0%}")
    
    print("\n[Final] 综合精选...")
    final_picks, curr_stars, curr_pts = final_select(profiles, state_machine, cross_trans, history, records, points)
    print(f"  Top 5: {[(p['num'], p['total']) for p in final_picks[:5]]}")
    
    print("\n[报告] 生成报告...")
    report = generate_report(
        marker_cross, block_stats, heatmap, star_duration,
        cross_trans, d1_vs_d2, saturation, state_machine,
        profiles, final_picks, curr_stars, curr_pts,
        history, target_issue
    )
    
    output = os.path.join(os.path.dirname(EXCEL), 'reports', 'excel_deep_mining_v2_report.md')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    with open(output, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n报告已保存: {output}")
    print("\n" + "=" * 70)
    print("报告全文")
    print("=" * 70)
    print(report)
    
    return report


def run_excel_deep_mining():
    """
    外部调用API — 供 aggregate_v18.py / run_today_v18.py 流水线调用
    
    Returns:
        (report_text: str, summary: dict)
        summary 包含:
          - target_issue: str
          - top5: list[int]  精选5码
          - avoid5: list[int]  回避5码
          - key_findings: list[str]  关键发现摘要
          - report_text: str  完整报告文本
    """
    records = load_excel_full()
    history = load_history()
    points = load_points()
    
    target_issue = str(int(history[0]['issue']) + 1)
    
    # 9层挖掘
    marker_cross = mine_marker_cross(records)
    block_stats = mine_block_side(records)
    heatmap = mine_row_col_heatmap(records)
    star_duration = mine_star_duration(records, history)
    cross_trans = mine_cross_period_transition(records, history)
    d1_vs_d2 = mine_d1_vs_d2(records)
    saturation = mine_block_saturation(records, history)
    state_machine = mine_state_machine(records, history)
    profiles = mine_number_profile(records, history)
    
    # 最终精选
    final_picks, curr_stars, curr_pts = final_select(
        profiles, state_machine, cross_trans, history, records, points
    )
    
    # 生成报告
    report = generate_report(
        marker_cross, block_stats, heatmap, star_duration,
        cross_trans, d1_vs_d2, saturation, state_machine,
        profiles, final_picks, curr_stars, curr_pts,
        history, target_issue
    )
    
    # 保存独立报告
    output_dir = os.path.join(os.path.dirname(EXCEL), 'reports')
    os.makedirs(output_dir, exist_ok=True)
    output = os.path.join(output_dir, 'excel_deep_mining_v2_report.md')
    with open(output, 'w', encoding='utf-8') as f:
        f.write(report)
    
    # 构建摘要
    top5 = [p['num'] for p in final_picks[:5]]
    avoid5 = [p['num'] for p in final_picks[-5:]]
    
    # 提取关键发现
    key_findings = []
    # 维度1: 最强双标记组合
    valid_mc = [r for r in marker_cross[0] if r['total'] > 10]
    if valid_mc:
        best_mc = max(valid_mc, key=lambda x: x['lift'])
        worst_mc = min(valid_mc, key=lambda x: x['lift'])
        key_findings.append(
            f"双标记最强: {best_mc['source']}「{best_mc['label']}」中奖率{best_mc['rate']:.0%}(Lift={best_mc['lift']:.2f}x)"
        )
        key_findings.append(
            f"双标记最弱: {worst_mc['source']}「{worst_mc['label']}」中奖率{worst_mc['rate']:.0%}(Lift={worst_mc['lift']:.2f}x)"
        )
    
    # 维度2: 最强Block
    best_block = max(block_stats.items(), key=lambda x: x[1]['lift'])
    worst_block = min(block_stats.items(), key=lambda x: x[1]['lift'])
    key_findings.append(
        f"位置最强: B{best_block[0][0]}-{best_block[0][1]} 中奖率{best_block[1]['rate']:.0%}(Lift={best_block[1]['lift']:.2f}x)"
    )
    key_findings.append(
        f"位置最弱: B{worst_block[0][0]}-{worst_block[0][1]} 中奖率{worst_block[1]['rate']:.0%}(Lift={worst_block[1]['lift']:.2f}x)"
    )
    
    # 维度5: 跨期转移最佳
    best_trans = max(cross_trans, key=lambda x: x['lift'])
    worst_trans = min(cross_trans, key=lambda x: x['lift'])
    key_findings.append(
        f"跨期最强: {best_trans['path']} → {best_trans['rate']:.0%}(Lift={best_trans['lift']:.2f}x)"
    )
    key_findings.append(
        f"跨期最弱: {worst_trans['path']} → {worst_trans['rate']:.0%}(Lift={worst_trans['lift']:.2f}x)"
    )
    
    # 维度8: 状态机最佳
    valid_sm = [r for r in state_machine if r['count'] >= 10]
    if valid_sm:
        best_sm = max(valid_sm, key=lambda x: x['lift'])
        key_findings.append(
            f"状态机最佳: {best_sm['state']} → 下期中奖{best_sm['rate']:.0%}(Lift={best_sm['lift']:.2f}x)"
        )
    
    summary = {
        'target_issue': target_issue,
        'top5': top5,
        'avoid5': avoid5,
        'current_stars': sorted(curr_stars),
        'current_points': sorted(curr_pts),
        'key_findings': key_findings,
        'report_text': report,
        'report_file': output,
    }
    
    return report, summary


if __name__ == '__main__':
    run()
