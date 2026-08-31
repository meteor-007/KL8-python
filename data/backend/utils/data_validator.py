# -*- coding: utf-8 -*-
"""
KL8 数据一致性校验引擎 v1.0
============================
核心功能：
  1. 校验所有数据源的期号、日期、开奖数据一致性
  2. 检测数据滞后、期号断裂、日期错位
  3. 自动修复可修复的问题（Excel同步）
  4. 输出结构化校验报告

数据源层级（单一真相源）：
  - kl8_history_final.txt = 开奖数据唯一真相源
  - daily_points.txt = 点位数据唯一真相源
  - 跟随+点位+开奖数据.xlsx = 展示层，必须与真相源对齐

校验项目：
  A. kl8_history_final.txt 数据新鲜度与完整性
  B. daily_points.txt 数据新鲜度与期号对齐
  C. Excel「开奖历史」Sheet vs kl8_history 期号对齐
  D. Excel「全量开奖数据」Sheet vs kl8_history 期号对齐
  E. Excel「跟随号码统计」Sheet 期号覆盖与顺序
  F. 热码统计文件 vs kl8_history 期号覆盖
"""
import os
import re
import sys
import logging
import datetime
from typing import Dict, List, Tuple, Optional

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()


HISTORY_FILE = os.path.join(_PROJ, 'kl8_history_final.txt')
POINTS_FILE = os.path.join(_PROJ, 'daily_points.txt')
EXCEL_FILE = os.path.join(_PROJ, '跟随+点位+开奖数据.xlsx')
HOT_DIR = data_path('热码统计')

logger = logging.getLogger("DataValidator")


def _parse_txt_history(filepath: str) -> List[Dict]:
    """解析txt历史文件为结构化列表"""
    history = []
    if not os.path.exists(filepath):
        return history
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if 'numbers:' not in line:
                continue
            parts = line.split(',')
            if len(parts) < 3:
                continue
            date_s = parts[0].split(':')[1] if ':' in parts[0] else ''
            issue_s = parts[1].split(':')[1] if ':' in parts[1] else ''
            nums_str = parts[2].split(':')[1].strip() if ':' in parts[2] else ''
            if not date_s or not issue_s or not nums_str:
                continue
            try:
                numbers = [int(n) for n in nums_str.split('-')]
            except ValueError:
                continue
            history.append({'date': date_s, 'issue': issue_s, 'numbers': numbers})
    return history


def _parse_txt_points(filepath: str) -> Dict[str, Dict]:
    """解析点位文件为 {issue: {date, points}} 字典"""
    points = {}
    if not os.path.exists(filepath):
        return points
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m_iss = re.search(r'period:(\d+)', line)
            m_date = re.search(r'date:(\d{4}-\d{2}-\d{2})', line)
            m_pts = re.search(r'points:([\d\s]+)', line)
            if m_iss and m_pts:
                pts = {int(p) for p in m_pts.group(1).strip().split() if p}
                points[m_iss.group(1)] = {
                    'date': m_date.group(1) if m_date else '',
                    'points': pts
                }
    return points


def _get_excel_sheet_latest_issue(sheet_name: str) -> Tuple[Optional[str], Optional[str], int]:
    """获取Excel指定Sheet的最新期号和日期

    Returns:
        (issue, date, row_count) 期号/日期/数据行数
    """
    import openpyxl
    if not os.path.exists(EXCEL_FILE):
        return None, None, 0
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return None, None, 0
            ws = wb[sheet_name]
            # 读取第2行（第1行是表头）
            row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            if not row2:
                return None, None, 0
            r = row2[0]
            issue = str(r[1]) if len(r) > 1 and r[1] else None
            date = str(r[0]) if len(r) > 0 and r[0] else None
            row_count = ws.max_row - 1 if ws.max_row > 1 else 0
            return issue, date, row_count
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"读取Excel Sheet '{sheet_name}' 失败: {e}")
        return None, None, 0


def _get_excel_跟随统计_issues() -> List[str]:
    """获取跟随号码统计Sheet中所有期号"""
    import openpyxl
    if not os.path.exists(EXCEL_FILE):
        return []
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        try:
            if '跟随号码统计' not in wb.sheetnames:
                return []
            ws = wb['跟随号码统计']
            issues = []
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                v = str(row[0] or '')
                m = re.search(r'(\d{7})期数据1', v)
                if m:
                    issues.append(m.group(1))
            return issues
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"读取跟随号码统计失败: {e}")
        return []


def _get_hot_dir_issues() -> List[str]:
    """获取热码统计目录中所有期号"""
    import glob
    issues = []
    for f in glob.glob(os.path.join(HOT_DIR, '*-热码统计.xlsx')):
        m = re.search(r'-(\d+)期', os.path.basename(f))
        if m:
            issues.append(m.group(1))
    return sorted(set(issues), key=int)


def validate_all(auto_fix: bool = False) -> Dict:
    """执行全量数据一致性校验

    Args:
        auto_fix: True时自动修复可修复的问题（Excel同步）

    Returns:
        校验报告字典
    """
    report = {
        'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'checks': {},
        'errors': [],
        'warnings': [],
        'fixes': [],
        'all_pass': True
    }

    # ═══════════════════════════════════════════
    # A. kl8_history_final.txt 数据新鲜度与完整性
    # ═══════════════════════════════════════════
    history = _parse_txt_history(HISTORY_FILE)
    check_a = {
        'exists': len(history) > 0,
        'total_periods': len(history),
        'latest_issue': history[0]['issue'] if history else '',
        'latest_date': history[0]['date'] if history else '',
    }

    if history:
        # 新鲜度：最新日期距今几天
        try:
            latest_dt = datetime.datetime.strptime(history[0]['date'], '%Y-%m-%d')
            day_gap = (datetime.datetime.now() - latest_dt).days
            check_a['day_gap'] = day_gap
            check_a['freshness_ok'] = day_gap <= 1
            if day_gap > 1:
                report['warnings'].append(
                    f"A. kl8_history最新日期={history[0]['date']}, 距今{day_gap}天，数据可能未更新"
                )
        except ValueError:
            check_a['freshness_ok'] = False
            report['errors'].append("A. 日期格式异常")

        # 号码完整性抽检（前10期+最后5期）
        num_errors = 0
        for h in history[:10] + history[-5:]:
            if len(h['numbers']) != 20:
                num_errors += 1
                report['errors'].append(f"A. 期号{h['issue']}号码数={len(h['numbers'])}≠20")
        check_a['number_completeness_ok'] = num_errors == 0

    else:
        report['errors'].append("A. kl8_history_final.txt 无数据或不存在")
        check_a['freshness_ok'] = False
        check_a['number_completeness_ok'] = False

    report['checks']['A_kl8_history'] = check_a

    # ═══════════════════════════════════════════
    # B. daily_points.txt 数据新鲜度与期号对齐
    # ═══════════════════════════════════════════
    points = _parse_txt_points(POINTS_FILE)
    check_b = {
        'exists': len(points) > 0,
        'total_periods': len(points),
    }

    if points:
        sorted_p_issues = sorted(points.keys(), key=int)
        check_b['latest_issue'] = sorted_p_issues[-1]
        check_b['latest_date'] = points[sorted_p_issues[-1]]['date']

        # 期号对齐：点位最新期应该是 history最新期 或 history最新期+1
        if history:
            hist_latest = int(history[0]['issue'])
            pts_latest = int(sorted_p_issues[-1])
            check_b['issue_aligned'] = pts_latest in (hist_latest, hist_latest + 1)
            if pts_latest < hist_latest:
                report['warnings'].append(
                    f"B. daily_points最新期={sorted_p_issues[-1]} < kl8_history最新期={history[0]['issue']}，点位数据滞后"
                )
            elif pts_latest > hist_latest + 1:
                report['warnings'].append(
                    f"B. daily_points最新期={sorted_p_issues[-1]} 远超 kl8_history最新期+1={hist_latest+1}，可能异常"
                )

        # 点位数量检查（每期应有20个点位）
        for iss in sorted_p_issues[-5:]:
            if len(points[iss]['points']) != 20:
                report['warnings'].append(
                    f"B. 期号{iss}点位数={len(points[iss]['points'])}≠20"
                )
    else:
        report['errors'].append("B. daily_points.txt 无数据或不存在")
        check_b['issue_aligned'] = False

    report['checks']['B_daily_points'] = check_b

    # ═══════════════════════════════════════════
    # C. Excel「开奖历史」Sheet vs kl8_history
    # ═══════════════════════════════════════════
    exc_issue_c, exc_date_c, exc_rows_c = _get_excel_sheet_latest_issue('开奖历史')
    check_c = {
        'sheet_exists': exc_issue_c is not None,
        'latest_issue': exc_issue_c,
        'latest_date': exc_date_c,
        'row_count': exc_rows_c,
    }

    if history and exc_issue_c:
        check_c['issue_aligned'] = exc_issue_c == history[0]['issue']
        check_c['date_aligned'] = exc_date_c == history[0]['date']
        check_c['row_count_ok'] = exc_rows_c == len(history)

        if exc_issue_c != history[0]['issue']:
            report['errors'].append(
                f"C. 开奖历史Sheet最新期号={exc_issue_c}, kl8_history最新={history[0]['issue']} ❌ 不一致！"
            )
        if exc_rows_c != len(history):
            report['warnings'].append(
                f"C. 开奖历史Sheet行数={exc_rows_c}, kl8_history={len(history)}期 ⚠️"
            )
    else:
        check_c['issue_aligned'] = False
        if not history:
            pass  # A已报错
        elif not exc_issue_c:
            report['errors'].append("C. 开奖历史Sheet不存在或为空")

    report['checks']['C_excel_开奖历史'] = check_c

    # ═══════════════════════════════════════════
    # D. Excel「全量开奖数据」Sheet vs kl8_history
    # ═══════════════════════════════════════════
    exc_issue_d, exc_date_d, exc_rows_d = _get_excel_sheet_latest_issue('全量开奖数据')
    check_d = {
        'sheet_exists': exc_issue_d is not None,
        'latest_issue': exc_issue_d,
        'latest_date': exc_date_d,
        'row_count': exc_rows_d,
    }

    if history and exc_issue_d:
        check_d['issue_aligned'] = exc_issue_d == history[0]['issue']
        check_d['date_aligned'] = exc_date_d == history[0]['date']
        check_d['row_count_ok'] = exc_rows_d == len(history)

        if exc_issue_d != history[0]['issue']:
            report['errors'].append(
                f"D. 全量开奖数据Sheet最新期号={exc_issue_d}, kl8_history最新={history[0]['issue']} ❌ 不一致！"
            )
    else:
        check_d['issue_aligned'] = False
        if history and not exc_issue_d:
            report['errors'].append("D. 全量开奖数据Sheet不存在或为空")

    report['checks']['D_excel_全量开奖数据'] = check_d

    # ═══════════════════════════════════════════
    # E. Excel「跟随号码统计」Sheet 期号覆盖
    # ═══════════════════════════════════════════
    follow_issues = _get_excel_跟随统计_issues()
    check_e = {
        'sheet_exists': len(follow_issues) > 0,
        'total_periods': len(follow_issues),
        'latest_issue': follow_issues[-1] if follow_issues else '',
        'oldest_issue': follow_issues[0] if follow_issues else '',
    }

    hot_issues = _get_hot_dir_issues()

    if history and follow_issues:
        # 要求：跟随统计应生成最新期号+1（目标预测期）
        hist_latest = int(history[0]['issue'])
        follow_latest = int(follow_issues[-1])
        check_e['latest_aligned'] = follow_latest >= hist_latest + 1
        if follow_latest < hist_latest + 1:
            # 热码已有目标期但 Excel 未写入 → 硬错误（会导致纯净池空跑）
            hot_latest = int(hot_issues[-1]) if hot_issues else 0
            if hot_latest >= hist_latest + 1:
                report['errors'].append(
                    f"E. 跟随号码统计最新期号={follow_issues[-1]} < 目标预测期={hist_latest+1}，"
                    f"但热码已有{hot_latest} — 同步断裂，禁止继续预测"
                )
            else:
                report['warnings'].append(
                    f"E. 跟随号码统计最新期号={follow_issues[-1]} < 目标预测期={hist_latest+1}，统计滞后"
                )

        # 期号顺序检查
        sorted_follow = sorted(follow_issues, key=int)
        check_e['order_ok'] = follow_issues == sorted_follow
        if follow_issues != sorted_follow:
            report['warnings'].append("E. 跟随号码统计期号顺序异常（非递增）")

    report['checks']['E_excel_跟随号码统计'] = check_e

    # ═══════════════════════════════════════════
    # F. 热码统计文件覆盖检查
    # ═══════════════════════════════════════════
    check_f = {
        'dir_exists': len(hot_issues) > 0,
        'total_files': len(hot_issues),
        'latest_issue': hot_issues[-1] if hot_issues else '',
        'oldest_issue': hot_issues[0] if hot_issues else '',
    }

    if history and hot_issues:
        hist_latest = int(history[0]['issue'])
        hot_latest = int(hot_issues[-1])
        check_f['latest_aligned'] = hot_latest >= hist_latest + 1
        if hot_latest < hist_latest + 1:
            report['warnings'].append(
                f"F. 热码统计最新期号={hot_issues[-1]} < 目标预测期={hist_latest+1}，需重新生成"
            )

    report['checks']['F_热码统计文件'] = check_f

    # ═══════════════════════════════════════════
    # 自动修复
    # ═══════════════════════════════════════════
    if auto_fix:
        # 修复C+D：Excel开奖历史和全量开奖数据不同步
        needs_excel_sync = False
        if history:
            if exc_issue_c != history[0]['issue'] or exc_rows_c != len(history):
                needs_excel_sync = True
            if exc_issue_d != history[0]['issue'] or exc_rows_d != len(history):
                needs_excel_sync = True

        if needs_excel_sync:
            logger.info("[自动修复] 检测到Excel与kl8_history不同步，执行sync_history_to_excel.py...")
            try:
                from data_acquisition.sync_history_to_excel import sync
                result = sync()
                if result:
                    report['fixes'].append("C+D. Excel开奖历史+全量开奖数据已重新同步 ✅")
                    # 修复后必须回读并清除已解决的 C/D 错误，避免假失败
                    exc_issue_c, exc_date_c, exc_rows_c = _get_excel_sheet_latest_issue('开奖历史')
                    exc_issue_d, exc_date_d, exc_rows_d = _get_excel_sheet_latest_issue('全量开奖数据')
                    check_c.update({
                        'latest_issue': exc_issue_c,
                        'latest_date': exc_date_c,
                        'row_count': exc_rows_c,
                        'issue_aligned': exc_issue_c == history[0]['issue'] if history and exc_issue_c else False,
                        'date_aligned': exc_date_c == history[0]['date'] if history and exc_date_c else False,
                        'row_count_ok': exc_rows_c == len(history) if history else False,
                    })
                    check_d.update({
                        'latest_issue': exc_issue_d,
                        'latest_date': exc_date_d,
                        'row_count': exc_rows_d,
                        'issue_aligned': exc_issue_d == history[0]['issue'] if history and exc_issue_d else False,
                        'date_aligned': exc_date_d == history[0]['date'] if history and exc_date_d else False,
                        'row_count_ok': exc_rows_d == len(history) if history else False,
                    })
                    report['checks']['C_excel_开奖历史'] = check_c
                    report['checks']['D_excel_全量开奖数据'] = check_d
                    report['errors'] = [
                        e for e in report['errors']
                        if not (e.startswith('C. 开奖历史') or e.startswith('D. 全量开奖数据'))
                    ]
                    report['warnings'] = [
                        w for w in report['warnings']
                        if not w.startswith('C. 开奖历史Sheet行数')
                    ]
                    if history and exc_issue_c and exc_issue_c != history[0]['issue']:
                        report['errors'].append(
                            f"C. 开奖历史Sheet最新期号={exc_issue_c}, kl8_history最新={history[0]['issue']} ❌ 不一致！"
                        )
                    if history and exc_issue_d and exc_issue_d != history[0]['issue']:
                        report['errors'].append(
                            f"D. 全量开奖数据Sheet最新期号={exc_issue_d}, kl8_history最新={history[0]['issue']} ❌ 不一致！"
                        )
                    if history and exc_rows_c != len(history):
                        report['warnings'].append(
                            f"C. 开奖历史Sheet行数={exc_rows_c}, kl8_history={len(history)}期 ⚠️"
                        )
                else:
                    report['fixes'].append("C+D. Excel同步失败 ❌")
            except Exception as e:
                report['fixes'].append(f"C+D. Excel同步异常: {e} ❌")

        # 修复F：热码统计缺失
        if history and hot_issues:
            hist_latest = int(history[0]['issue'])
            hot_latest = int(hot_issues[-1])
            if hot_latest < hist_latest + 1:
                logger.info(f"[自动修复] 热码统计滞后({hot_latest} < 目标预测期{hist_latest+1})，执行generate_hot_excel.py...")
                try:
                    import subprocess
                    script = os.path.join(_PROJ, 'data_acquisition', 'generate_hot_excel.py')
                    # 执行两次：一次 fill-missing 补齐，一次普通生成来生成目标期
                    subprocess.run(
                        [sys.executable, script, '--fill-missing'],
                        capture_output=True, text=True, encoding='utf-8', cwd=_PROJ, timeout=120
                    )
                    result = subprocess.run(
                        [sys.executable, script],
                        capture_output=True, text=True, encoding='utf-8', cwd=_PROJ, timeout=120
                    )
                    if result.returncode == 0:
                        report['fixes'].append("F. 热码统计已自动补生成(至目标预测期) ✅")
                        # 刷新热码期号列表，并清除已解决的 F 警告
                        hot_issues[:] = _get_hot_dir_issues()
                        hot_latest = int(hot_issues[-1]) if hot_issues else hot_latest
                        check_f['latest_issue'] = hot_issues[-1] if hot_issues else ''
                        check_f['total_files'] = len(hot_issues)
                        check_f['latest_aligned'] = hot_latest >= hist_latest + 1
                        report['checks']['F_热码统计文件'] = check_f
                        report['warnings'] = [
                            w for w in report['warnings'] if not w.startswith('F. 热码统计')
                        ]
                        if hot_latest < hist_latest + 1:
                            report['warnings'].append(
                                f"F. 热码统计最新期号={hot_issues[-1] if hot_issues else '无'} < 目标预测期={hist_latest+1}，需重新生成"
                            )
                    else:
                        report['fixes'].append(f"F. 热码统计补生成失败: {result.stderr[:100]} ❌")
                except Exception as e:
                    report['fixes'].append(f"F. 热码统计补生成异常: {e} ❌")

        # 修复E：热码已有目标期但跟随Sheet未写入 → 同进程同步
        if history:
            hist_latest = int(history[0]['issue'])
            target = hist_latest + 1
            follow_latest = int(follow_issues[-1]) if follow_issues else 0
            hot_latest = int(hot_issues[-1]) if hot_issues else 0
            if hot_latest >= target and follow_latest < target:
                logger.info(
                    f"[自动修复] 跟随Sheet滞后({follow_latest} < {target})，热码已有{hot_latest}，执行process_hot_numbers..."
                )
                try:
                    from data_acquisition.process_hot_numbers import process
                    process(target_period=str(target))
                    follow_issues[:] = _get_excel_跟随统计_issues()
                    if follow_issues and int(follow_issues[-1]) >= target:
                        report['fixes'].append(f"E. 跟随号码统计已同步至{follow_issues[-1]} ✅")
                        # 清除对应 errors / lag warnings
                        report['errors'] = [
                            e for e in report['errors']
                            if not e.startswith('E. 跟随号码统计')
                        ]
                        report['warnings'] = [
                            w for w in report['warnings']
                            if not w.startswith('E. 跟随号码统计最新期号')
                        ]
                        check_e['latest_issue'] = follow_issues[-1]
                        check_e['latest_aligned'] = True
                        check_e['total_periods'] = len(follow_issues)
                        report['checks']['E_excel_跟随号码统计'] = check_e
                    else:
                        report['fixes'].append(
                            f"E. 跟随同步后仍滞后(最新={follow_issues[-1] if follow_issues else '无'}) ❌"
                        )
                except Exception as e:
                    report['fixes'].append(f"E. 跟随号码统计同步异常: {e} ❌")

    # ═══════════════════════════════════════════
    # 汇总
    # ═══════════════════════════════════════════
    report['all_pass'] = len(report['errors']) == 0

    return report


def print_report(report: Dict) -> None:
    """格式化输出校验报告"""
    print("\n" + "=" * 70)
    print(f"  KL8 数据一致性校验报告 — {report['timestamp']}")
    print("=" * 70)

    checks = report['checks']

    # A
    a = checks.get('A_kl8_history', {})
    print(f"\n[A] kl8_history_final.txt")
    print(f"    期数: {a.get('total_periods', 0)}")
    print(f"    最新: 期号={a.get('latest_issue', 'N/A')}, 日期={a.get('latest_date', 'N/A')}")
    print(f"    新鲜度: {'✅' if a.get('freshness_ok') else '❌'} (距今{a.get('day_gap', '?')}天)")
    print(f"    号码完整性: {'✅' if a.get('number_completeness_ok') else '❌'}")

    # B
    b = checks.get('B_daily_points', {})
    print(f"\n[B] daily_points.txt")
    print(f"    期数: {b.get('total_periods', 0)}")
    print(f"    最新: 期号={b.get('latest_issue', 'N/A')}, 日期={b.get('latest_date', 'N/A')}")
    print(f"    期号对齐: {'✅' if b.get('issue_aligned') else '❌'}")

    # C
    c = checks.get('C_excel_开奖历史', {})
    print(f"\n[C] Excel「开奖历史」Sheet")
    print(f"    存在: {'✅' if c.get('sheet_exists') else '❌'}")
    print(f"    最新: 期号={c.get('latest_issue', 'N/A')}, 日期={c.get('latest_date', 'N/A')}")
    print(f"    行数: {c.get('row_count', 0)}")
    print(f"    期号对齐: {'✅' if c.get('issue_aligned') else '❌'}")
    print(f"    日期对齐: {'✅' if c.get('date_aligned') else '❌'}")

    # D
    d = checks.get('D_excel_全量开奖数据', {})
    print(f"\n[D] Excel「全量开奖数据」Sheet")
    print(f"    存在: {'✅' if d.get('sheet_exists') else '❌'}")
    print(f"    最新: 期号={d.get('latest_issue', 'N/A')}, 日期={d.get('latest_date', 'N/A')}")
    print(f"    行数: {d.get('row_count', 0)}")
    print(f"    期号对齐: {'✅' if d.get('issue_aligned') else '❌'}")

    # E
    e = checks.get('E_excel_跟随号码统计', {})
    print(f"\n[E] Excel「跟随号码统计」Sheet")
    print(f"    存在: {'✅' if e.get('sheet_exists') else '❌'}")
    print(f"    期数: {e.get('total_periods', 0)}")
    print(f"    范围: {e.get('oldest_issue', 'N/A')} → {e.get('latest_issue', 'N/A')}")
    print(f"    最新对齐: {'✅' if e.get('latest_aligned') else '❌'}")
    print(f"    顺序正确: {'✅' if e.get('order_ok') else '❌'}")

    # F
    f = checks.get('F_热码统计文件', {})
    print(f"\n[F] 热码统计文件目录")
    print(f"    文件数: {f.get('total_files', 0)}")
    print(f"    范围: {f.get('oldest_issue', 'N/A')} → {f.get('latest_issue', 'N/A')}")
    print(f"    最新对齐: {'✅' if f.get('latest_aligned') else '❌'}")

    # 错误汇总
    if report['errors']:
        print(f"\n{'='*70}")
        print(f"  ❌ 严重错误 ({len(report['errors'])}项)：")
        for err in report['errors']:
            print(f"    - {err}")

    if report['warnings']:
        print(f"\n  ⚠️ 警告 ({len(report['warnings'])}项)：")
        for w in report['warnings']:
            print(f"    - {w}")

    if report['fixes']:
        print(f"\n  🔧 自动修复 ({len(report['fixes'])}项)：")
        for fix in report['fixes']:
            print(f"    - {fix}")

    # 最终结论
    print(f"\n{'='*70}")
    if report['all_pass']:
        print("  ✅ 全部校验通过！数据一致性良好。")
    else:
        print("  ❌ 校验未通过！请修复上述错误后重试。")
    print("="*70)


if __name__ == '__main__':
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    import argparse
    logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')

    parser = argparse.ArgumentParser(description='KL8 数据一致性校验引擎')
    parser.add_argument('--auto-fix', action='store_true',
                        help='自动修复可修复的问题（Excel同步/热码统计补生成）')
    args = parser.parse_args()

    report = validate_all(auto_fix=args.auto_fix)
    print_report(report)
