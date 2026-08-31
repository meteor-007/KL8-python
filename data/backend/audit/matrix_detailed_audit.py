import openpyxl
import re
import collections

from utils.excel_lock import excel_lock

def detailed_matrix_audit():
    """矩阵详细审计"""
    from core.feature_optimizer import EXCEL_FILE, POINT_FILL, BORDER_CLR
    with excel_lock(EXCEL_FILE, timeout=60):
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        try:
            ws = wb['跟随号码统计']
            data2_rows = {}
            for r in range(1, ws.max_row + 1):
                val = str(ws.cell(row=r, column=1).value or "").strip()
                if "数据2" in val:
                    m = re.search(r'(\d+)', val)
                    if m: data2_rows[m.group(1)] = r
            print(f"数据2总期数: {len(data2_rows)}")
            stealth_counter = collections.Counter()
            offsets = [1, 6, 11, 16]
            for iss in sorted(data2_rows.keys()):
                start = data2_rows[iss]
                for b_idx, offset in enumerate(offsets):
                    for row_off in range(4):
                        ri = start + offset + row_off
                        if ri > ws.max_row: continue
                        for col in list(range(1, 5)) + list(range(6, 10)):
                            cell = ws.cell(row=ri, column=col)
                            val = str(cell.value or "").strip().replace('*', '')
                            if val.isdigit():
                                is_point = (cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == POINT_FILL)
                                b = cell.border
                                is_win = b and any([
                                    b.left and b.left.color and b.left.color.rgb == BORDER_CLR,
                                    b.right and b.right.color and b.right.color.rgb == BORDER_CLR,
                                    b.top and b.top.color and b.top.color.rgb == BORDER_CLR,
                                    b.bottom and b.bottom.color and b.bottom.color.rgb == BORDER_CLR
                                ])
                                if is_win and not is_point:
                                    stealth_counter[int(val)] += 1
        finally:
            wb.close()
    print(f"Top 15 隐码: {stealth_counter.most_common(15)}")
    return stealth_counter

if __name__ == '__main__':
    detailed_matrix_audit()
