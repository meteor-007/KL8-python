#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""日报数据管线 — Excel + 历史 + 点位加载"""
import glob
import logging
import os
import re
import subprocess
import sys

if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, _ensure_project_path, data_path
_ensure_project_path()
_PROJ = get_project_root()

from utils.excel_lock import excel_lock

logger = logging.getLogger("LotteryEngine")


def _parse_period_from_hot_name(path: str):
    m = re.search(r'-(\d+)期', os.path.basename(path))
    return m.group(1) if m else None


def _excel_has_follow_issue(output_file: str, issue: str) -> bool:
    """全表扫描「跟随号码统计」是否已含指定期号（不可只扫前 N 行）。"""
    import openpyxl
    if not os.path.exists(output_file):
        return False
    with excel_lock(output_file, timeout=60):
        wb = openpyxl.load_workbook(output_file, read_only=True)
        try:
            if '跟随号码统计' not in wb.sheetnames:
                return False
            ws = wb['跟随号码统计']
            needle = f'{issue}期数据'
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                if needle in str(row[0] or ''):
                    return True
            return False
        finally:
            wb.close()


def _latest_hot_file_by_period(hot_dir: str):
    hot_files = glob.glob(os.path.join(hot_dir, '*-热码统计.xlsx'))
    if not hot_files:
        return None, None
    ranked = []
    for f in hot_files:
        issue = _parse_period_from_hot_name(f)
        if issue and issue.isdigit():
            ranked.append((int(issue), f, issue))
    if not ranked:
        return None, None
    ranked.sort(key=lambda x: x[0])
    _, path, issue = ranked[-1]
    return path, issue


class DataCenter:
    """数据中心：加载 Excel + 历史 + 点位（日报前置数据管线）"""
    _instance = None
    _lock = __import__('threading').Lock()

    def __new__(cls):
        with cls._lock:
            if cls._instance is None:
                cls._instance = super().__new__(cls)
                cls._instance.initialized = False
        return cls._instance

    @classmethod
    def reset_singleton(cls):
        """强制重置单例（重跑日报前清空）"""
        with cls._lock:
            cls._instance = None

    def _sync_hot_numbers_to_excel(self):
        """将最新热码同步到主 Excel；失败必须抛错，禁止静默跳过。"""
        hot_dir = data_path('热码统计')
        output_file = data_path('跟随+点位+开奖数据.xlsx')
        latest_file, issue = _latest_hot_file_by_period(hot_dir)
        if not latest_file or not issue:
            logger.warning("热码统计目录无可用文件，跳过同步")
            return

        try:
            if _excel_has_follow_issue(output_file, issue):
                logger.info(f"跟随号码统计已含 {issue}，跳过热码同步")
                return
        except Exception as e:
            logger.warning(f"检查主Excel期号失败，将强制同步: {e}")

        # 关键：必须同进程 import 调用，禁止在父进程已持锁时再开子进程抢锁
        from data_acquisition.process_hot_numbers import process_single
        logger.info(f"同步热码 → 跟随号码统计: {issue} ({os.path.basename(latest_file)})")
        process_single(latest_file)

        if not _excel_has_follow_issue(output_file, issue):
            raise RuntimeError(
                f"热码同步后主Excel仍缺少 {issue}期数据！"
                f"请检查 process_hot_numbers / Excel 锁 / 文件是否被占用。"
            )
        logger.info(f"热码同步校验通过: {issue} 已写入跟随号码统计")

        # Excel 已变，必须作废 JSON 高速缓存，避免 load_all_data 读到旧期
        cache_file = os.path.join(_PROJ, 'data_cache.json')
        if os.path.exists(cache_file):
            try:
                os.remove(cache_file)
                logger.info("已删除过期 data_cache.json")
            except OSError as e:
                logger.warning(f"删除 data_cache.json 失败: {e}")

    def _ensure_follow_covers_target(self):
        """硬门禁：跟随最新期必须覆盖目标预测期(history最新+1)。"""
        from utils.data_validator import _get_excel_跟随统计_issues, _get_hot_dir_issues, _parse_txt_history
        history = _parse_txt_history(os.path.join(_PROJ, 'kl8_history_final.txt'))
        if not history:
            return
        target = str(int(history[0]['issue']) + 1)
        follow_issues = _get_excel_跟随统计_issues()
        follow_latest = follow_issues[-1] if follow_issues else ''
        hot_issues = _get_hot_dir_issues()
        hot_latest = hot_issues[-1] if hot_issues else ''

        if hot_latest and int(hot_latest) >= int(target):
            if not follow_latest or int(follow_latest) < int(target):
                raise RuntimeError(
                    f"跟随号码统计滞后: Excel最新={follow_latest or '无'} < 目标期={target} "
                    f"(热码已有 {hot_latest})。禁止带着残缺跟随数据跑预测。"
                )

    def _apply_excel_formats(self):
        script_path = os.path.join(_PROJ, 'format', 'apply_formats.py')
        if not os.path.exists(script_path):
            return
        # 同进程调用，避免子进程与锁冲突
        from format.apply_formats import apply
        apply(full=False)

    def _validate_data_consistency(self):
        try:
            from utils.data_validator import validate_all
            report = validate_all(auto_fix=True)
            if not report['all_pass']:
                for err in report.get('errors', []):
                    logger.warning(f"  ❌ {err}")
            a_check = report['checks'].get('A_kl8_history', {})
            if not a_check.get('exists', False):
                raise RuntimeError("kl8_history_final.txt 无数据，无法继续！")
            return report['all_pass']
        except ImportError:
            return True
        except RuntimeError:
            raise
        except Exception as e:
            logger.warning(f"数据一致性校验异常: {e}")
            return True

    def _run_garbage_collection(self):
        script_path = os.path.join(_PROJ, 'utils', 'garbage_collector.py')
        if os.path.exists(script_path):
            subprocess.run(
                [sys.executable, script_path],
                capture_output=True, text=True, encoding='utf-8', errors='replace', cwd=_PROJ,
            )

    def initialize(self):
        if self.initialized:
            return
        self._run_garbage_collection()
        self._validate_data_consistency()
        self._sync_hot_numbers_to_excel()
        self._ensure_follow_covers_target()
        self._apply_excel_formats()

        from core import feature_optimizer as fo
        fo.clear_data_cache()
        cache_file = os.path.join(_PROJ, 'data_cache.json')
        if os.path.exists(cache_file):
            try:
                os.remove(cache_file)
                logger.info("已清除 data_cache.json，强制从 Excel 重载")
            except OSError as e:
                logger.warning(f"删除 data_cache.json 失败: {e}")
        logger.info("正在执行全量数据加载 (Excel + Txt)...")
        self.data1, self.data2, self.d1_stars, self.history, self.points = fo.load_all_data()
        self.history.sort(key=lambda h: h['issue'], reverse=True)
        if self.history:
            self.latest_issue = self.history[0]['issue']
        else:
            self.latest_issue = "000000"
        self.latest_data2_issue = max((str(k) for k in self.data2.keys()), key=int) if self.data2 else "000000"

        # 目标期跟随数据必须已进 DataCenter
        try:
            target_issue = str(int(self.latest_issue) + 1)
        except (ValueError, TypeError):
            target_issue = None
        if target_issue and target_issue not in self.data1:
            raise RuntimeError(
                f"DataCenter 加载后缺少目标期 {target_issue} 的 data1。"
                f"当前 data2 最新={self.latest_data2_issue}。请检查热码同步。"
            )

        self.initialized = True
        logger.info(
            f"数据中心加载成功: 历史={len(self.history)}期, 开奖最新={self.latest_issue}, "
            f"跟随最新={self.latest_data2_issue}, 目标期={target_issue}"
        )
