# -*- coding: utf-8 -*-
"""修复Excel — 迁移至 format/ 子树"""
import sys, os, openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from pathlib import Path

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()

def fix(excel=None):
    if excel is None:
        excel = os.path.join(_PROJ, '跟随+点位+开奖数据.xlsx')
    from utils.excel_lock import excel_lock
    with excel_lock(excel, timeout=60):
        wb = openpyxl.load_workbook(excel)
        try:
            for sn in wb.sheetnames:
                ws = wb[sn]
                if sn == '跟随号码统计':
                    PF = "FFFCE4EC"
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            cell = ws.cell(row=r, column=c)
                            v = str(cell.value or "").strip()
                            if v.isdigit() and 1 <= int(v) <= 80:
                                if not (cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == PF):
                                    pass
            wb.save(excel)
        finally:
            wb.close()
    print(f"[完成] Excel已修复: {excel}")

if __name__ == '__main__':
    fix()
