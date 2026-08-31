# -*- coding: utf-8 -*-
"""
Excel 自动化美化引擎 (v3.1 增量双期版)
=====================================
v3.1 核心修复:
  增量模式同时处理「最新预测期」+「最近已开奖期」:
    - 最新期: 点位底色 (当日点位)
    - 已开奖上期: 中奖边框 FFD966B3 (开奖号复盘标记)
  解决 v3.0 只扫最新1期导致「开奖进 history 后上期永远补不上边框」的时序缺口。

v3.0 优化保留:
  1. 增量只扫约2个数据块(~84行), 仍远快于全量
  2. 支持 --full 强制全量格式化
  3. 点位底色(FFFCE4EC) + 中奖边框(FFD966B3)

格式化规则:
  - 点位底色 (FFFCE4EC): 基于 daily_points.txt, 标记该期点位号
  - 中奖边框 (FFD966B3): 基于 kl8_history_final.txt, 标记该期实际开奖号
  - 标准边框: 所有数字单元格应用thin边框+居中对齐
"""
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os
import re
import sys

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()

from utils.excel_lock import excel_lock
from utils.paths import data_path
EXCEL_FILE = data_path('跟随+点位+开奖数据.xlsx')
POINTS_FILE = data_path('daily_points.txt')
HISTORY_FILE = data_path('kl8_history_final.txt')

POINT_FILL = PatternFill("solid", fgColor="FFFCE4EC")
WIN_SIDE = Side(style='thick', color="FFD966B3")
WIN_BORDER = Border(left=WIN_SIDE, right=WIN_SIDE, top=WIN_SIDE, bottom=WIN_SIDE)
NORMAL_SIDE = Side(style='thin', color="D9D9D9")
NORMAL_BORDER = Border(left=NORMAL_SIDE, right=NORMAL_SIDE, top=NORMAL_SIDE, bottom=NORMAL_SIDE)
CLEAR_FILL = PatternFill(fill_type=None)


def load_points():
    """加载点位数据: {期号: set(号码)}"""
    points = {}
    if os.path.exists(POINTS_FILE):
        with open(POINTS_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                m_iss = re.search(r'period:(\d+)', line)
                m_pts = re.search(r'points:([\d\s]+)', line)
                if m_iss and m_pts:
                    points[m_iss.group(1)] = set(int(n) for n in m_pts.group(1).split())
    return points


def load_history():
    """加载开奖历史: {期号: set(号码)}"""
    history = {}
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                m_iss = re.search(r'period:(\d+)', line)
                m_nums = re.search(r'numbers:([\d-]+)', line)
                if m_iss and m_nums:
                    history[m_iss.group(1)] = set(int(n) for n in m_nums.group(1).split('-'))
    return history


def _find_latest_period_range(ws):
    """定位最新一期数据块的行范围

    扫描A列找到最后一个 'XXX期数据1' 和 'XXX期数据2' 标记,
    返回 (start_row, end_row, issue)

    数据块结构: 数据1标题行 → 4周期×(4行+1空行)=20行 → 数据2标题行 → 20行 → 空行
    总计约42行
    """
    last_d1_row = None
    last_d2_row = None
    last_issue = None
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        m1 = re.search(r'(\d+)期数据1', v)
        m2 = re.search(r'(\d+)期数据2', v)
        if m1:
            last_d1_row = r
            last_issue = m1.group(1)
        if m2:
            last_d2_row = r
    if last_d1_row is None:
        return None, None, None
    # 数据2结束后还有约20行数据, 然后是空行
    end_row = (last_d2_row + 22) if last_d2_row else (last_d1_row + 42)
    end_row = min(end_row, ws.max_row)
    return last_d1_row, end_row, last_issue


def _find_period_range(ws, issue):
    """定位指定期号数据块行范围, 找不到返回 (None, None)"""
    issue = str(issue)
    d1_row = None
    d2_row = None
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        if re.search(rf'{issue}期数据1', v):
            d1_row = r
        elif re.search(rf'{issue}期数据2', v):
            d2_row = r
    if d1_row is None:
        return None, None
    end_row = (d2_row + 22) if d2_row else (d1_row + 42)
    end_row = min(end_row, ws.max_row)
    return d1_row, end_row


def _collect_sheet_issues(ws):
    """收集跟随号码统计中出现过的期号(按出现顺序)"""
    issues = []
    seen = set()
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        m = re.search(r'(\d+)期数据1', v)
        if m and m.group(1) not in seen:
            seen.add(m.group(1))
            issues.append(m.group(1))
    return issues


def _resolve_drawn_issue(latest_issue, history_map, sheet_issues):
    """解析需要对齐中奖边框的已开奖期号

    优先: latest-1(若已在history且Excel有块)
    回退: sheet中 < latest 且已在history的最大期号
    """
    if not latest_issue:
        return None
    try:
        latest_int = int(latest_issue)
    except ValueError:
        return None

    candidate = str(latest_int - 1)
    if candidate in history_map and candidate in sheet_issues:
        return candidate

    drawn = [iss for iss in sheet_issues if iss in history_map and int(iss) < latest_int]
    if not drawn:
        return None
    return max(drawn, key=int)


def _format_block(ws, start_row, end_row, current_issue, points_map, history_map):
    """格式化指定行范围内的数据块"""
    win_marked = 0
    point_marked = 0
    for r in range(start_row, end_row + 1):
        v1 = str(ws.cell(row=r, column=1).value or "").strip()
        # 跳过期号标题行
        if re.search(r'\d+期数据', v1):
            continue
        # 处理该行的号码
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            val_str = str(cell.value or "").strip().replace('*', '')
            if val_str.isdigit():
                num = int(val_str)
                # 1. 基础边框与对齐
                cell.border = NORMAL_BORDER
                cell.alignment = Alignment(horizontal='center')
                # 2. 点位底色
                if current_issue in points_map and num in points_map[current_issue]:
                    cell.fill = POINT_FILL
                    point_marked += 1
                else:
                    cell.fill = CLEAR_FILL
                # 3. 中奖边框
                if current_issue in history_map and num in history_map[current_issue]:
                    cell.border = WIN_BORDER
                    win_marked += 1
    return point_marked, win_marked


def apply(full=False):
    """执行格式化

    Args:
        full: True=全量格式化所有期号; False=增量双期(最新预测期+最近已开奖期)
    """
    print(f"[格式] 开始处理: {os.path.basename(EXCEL_FILE)} ({'全量' if full else '增量双期'}模式)")
    points_map = load_points()
    history_map = load_history()

    with excel_lock(EXCEL_FILE, timeout=60):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if '跟随号码统计' not in wb.sheetnames:
            wb.close()
            return
        ws = wb['跟随号码统计']

        if full:
            # ── 全量模式: 扫描所有行 ──
            current_issue = None
            for r in range(1, ws.max_row + 1):
                v1 = str(ws.cell(row=r, column=1).value or "").strip()
                m = re.search(r'(\d+)期数据', v1)
                if m:
                    current_issue = m.group(1)
                    print(f"  [扫描] 正在处理 {current_issue} 期区块...")
                    continue
                if not current_issue:
                    continue
                for c in range(1, 11):
                    cell = ws.cell(row=r, column=c)
                    val_str = str(cell.value or "").strip().replace('*', '')
                    if val_str.isdigit():
                        num = int(val_str)
                        cell.border = NORMAL_BORDER
                        cell.alignment = Alignment(horizontal='center')
                        if current_issue in points_map and num in points_map[current_issue]:
                            cell.fill = POINT_FILL
                        else:
                            cell.fill = CLEAR_FILL
                        if current_issue in history_map and num in history_map[current_issue]:
                            cell.border = WIN_BORDER
        else:
            # ── 增量双期: 最新预测期(点位) + 最近已开奖期(中奖边框复盘) ──
            start_row, end_row, latest_issue = _find_latest_period_range(ws)
            if start_row is None:
                print("[格式] 未找到任何期号数据，跳过")
                wb.close()
                return

            sheet_issues = set(_collect_sheet_issues(ws))
            drawn_issue = _resolve_drawn_issue(latest_issue, history_map, sheet_issues)
            targets = []
            if drawn_issue and drawn_issue != latest_issue:
                targets.append(('已开奖复盘', drawn_issue))
            targets.append(('最新预测', latest_issue))

            # 去重保序
            seen = set()
            ordered = []
            for role, iss in targets:
                if iss not in seen:
                    seen.add(iss)
                    ordered.append((role, iss))

            for role, iss in ordered:
                if iss == latest_issue:
                    s, e = start_row, end_row
                else:
                    s, e = _find_period_range(ws, iss)
                if s is None:
                    print(f"  [跳过] {role}期 {iss}: Excel中无数据块")
                    continue
                pts, wins = _format_block(ws, s, e, iss, points_map, history_map)
                in_hist = "有开奖" if iss in history_map else "未开奖"
                print(
                    f"  [增量] {role}期 {iss} (行{s}-{e}, {in_hist}) "
                    f"→ 点位底色{pts}格, 中奖边框{wins}格"
                )

        wb.save(EXCEL_FILE)
        wb.close()
    print(f"[成功] 格式与点位同步完成 ({'全量' if full else '增量双期'}模式)")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Excel格式化引擎 v3.1')
    parser.add_argument('--full', action='store_true', help='强制全量格式化所有期号')
    args = parser.parse_args()
    apply(full=args.full)
