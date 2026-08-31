#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
快乐8 热码统计生成器 - 完美还原官方格式与逻辑版 (v3.1)
=====================================================
1. 格式：16列布局 (All, 50, 25, 10期)，包含 HITS/RANK/RATIO
2. 逻辑：多维度加权共振星标 (Focus Pool Resonance Scoring)
3. 兼容：生成包含 Sheet1/2/3 的标准 Excel
4. [v3.1] 修复: 点位数据不再膨胀HITS; 复合排名解决短窗口并列
   - HITS列 = 纯开奖命中数 (不含点位)
   - RANK列 = 复合排名 (短窗口并列时用长窗口排名破序)
   - RATIO列 = 加权比率 (开奖+点位加权计算)
"""
import pandas as pd
import os
import sys
import collections
import datetime
import re
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# 强制 stdout 编码为 utf-8 以防 Windows 命令行乱码或报错
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# ── 配置 — 自动上溯到项目根目录 ──
import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()
HISTORY_FILE = os.path.join(_PROJ, 'kl8_history_final.txt')
POINTS_FILE = os.path.join(_PROJ, 'daily_points.txt')
OUTPUT_DIR = data_path('热码统计')

NUM_TOTAL = 80
WINDOWS = [("全量", None), ("50期", 50), ("25期", 25), ("10期", 10)]

# 样式定义
THIN_SIDE = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
SUB_FILL = PatternFill("solid", fgColor="FFD9EAF7")

def load_data():
    """加载历史开奖数据 + 点位数据（分离开来，不再混合）
    
    v3.1 修复: 历史数据中的 numbers 不再包含点位数据
    - numbers: 纯开奖号码 (用于 HITS 列统计)
    - points_map: 点位数据 (用于 RATIO 加权计算)
    """
    history = []
    if not os.path.exists(HISTORY_FILE):
        print(f"[错误] 未找到历史文件: {HISTORY_FILE}")
        return history, {}
    with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if 'numbers:' not in line:
                continue
            parts = line.split(',')
            history.append({
                'date': parts[0].split(':')[1],
                'period': parts[1].split(':')[1],
                'numbers': [int(n) for n in parts[2].split(':')[1].strip().split('-')]
            })
            
    # 加载点位数据（独立存储，不混入 history.numbers）
    points_map = {}
    if os.path.exists(POINTS_FILE):
        with open(POINTS_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line: continue
                m_iss = re.search(r'period:(\d+)', line)
                m_pts = re.search(r'points:([\d\s]+)', line)
                if m_iss and m_pts:
                    points_map[m_iss.group(1)] = [int(x) for x in m_pts.group(1).split()]
    
    # v3.1: 不再将点位数据 extend 进 history.numbers
    # 原逻辑导致 HITS 列被膨胀（开奖+点位叠加），号码同时出现于开奖和点位时命中数翻倍
    # 现改为: HITS = 纯开奖命中, RATIO = 加权比率(开奖权重0.7 + 点位权重0.3)
    
    return history, points_map

def rank_with_ties(values: dict[int, int]) -> dict[int, int]:
    """带有并列排名的排序引擎（标准竞赛排名法）"""
    ranked = sorted(values.items(), key=lambda item: (-item[1], item[0]))
    result: dict[int, int] = {}
    prev_value = None
    prev_rank = 0
    for index, (number, value) in enumerate(ranked, start=1):
        if value == prev_value:
            result[number] = prev_rank
        else:
            result[number] = index
            prev_rank = index
            prev_value = value
    return result


def rank_composite(
    primary_scores: dict[int, float],
    tiebreak_scores: dict[int, float],
) -> dict[int, int]:
    """复合排名引擎 — 解决短窗口大量并列问题
    
    当多个号码的 primary_score 相同时，使用 tiebreak_scores 破序：
    - primary_score 降序排（越高越热）
    - 同分时按 tiebreak_scores 降序排（长窗口排名作为破序键）
    - 再同分按号码升序排
    
    返回标准竞赛排名（并列则跳号）
    """
    ranked = sorted(
        primary_scores.items(),
        key=lambda item: (-item[1], -tiebreak_scores.get(item[0], 0), item[0])
    )
    result: dict[int, int] = {}
    prev_value = None
    prev_rank = 0
    for index, (number, value) in enumerate(ranked, start=1):
        # 判断是否真正并列：primary_score 和 tiebreak_score 都相同
        if value == prev_value and tiebreak_scores.get(number, 0) == tiebreak_scores.get(ranked[index-2][0] if index >= 2 else number, 0):
            result[number] = prev_rank
        else:
            result[number] = index
            prev_rank = index
            prev_value = value
    return result

def build_hot_windows(
    records: list[dict[str, Any]],
    points_map: dict[str, list[int]] | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """构建四个统计窗口的数据
    
    v3.1 修复:
    - HITS: 纯开奖命中数（不含点位）
    - RANK: 复合排名（短窗口并列时用长窗口排名破序）
    - RATIO: 加权比率 = (开奖命中 + 点位权重*点位命中) / 期望 * 100
             点位权重 = 0.3（避免过度膨胀）
    """
    if points_map is None:
        points_map = {}
    
    hist = records[::-1]
    
    # ── 第一步：计算各窗口的纯开奖命中数 ──
    window_draw_counts: dict[str, dict[int, int]] = {}
    window_point_counts: dict[str, dict[int, int]] = {}
    window_sizes: dict[str, int] = {}
    
    for label, window in WINDOWS:
        subset = hist if window is None else hist[-window:]
        window_sizes[label] = len(subset)
        
        # 纯开奖命中
        draw_counts = collections.Counter(
            number for record in subset for number in record["numbers"]
        )
        window_draw_counts[label] = draw_counts
        
        # 点位命中统计
        point_counts: dict[int, int] = collections.Counter()
        for record in subset:
            iss = record['period']
            if iss in points_map:
                point_counts.update(points_map[iss])
        window_point_counts[label] = point_counts
    
    # ── 第二步：计算加权得分（用于排名和RATIO）
    # 加权得分 = 开奖命中 + 0.3 * 点位命中
    # 这样点位对排名有影响但不会导致膨胀
    POINTS_WEIGHT = 0.3
    
    window_weighted_scores: dict[str, dict[int, float]] = {}
    for label in [w[0] for w in WINDOWS]:
        scores = {}
        for number in range(1, 81):
            draw_hits = window_draw_counts[label].get(number, 0)
            point_hits = window_point_counts[label].get(number, 0)
            scores[number] = draw_hits + POINTS_WEIGHT * point_hits
        window_weighted_scores[label] = scores
    
    # ── 第三步：复合排名 ──
    # 全量窗口用自身排名（数据量足够，并列很少）
    # 短窗口用更长窗口的排名作为破序键
    all_weighted = window_weighted_scores["全量"]
    
    window_ranks: dict[str, dict[int, int]] = {}
    # 全量窗口：直接排名
    window_ranks["全量"] = rank_with_ties(all_weighted)
    
    # 50期窗口：用全量排名破序
    window_ranks["50期"] = rank_composite(
        window_weighted_scores["50期"],
        all_weighted,
    )
    
    # 25期窗口：用50期加权得分破序
    window_ranks["25期"] = rank_composite(
        window_weighted_scores["25期"],
        window_weighted_scores["50期"],
    )
    
    # 10期窗口：用25期加权得分破序
    window_ranks["10期"] = rank_composite(
        window_weighted_scores["10期"],
        window_weighted_scores["25期"],
    )
    
    # ── 第四步：组装输出 ──
    result: dict[str, list[dict[str, Any]]] = {}
    for label, window in WINDOWS:
        window_size = window_sizes[label]
        expected = max(window_size * 0.25, 1e-9)
        rows = []
        for number in range(1, 81):
            draw_hits = window_draw_counts[label].get(number, 0)
            weighted_score = window_weighted_scores[label][number]
            # RATIO 基于加权得分计算（保留点位的热度贡献）
            ratio = round(weighted_score / expected * 100, 1)
            rows.append({
                "number": number,
                "hits": draw_hits,  # v3.1: HITS = 纯开奖命中数
                "rank": window_ranks[label][number],
                "ratio": ratio,
            })
        rows.sort(key=lambda item: (item["rank"], item["number"]))
        result[label] = rows
    return result

def build_focus_hot_pool(hot_windows: dict[str, list[dict[str, Any]]]) -> list[int]:
    """多窗口交集过滤精选号码"""
    star_set = set()
    top_5_all = set([item["number"] for item in hot_windows["全量"] if item["rank"] <= 5])
    star_set.update(top_5_all)
    
    window_candidates = {}
    for label in ["10期", "25期", "50期"]:
        window_candidates[label] = set([item["number"] for item in hot_windows[label] if item["rank"] <= 12])
    
    intersection_2plus = set()
    all_nums = (window_candidates["10期"] | window_candidates["25期"] | window_candidates["50期"])
    for num in all_nums:
        count = sum(1 for label in ["10期", "25期", "50期"] if num in window_candidates[label])
        if count >= 2:
            intersection_2plus.add(num)
    
    star_set.update(intersection_2plus)
    if len(star_set) < 20:
        for label in ["10期", "25期", "50期"]:
            top_8_short = [item["number"] for item in hot_windows[label] if item["rank"] <= 8]
            star_set.update(top_8_short)
    
    return sorted(list(star_set))

def generate_hot_excel(history, target_issue, target_date, points_map=None):
    """生成 Excel 文件
    
    v3.1: 新增 points_map 参数，用于加权RATIO计算（不影响HITS）
    """
    hot_windows = build_hot_windows(history, points_map=points_map)
    star_set = set(build_focus_hot_pool(hot_windows))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    col_widths = {
        1: 8, 2: 10, 3: 8, 4: 10,
        5: 8, 6: 10, 7: 8, 8: 10,
        9: 8, 10: 10, 11: 8, 12: 10,
        13: 8, 14: 10, 15: 8, 16: 10
    }
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    window_titles = {
        "全量": f"{len(history)} Game Chart",
        "50期": "50 Game Chart",
        "25期": "25 Game Chart",
        "10期": "10 Game Chart"
    }
    title_cols = [1, 5, 9, 13]
    for col, label in zip(title_cols, ["全量", "50期", "25期", "10期"]):
        ws.cell(row=1, column=col, value=window_titles[label])
        
    headers = ["##", "HITS", "RANK", "RATIO"] * 4
    for col, head in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=head)
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    labels = ["全量", "50期", "25期", "10期"]
    for idx in range(80):
        row = 3 + idx
        for w_idx, label in enumerate(labels):
            item = hot_windows[label][idx]
            start_col = w_idx * 4 + 1
            num_text = str(item['number'])
            if item['number'] in star_set:
                num_text += "*"
                
            vals = [num_text, item["hits"], item["rank"], item["ratio"]]
            for offset, val in enumerate(vals):
                cell = ws.cell(row=row, column=start_col + offset, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER

    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"{target_date}-{target_issue}期-热码统计.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    
    print(f"\n{'=' * 60}")
    print(f"✅ 热码统计文件完美复刻成功")
    print(f"   文件: {filename}")
    print(f"   星标数: {len(star_set)} 个 (加权共振池)")
    print(f"{'=' * 60}")
    return filepath

def get_existing_hot_issues():
    """扫描热码统计目录，返回已存在的期号集合"""
    import glob
    existing = set()
    for f in glob.glob(os.path.join(OUTPUT_DIR, '*-热码统计.xlsx')):
        m = re.search(r'-(\d+)期', os.path.basename(f))
        if m:
            existing.add(m.group(1))
    return existing


def _validate_issue_date_consistency(history, target_issue, target_date):
    """校验目标期号与日期的一致性

    规则：target_issue应该是history最新期号+1，target_date应该是最新日期+1天
    允许±1天误差（开奖可能不严格按日历日）

    Returns:
        (ok: bool, warning: str)
    """
    if not history:
        return True, ""

    latest_issue = int(history[0]['period'])
    expected_issue = latest_issue + 1
    actual_issue = int(target_issue)

    if actual_issue != expected_issue:
        warning = (f"目标期号{target_issue}≠预期{expected_issue}(最新期{latest_issue}+1)。"
                   f"可能存在期号断裂或数据未更新！")
        # 仅警告，不阻止（允许手动指定期号）
        return False, warning

    # 日期对齐检查
    try:
        latest_date = datetime.datetime.strptime(history[0]['date'], '%Y-%m-%d')
        target_date_obj = datetime.datetime.strptime(target_date, '%Y%m%d')
        day_diff = (target_date_obj - latest_date).days
        if day_diff < 0 or day_diff > 3:
            warning = (f"目标日期{target_date}与最新开奖日期{history[0]['date']}相差{day_diff}天，"
                       f"可能存在日期异常！")
            return False, warning
    except ValueError:
        return False, f"日期格式异常: target_date={target_date}, latest_date={history[0]['date']}"

    return True, ""


def generate_for_period(history, points_map, target_issue, target_date):
    """为指定期号生成热码统计（v3.2: 增加期号日期一致性校验）"""
    # ── 校验期号与日期一致性 ──
    ok, warning = _validate_issue_date_consistency(history, target_issue, target_date)
    if not ok:
        print(f"[校验] ⚠️ {warning}")
        print(f"[校验] 继续生成，但请确认数据是否最新！")
    else:
        print(f"[校验] ✅ 期号{target_issue}日期{target_date}与历史数据一致")

    # 复制history避免修改原始数据
    work_history = list(history)
    
    target_date_obj = datetime.datetime.strptime(target_date, '%Y%m%d')
    
    # v3.1: 目标期的点位数据作为"预热期"加入统计窗口
    # 注意：该预热期的 numbers 为空列表（HITS不受影响），
    # 但点位数据会通过 points_map 参与 RATIO 加权计算
    if target_issue in points_map:
        work_history.insert(0, {
            'date': target_date_obj.strftime('%Y-%m-%d'),
            'period': target_issue,
            'numbers': []  # v3.1: 不再填入点位数据到 numbers
        })
        print(f"[加权] 目标期 {target_issue} 点位数据已加入RATIO加权计算 (权重0.3)")

    generate_hot_excel(work_history, target_issue, target_date, points_map=points_map)


def fill_missing_periods(history, points_map):
    """自动检测并补生成缺失期号的热码统计"""
    existing_issues = get_existing_hot_issues()
    
    # 构建history的期号→日期映射
    issue_date_map = {h['period']: h['date'] for h in history}
    
    # 只检查最近30期是否有缺失（避免扫描全量历史）
    recent_history = history[:30]
    
    # 从最近history中找出缺失的热码统计期号
    missing = []
    for h in recent_history:
        source_issue = int(h['period'])
        target_issue = str(source_issue + 1)
        
        # 计算目标日期
        try:
            source_date = datetime.datetime.strptime(h['date'], '%Y-%m-%d')
            target_date_obj = source_date + datetime.timedelta(days=1)
            target_date = target_date_obj.strftime('%Y%m%d')
        except ValueError:
            continue
        
        if target_issue not in existing_issues:
            # 截取history到该期（含该期及之前的所有数据）
            truncated = [rec for rec in history if int(rec['period']) <= source_issue]
            if truncated:
                missing.append((target_issue, target_date, truncated))
    
    if not missing:
        print("[检查] 所有期号热码统计完整，无需补生成")
        return 0
    
    print(f"[补生成] 发现 {len(missing)} 个缺失期号: {[m[0] for m in missing]}")
    for target_issue, target_date, truncated_history in missing:
        print(f"  补生成 {target_issue}期 (日期{target_date})...")
        generate_for_period(truncated_history, points_map, target_issue, target_date)
        existing_issues.add(target_issue)
    
    return len(missing)


def main():
    import argparse
    parser = argparse.ArgumentParser(description='快乐8 热码统计生成器')
    parser.add_argument('--target-period', type=str, default=None,
                        help='手动指定目标期号 (如 2026134)，默认自动推算')
    parser.add_argument('--fill-missing', action='store_true',
                        help='自动检测并补生成缺失期号')
    args = parser.parse_args()
    
    print("=" * 60)
    print("  快乐8 热码统计生成器 v3.2 (校验+缺期补偿版)")
    print("=" * 60)
    history, points_map = load_data()
    if not history: return

    # ── 数据新鲜度校验 ──
    print(f"[数据] 历史{len(history)}期, 最新={history[0]['period']}({history[0]['date']})")

    # 检查kl8_history是否需要更新
    today_str = datetime.datetime.now().strftime('%Y-%m-%d')
    latest_date = history[0]['date']
    day_gap = (datetime.datetime.now() - datetime.datetime.strptime(latest_date, '%Y-%m-%d')).days
    if day_gap > 1:
        print(f"[警告] ⚠️ kl8_history最新日期={latest_date}, 距今{day_gap}天，数据可能未更新！")
        print(f"[建议] 先执行 python data_acquisition/fetch_kl8_history.py 更新数据")
    
    # ── 补生成缺失期号 ──
    if args.fill_missing:
        filled = fill_missing_periods(history, points_map)
        if filled > 0:
            print(f"[补生成] 共补生成 {filled} 个缺失期号")
    
    # ── 正常生成当期 ──
    if args.target_period:
        # 手动指定期号模式
        target_issue = args.target_period
        # 从history中查找该期号前一期来推算日期
        prev_issue = str(int(target_issue) - 1)
        prev_record = next((h for h in history if h['period'] == prev_issue), None)
        if prev_record:
            prev_date = datetime.datetime.strptime(prev_record['date'], '%Y-%m-%d')
            target_date = (prev_date + datetime.timedelta(days=1)).strftime('%Y%m%d')
        else:
            target_date = datetime.datetime.now().strftime('%Y%m%d')
            print(f"[警告] 无法推算日期，使用当前日期: {target_date}")
        
        # 截取history到前一期
        truncated = [h for h in history if int(h['period']) < int(target_issue)]
        generate_for_period(truncated, points_map, target_issue, target_date)
    else:
        # 自动推算模式（原始逻辑）
        latest_issue = int(history[0]['period'])
        target_issue = str(latest_issue + 1)
        
        last_date_str = history[0]['date']
        last_date = datetime.datetime.strptime(last_date_str, '%Y-%m-%d')
        target_date = (last_date + datetime.timedelta(days=1)).strftime('%Y%m%d')
        
        # fill-missing 已覆盖 latest+1 时跳过，避免同一期写两遍
        if args.fill_missing and target_issue in get_existing_hot_issues():
            print(f"[跳过] {target_issue} 已由缺期补偿生成，无需重复写入")
        else:
            generate_for_period(history, points_map, target_issue, target_date)

if __name__ == '__main__':
    main()
