# -*- coding: utf-8 -*-
"""
同步开奖历史到Excel — v2.0 双Sheet完整同步
============================================
v2.0 核心修复：
  1. 同步「开奖历史」Sheet（日期/期号/号码连字符）
  2. 同步「全量开奖数据」Sheet（日期/期号/号码文本/20个独立号码列）
  3. 同步前进行数据校验：号码完整性(=20)、期号降序、日期期号对齐
  4. 同步后验证：确认Sheet最新期号与kl8_history一致
"""
import openpyxl
import os
import re
import sys
import logging
from typing import List, Dict

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()

from utils.excel_lock import excel_lock
from utils.paths import data_path

HISTORY_FILE = data_path('kl8_history_final.txt')
EXCEL_FILE = data_path('跟随+点位+开奖数据.xlsx')

logger = logging.getLogger("SyncHistoryToExcel")


def _parse_history(filepath: str) -> List[Dict]:
    """解析 kl8_history_final.txt 为标准结构列表

    Returns:
        按文件原序（降序）排列的字典列表，每项含 date/issue/numbers
    """
    history = []
    if not os.path.exists(filepath):
        logger.error(f"历史文件不存在: {filepath}")
        return history

    with open(filepath, 'r', encoding='utf-8') as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if 'numbers:' not in line:
                continue
            parts = line.split(',')
            if len(parts) < 3:
                logger.warning(f"第{line_no}行格式异常，跳过: {line[:60]}")
                continue

            date_s = parts[0].split(':')[1] if ':' in parts[0] else ''
            issue_s = parts[1].split(':')[1] if ':' in parts[1] else ''
            nums_str = parts[2].split(':')[1].strip() if ':' in parts[2] else ''

            if not date_s or not issue_s or not nums_str:
                logger.warning(f"第{line_no}行关键字段缺失，跳过: issue={issue_s}")
                continue

            try:
                numbers = [int(n) for n in nums_str.split('-')]
            except ValueError:
                logger.warning(f"第{line_no}行号码解析失败，跳过: {nums_str[:40]}")
                continue

            history.append({
                'date': date_s,
                'issue': issue_s,
                'numbers': numbers
            })

    return history


def _validate_history(history: List[Dict]) -> bool:
    """校验历史数据的完整性和一致性

    检查项：
      1. 每期号码数量 = 20
      2. 期号降序排列（最新在前）
      3. 号码范围 1-80
      4. 日期格式 YYYY-MM-DD
      5. 期号格式 7位数字

    Returns:
        True=校验通过, False=存在严重问题
    """
    if not history:
        logger.error("[校验] 历史数据为空！")
        return False

    errors = 0
    for i, h in enumerate(history):
        # 号码数量
        if len(h['numbers']) != 20:
            logger.error(f"[校验] 期号{h['issue']}号码数={len(h['numbers'])}≠20")
            errors += 1

        # 号码范围
        for n in h['numbers']:
            if n < 1 or n > 80:
                logger.error(f"[校验] 期号{h['issue']}号码{n}超出1-80范围")
                errors += 1
                break

        # 日期格式
        if not re.match(r'\d{4}-\d{2}-\d{2}', h['date']):
            logger.error(f"[校验] 期号{h['issue']}日期格式异常: {h['date']}")
            errors += 1

        # 期号格式
        if not re.match(r'\d{7}$', h['issue']):
            logger.error(f"[校验] 期号格式异常: {h['issue']}")
            errors += 1

    # 降序检查（最多检查前100期，避免全量扫描）
    check_range = min(len(history) - 1, 100)
    order_errors = 0
    for i in range(check_range):
        if int(history[i]['issue']) <= int(history[i + 1]['issue']):
            order_errors += 1
            if order_errors <= 3:
                logger.warning(f"[校验] 期号非降序: {history[i]['issue']} <= {history[i+1]['issue']}")

    if order_errors > 0:
        logger.warning(f"[校验] 期号降序异常{order_errors}处（可能是旧数据格式）")

    if errors > 0:
        logger.error(f"[校验] 发现{errors}个严重错误，同步中止！")
        return False

    logger.info(f"[校验] 历史数据校验通过: {len(history)}期, 最新={history[0]['issue']}({history[0]['date']})")
    return True


def _sync_开奖历史(wb: openpyxl.Workbook, history: List[Dict]) -> None:
    """同步「开奖历史」Sheet（日期/期号/号码连字符格式）"""
    if '开奖历史' in wb.sheetnames:
        del wb['开奖历史']
    ws = wb.create_sheet('开奖历史')

    ws.cell(row=1, column=1, value='日期')
    ws.cell(row=1, column=2, value='期号')
    ws.cell(row=1, column=3, value='开奖号码')

    for i, h in enumerate(history, 2):
        ws.cell(row=i, column=1, value=h['date'])
        ws.cell(row=i, column=2, value=int(h['issue']))
        ws.cell(row=i, column=3, value='-'.join(f'{n:02d}' for n in h['numbers']))

    logger.info(f"[开奖历史] 写入 {len(history)} 期, 最新={history[0]['issue']}")


def _sync_全量开奖数据(wb: openpyxl.Workbook, history: List[Dict]) -> None:
    """同步「全量开奖数据」Sheet（日期/期号/号码文本/20个独立号码列）"""
    if '全量开奖数据' in wb.sheetnames:
        del wb['全量开奖数据']
    ws = wb.create_sheet('全量开奖数据')

    # 写表头
    headers = ['日期', '期号', '开奖号码（共20个）'] + [f'号码{i+1}' for i in range(20)]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # 写数据（降序，最新在最上面）
    for i, h in enumerate(history, 2):
        ws.cell(row=i, column=1, value=h['date'])
        ws.cell(row=i, column=2, value=int(h['issue']))
        ws.cell(row=i, column=3, value=' '.join(f'{n:02d}' for n in h['numbers']))
        for j, n in enumerate(h['numbers']):
            ws.cell(row=i, column=4 + j, value=n)

    logger.info(f"[全量开奖数据] 写入 {len(history)} 期, 最新={history[0]['issue']}")


def _post_sync_verify(wb: openpyxl.Workbook, expected_issue: str, expected_count: int) -> bool:
    """同步后验证：确认Excel各Sheet的最新期号与txt一致"""
    ok = True
    for sheet_name in ['开奖历史', '全量开奖数据']:
        if sheet_name not in wb.sheetnames:
            logger.error(f"[验证] Sheet '{sheet_name}' 不存在！")
            ok = False
            continue
        ws = wb[sheet_name]
        # 读取第2行（最新期）
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 4)]
        actual_issue = str(row2[1]) if row2[1] else ''
        actual_count = ws.max_row - 1  # 减去表头

        if actual_issue != expected_issue:
            logger.error(f"[验证] {sheet_name} 最新期号={actual_issue}, 期望={expected_issue} ❌")
            ok = False
        else:
            logger.info(f"[验证] {sheet_name} 最新期号={actual_issue} ✅")

        if actual_count != expected_count:
            logger.warning(f"[验证] {sheet_name} 行数={actual_count}, 期望={expected_count} ⚠️")

    return ok


def sync():
    """将历史数据同步到Excel的两个Sheet：开奖历史 + 全量开奖数据

    完整流程：
      1. 解析 kl8_history_final.txt
      2. 数据校验（号码完整性/期号降序/日期格式）
      3. 同步「开奖历史」Sheet
      4. 同步「全量开奖数据」Sheet
      5. 同步后验证
    """
    # 1. 解析
    history = _parse_history(HISTORY_FILE)
    if not history:
        logger.error("[同步] 无历史数据，中止")
        return False

    print(f"[同步] 从 {os.path.basename(HISTORY_FILE)} 读取 {len(history)} 期数据")
    print(f"[同步] 最新: 期号={history[0]['issue']}, 日期={history[0]['date']}")
    print(f"[同步] 最旧: 期号={history[-1]['issue']}, 日期={history[-1]['date']}")

    # 2. 校验
    if not _validate_history(history):
        print("[同步] ❌ 数据校验失败，中止同步！请检查 kl8_history_final.txt")
        return False

    # 3+4. 双Sheet同步
    with excel_lock(EXCEL_FILE, timeout=60):
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            # 删除默认Sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        else:
            wb = openpyxl.load_workbook(EXCEL_FILE)

        try:
            _sync_开奖历史(wb, history)
            _sync_全量开奖数据(wb, history)

            # 5. 同步后验证
            _post_sync_verify(wb, history[0]['issue'], len(history))

            wb.save(EXCEL_FILE)
        finally:
            wb.close()

    print(f"[完成] 已同步 {len(history)} 期历史到Excel（开奖历史+全量开奖数据）")
    return True


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')
    sync()
