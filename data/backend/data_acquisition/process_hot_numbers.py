# -*- coding: utf-8 -*-
"""
快乐8 热码同步引擎 (Stride 4 专业版)
=================================
1. 纵向读取 (Stride 4): 80个号码按 0,4,8... 分解为 4 行，每行 20 个数。
2. 四维度独立: All, 50, 25, 10 每个窗口独立生成 4x20 块。
3. 数据 1 (热码精准提取): 提取每行 20 个数中的带 * 号码，前 4 后 4 布局。
4. 数据 2 (位置精准提取): 提取每行 20 个数的前 4 个和最后 4 个。
5. 排版: 4行一周期（All, 50, 25, 10），周期后留空行。
"""
import pandas as pd
import openpyxl
import os
import re
import sys

# 强制 stdout 编码为 utf-8 以防 Windows 命令行乱码或报错
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()

from utils.excel_lock import excel_lock
from utils.paths import data_path

HOT_DIR = data_path('热码统计')
OUTPUT_FILE = data_path('跟随+点位+开奖数据.xlsx')
TARGET_SHEET = '跟随号码统计'

def get_hot_files_by_period(target_period=None):
    """获取热码统计文件，支持指定期号或获取所有未同步文件"""
    import glob
    files = glob.glob(os.path.join(HOT_DIR, '*-热码统计.xlsx'))
    if not files: return []
    
    if target_period:
        # 指定期号模式
        matched = [f for f in files if f'-{target_period}期-' in os.path.basename(f)]
        return matched
    else:
        # 返回按期号排序的所有文件（而非修改时间，确保顺序稳定）
        result = []
        for f in files:
            m = re.search(r'-(\d+)期', os.path.basename(f))
            if m:
                result.append((int(m.group(1)), f))
        result.sort(key=lambda x: x[0])
        return [f for _, f in result]

def get_latest_hot_file():
    """按期号取最新热码文件（不用 mtime，避免批量触摸导致选错期）。"""
    import glob
    files = glob.glob(os.path.join(HOT_DIR, '*-热码统计.xlsx'))
    if not files:
        return None
    ranked = []
    for f in files:
        m = re.search(r'-(\d+)期', os.path.basename(f))
        if m:
            ranked.append((int(m.group(1)), f))
    if not ranked:
        return max(files, key=os.path.getmtime)
    ranked.sort(key=lambda x: x[0])
    return ranked[-1][1]

def extract_raw_columns(df):
    """提取 All, 50, 25, 10 四列原始数据 (80个/列)"""
    # 假设列索引: 0, 4, 8, 12
    cols = []
    max_row = min(df.shape[0], 82)  # 防止越界：有些老文件行数不足82行
    for c_idx in [0, 4, 8, 12]:
        if c_idx >= df.shape[1]:
            cols.append([""] * 80)
            continue
        col_data = []
        for i in range(2, max_row):  # 跳过标题和表头 (第1行标题, 第2行 ## HITS...)
            val = str(df.iloc[i, c_idx]).strip()
            if val == 'nan': val = ""
            col_data.append(val)
        # 补齐到80行
        while len(col_data) < 80:
            col_data.append("")
        cols.append(col_data[:80])
    return cols

def _find_insert_row(ws, issue):
    """智能定位新数据应插入的行号
    
    策略：
      1. 如果issue已存在，清除旧数据并返回其起始行
      2. 否则，找到最后一个期号数据块的结束行，+3作为新起始行
      3. 确保按期号顺序插入（如果新期号小于末尾期号，找到正确位置插入）
    """
    # 收集所有期号标记及其行号
    markers = []  # [(row, issue_int, data_type)]
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "")
        m = re.search(r'(\d+)期数据(1|2)', v)
        if m:
            markers.append((r, int(m.group(1)), int(m.group(2))))
    
    issue_int = int(issue)
    
    # 检查是否已存在
    existing_data1_row = None
    existing_data2_row = None
    for r, iss, dtype in markers:
        if iss == issue_int and dtype == 1:
            existing_data1_row = r
        if iss == issue_int and dtype == 2:
            existing_data2_row = r
    
    if existing_data1_row is not None:
        # 已存在：清除旧数据
        # 数据块范围: data1行到data2行+22行(数据2内容)
        clear_start = existing_data1_row
        clear_end = (existing_data2_row or existing_data1_row) + 22
        for dr in range(clear_start, clear_end + 1):
            for dc in range(1, 11):
                ws.cell(row=dr, column=dc, value=None)
        return clear_start
    
    # 不存在：找到正确插入位置（按期号顺序）
    data1_markers = [(r, iss) for r, iss, dtype in markers if dtype == 1]
    data1_markers.sort(key=lambda x: x[1])  # 按期号排序
    
    if not data1_markers:
        # 空Sheet，从第1行开始
        return 1
    
    # 找到应插入在哪个已有期号之后
    insert_after = None
    for r, iss in data1_markers:
        if iss < issue_int:
            insert_after = (r, iss)
        else:
            break
    
    if insert_after is None:
        # 新期号比所有已存在的都小，插入到最前面（第1行）
        # 但这种情况通常不应发生
        return 1
    
    # 找到该期号的data2标记行
    after_row, after_iss = insert_after
    after_data2_row = None
    for r, iss, dtype in markers:
        if iss == after_iss and dtype == 2:
            after_data2_row = r
            break
    
    if after_data2_row:
        # 数据2占: 标题1行 + 4周期*(4行+1空行) = 1+20=21行
        return after_data2_row + 21 + 2  # +2是期号间间隔
    else:
        # 找不到data2标记，回退到最后一行有数据的行+3
        last_row = 0
        for r in range(ws.max_row, 0, -1):
            if ws.cell(row=r, column=1).value:
                last_row = r
                break
        return last_row + 3


def process_single(hot_file):
    """处理单个热码统计文件并同步到Excel"""
    print(f"[处理] Stride-4 同步: {os.path.basename(hot_file)}")
    df = pd.read_excel(hot_file, header=None)
    
    if df.shape[0] < 82 or df.shape[1] < 13:
        msg = (
            f"[安全降级] {os.path.basename(hot_file)} 维度异常 "
            f"(行数{df.shape[0]} < 82 或 列数{df.shape[1]} < 13)。"
            f"拒绝写入以防脏数据污染！"
        )
        print(msg)
        raise RuntimeError(msg)
    
    match = re.search(r'-(\d+)期', os.path.basename(hot_file))
    issue = match.group(1) if match else "0000"

    raw_cols = extract_raw_columns(df) # [[80], [80], [80], [80]]

    # 构建 4 个 4x20 的矩阵
    matrices = [] # [[[20]*4], [[20]*4], ...]
    for col in raw_cols:
        matrix = []
        for start in range(4):
            row = [col[i] for i in range(start, 80, 4)]
            matrix.append(row)
        matrices.append(matrix)

    with excel_lock(OUTPUT_FILE, timeout=60):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        try:
            ws = wb[TARGET_SHEET]

            # 智能定位插入行
            start_row = _find_insert_row(ws, issue)

            # --- 写入数据 1 (4 Cycle x 4 Window) ---
            ws.cell(row=start_row, column=1, value=f"{issue}期数据1")
            curr = start_row + 1
            for cycle in range(4):
                for win_idx in range(4):
                    row_20 = matrices[win_idx][cycle]
                    stars = [n for n in row_20 if '*' in n]
                    for i in range(min(4, len(stars))):
                        ws.cell(row=curr, column=i+1, value=stars[i])
                    for i in range(min(4, len(stars) - 4)):
                        ws.cell(row=curr, column=i+6, value=stars[i+4])
                    curr += 1
                curr += 1

            # --- 写入数据 2 (4 Cycle x 4 Window) ---
            data2_start = curr + 1
            ws.cell(row=data2_start, column=1, value=f"{issue}期数据2")
            curr = data2_start + 1
            for cycle in range(4):
                for win_idx in range(4):
                    row_20 = matrices[win_idx][cycle]
                    for i in range(4):
                        ws.cell(row=curr, column=i+1, value=row_20[i])
                    for i in range(4):
                        ws.cell(row=curr, column=i+6, value=row_20[16+i])
                    curr += 1
                curr += 1

            wb.save(OUTPUT_FILE)
        finally:
            wb.close()
    print(f"[成功] {issue} 期 Stride-4 逻辑同步完成")


def process(target_period=None, sync_all_missing=False):
    """同步热码数据到主Excel
    
    Args:
        target_period: 指定只同步某期号 (如 '2026134')
        sync_all_missing: True时同步所有期号（按时间排序）
    """
    if target_period:
        files = get_hot_files_by_period(target_period)
        if not files:
            print(f"[警告] 未找到期号 {target_period} 的热码统计文件")
            return
        for f in files:
            process_single(f)
    elif sync_all_missing:
        files = get_hot_files_by_period()
        if not files:
            print("[警告] 未找到任何热码统计文件")
            return
        # 检查主Excel中已有哪些期号
        existing_issues = set()
        if os.path.exists(OUTPUT_FILE):
            try:
                with excel_lock(OUTPUT_FILE, timeout=30):
                    wb = openpyxl.load_workbook(OUTPUT_FILE, read_only=True)
                    try:
                        if TARGET_SHEET in wb.sheetnames:
                            ws = wb[TARGET_SHEET]
                            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                                val = str(row[0] or "")
                                m = re.search(r'(\d+)期数据1', val)
                                if m:
                                    existing_issues.add(m.group(1))
                    finally:
                        wb.close()
            except Exception as e:
                print(f"[警告] 读取主Excel失败: {e}")
        
        # 只同步未存在于主Excel的期号
        for f in files:
            m = re.search(r'-(\d+)期', os.path.basename(f))
            if m and m.group(1) not in existing_issues:
                process_single(f)
            elif m and m.group(1) in existing_issues:
                print(f"[跳过] {m.group(1)}期已存在于主Excel")
    else:
        # 原始逻辑：只同步最新
        latest = get_latest_hot_file()
        if latest:
            process_single(latest)

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='快乐8 热码同步引擎')
    parser.add_argument('positional_args', nargs='*', help='兼容位置参数 [issue, file_path]')
    parser.add_argument('--target-period', type=str, default=None,
                        help='指定只同步某期号 (如 2026134)')
    parser.add_argument('--sync-all-missing', action='store_true',
                        help='同步所有尚未同步到主Excel的期号')
    args = parser.parse_args()
    
    # 优先兼容位置参数 [issue, file_path]
    if args.positional_args and len(args.positional_args) >= 2:
        file_path = args.positional_args[1]
        if os.path.exists(file_path):
            process_single(file_path)
        else:
            print(f"[错误] 找不到热码统计文件: {file_path}")
    else:
        process(target_period=args.target_period, sync_all_missing=args.sync_all_missing)
