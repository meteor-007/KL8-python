# -*- coding: utf-8 -*-
"""跟随号码统计 — v6: 秩带效应 + 跟随对(X→Y)分析
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if 2026089 <= i <= 2026209)

# ═══ 1. 秩带效应: block+side+col → 绝对秩 ═══
all_rows = list(ws.iter_rows())
print('═══ 1. 数据1 星号位置与命中率 (行内星号序号) ═══')
slot = defaultdict(lambda: [0,0])
for iss in issues:
    # 找到该期数据1块
    for idx in range(len(all_rows)):
        row = all_rows[idx]
        v = str(row[0].value or "").strip()
        m = re.search(r'(\d{7})期[\s\S]*?数据1', v)
        if not m or int(m.group(1)) != iss: continue
        # 读16行
        seq = []
        for off in [1, 6, 11, 16]:
            for ro in range(4):
                if idx+off+ro >= len(all_rows): continue
                trow = all_rows[idx+off+ro]
                row_cells = []
                for c in range(min(10, len(trow))):
                    cell = trow[c]
                    cv = str(cell.value or "").strip()
                    if cv and cv != 'nan':
                        row_cells.append((c, int(cv.replace('*','')), '*' in cv))
                seq.append(row_cells)
        # 每行的星号按col排序 → 序号
        for row_cells in seq:
            stars = [x for x in row_cells if x[2]]
            stars_cols = [x[0] for x in stars]
            # 行内星号序号 = 该星在 stars 中的索引
            for si, (c, n, st) in enumerate(stars):
                hit = 1 if n in draws[iss] else 0
                slot[si][1] += 1; slot[si][0] += hit
        break
print(f"{'星号序号':>8s} {'n':>6s} {'命中率':>8s}")
for si in range(8):
    h, n = slot[si]
    if n: print(f'  {si+1}号星: n={n:5d} 率={h/n:.4f}')

# column 直读
col_hit = defaultdict(lambda: [0,0])
for iss in issues:
    for idx, row in enumerate(all_rows):
        v = str(row[0].value or "").strip()
        m = re.search(r'(\d{7})期[\s\S]*?数据1', v)
        if not m or int(m.group(1)) != iss: continue
        for off in [1, 6, 11, 16]:
            for ro in range(4):
                trow = all_rows[idx+off+ro]
                for c in range(min(10, len(trow))):
                    cell = trow[c]
                    cv = str(cell.value or "").strip()
                    if cv and cv != 'nan' and cv != '*' and '*' in cv:
                        n = int(cv.replace('*',''))
                        col_hit[c][1] += 1; col_hit[c][0] += 1 if n in draws[iss] else 0
        break
print('\n数据1 列直读命中率:')
for c in range(10):
    h, n = col_hit[c]
    if n: print(f'  col{c}: n={n:5d} 率={h/n:.4f}')

# ═══ 2. 跟随对分析: X开出 → Y 跟出概率 ═══
print('\n═══ 2. 跟随对 (X上期开出 → Y本期开出) ═══')
paired = defaultdict(lambda: [0,0])  # (X,Y) -> [count X drew, Y drew after]
for i_iss in range(len(issues)-1):
    iss = issues[i_iss]; nxt = issues[i_iss+1]
    prev = draws[iss]; cur = draws[nxt]
    for x in prev:
        for y in cur:
            paired[(x,y)][1] += 1
            paired[(x,y)][0] += 1  # 计数
# 实际逻辑: P(Y | X) = #(X then Y) / #X
x_count = defaultdict(int)
pair_count = defaultdict(int)
for i_iss in range(len(issues)-1):
    iss = issues[i_iss]; nxt = issues[i_iss+1]
    prev = draws[iss]; cur = draws[nxt]
    for x in prev:
        x_count[x] += 1
        for y in cur:
            pair_count[(x,y)] += 1

# 计算跟随率 = pair / x_count; 基线 = y频率
y_freq = defaultdict(int)
for iss in issues:
    for y in draws[iss]: y_freq[y] += 1
total_draws = sum(y_freq.values())
base_p = {y: c/total_draws for y, c in y_freq.items()}

print(f'对X: 计算 P(Y本期 | X上期) / P(Y) 的 lift')
follow_lift = []
for (x, y), c in pair_count.items():
    if x_count[x] == 0: continue
    p_yx = c / x_count[x]
    p_y = base_p[y]
    if p_y == 0: continue
    lift = p_yx / p_y
    follow_lift.append((lift, x, y, c, x_count[x], p_yx, p_y))
follow_lift.sort(reverse=True)
print('Top 20 跟随对 (lift最高, 需成对次数≥8):')
shown = 0
for lift, x, y, c, xc, p_yx, p_y in follow_lift:
    if c < 8: continue
    print(f'  {x:02d}→{y:02d}: lift={lift:.2f} 次数={c} X出现={xc} P(Y|X)={p_yx:.3f} P(Y)={p_y:.3f}')
    shown += 1
    if shown >= 20: break

# 反向: 最弱跟随对 (杀号)
print('\nBottom 20 跟随对 (lift最低, 应回避):')
shown = 0
for lift, x, y, c, xc, p_yx, p_y in reversed(follow_lift):
    if c < 8: continue
    print(f'  {x:02d}→{y:02d}: lift={lift:.2f} 次数={c} X出现={xc} P(Y|X)={p_yx:.3f} P(Y)={p_y:.3f}')
    shown += 1
    if shown >= 20: break

# ═══ 3. 验证跟随对的无前视价值 ═══
print('\n═══ 3. 跟随对 无前视验证 ═══')
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]
# 用 ≤N-1 期的对计数, 预测 N 期: 对每个Y, 评分 = sum over X in prev draw of lift(X,Y)
def train_pairs(ref_issues):
    xc = defaultdict(int); pc = defaultdict(int)
    for i in range(len(ref_issues)-1):
        a = ref_issues[i]; b = ref_issues[i+1]
        for x in draws[a]:
            xc[x] += 1
            for y in draws[b]:
                pc[(x,y)] += 1
    yf = defaultdict(int)
    for i in ref_issues:
        for y in draws[i]: yf[y] += 1
    tot = sum(yf.values())
    lifts = {}
    for (x,y), c in pc.items():
        if xc[x] >= 3 and yf[y] >= 3:
            lifts[(x,y)] = (c/xc[x]) / (yf[y]/tot)
    return lifts

per = []
for iss in test_issues:
    if iss-1 not in draws: continue
    ref = [i for i in issues if i < iss]
    lifts = train_pairs(ref)
    prev = draws[iss-1]
    yscore = defaultdict(float)
    for y in draws[iss]: pass
    # 对上一期每个开出X, 累加Y的lift; 只考虑未在上期开出的Y
    for y in range(1, 81):
        if y in prev: continue
        s = 0
        for x in prev:
            s += lifts.get((x,y), 1.0)
        yscore[y] = s
    # 选Top-K (排除上期已出)
    top = sorted(yscore.items(), key=lambda kv: -kv[1])[:20]
    picks = [y for y, _ in top]
    per.append(len(set(picks) & draws[iss]))
print(f'跟随对Top20(排除上期已出): 期均命中 {statistics.mean(per):.2f} (随机期望5.0)')

# 随机基线
import random
rng = random.Random(3)
rnd = []
for iss in test_issues:
    prev = draws.get(iss-1, set())
    alln = [i for i in range(1,81) if i not in prev]
    picks = rng.sample(alln, 20)
    rnd.append(len(set(picks) & draws[iss]))
print(f'随机20码(排除上期已出): 期均 {statistics.mean(rnd):.2f}')