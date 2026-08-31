# -*- coding: utf-8 -*-
"""跟随号码统计 — v5: 点位 / 承接 / 错位 / 数据2全池排序
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)

# 点位
points = {}
with open(r'D:\Dpanqianyi\Python-Project\data\daily_points.txt', encoding='utf-8') as f:
    for line in f:
        m = re.search(r'period:(\d+)', line)
        pm = re.search(r'points:([\d\s]+)', line)
        if m and pm:
            points[int(m.group(1))] = set(int(p) for p in pm.group(1).split() if p)

records = []
all_rows = list(ws.iter_rows())
BLOCK_OFFSETS = [1, 6, 11, 16]
for r_idx, row in enumerate(all_rows):
    first_val = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
    if not m: continue
    issue = int(m.group(1)); dtype = int(m.group(2))
    for b_idx, offset in enumerate(BLOCK_OFFSETS):
        for row_off in range(4):
            ri = r_idx + offset + row_off
            if ri >= len(all_rows): continue
            trow = all_rows[ri]
            for col_idx in list(range(4)) + list(range(5, 9)):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                side = 'L' if col_idx < 4 else 'R'
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    block=b_idx, side=side, row=row_off,
                    col=col_idx if side=='L' else col_idx-5))

def uniq_pool(iss, star_only=True, dtypes=(1,2)):
    d = {}
    for r in records:
        if r['issue'] != iss: continue
        if r['dtype'] not in dtypes: continue
        if star_only and not r['star']: continue
        if r['num'] not in d:
            d[r['num']] = dict(blocks=set(), sides=set(), points=0, dtypes=set())
        d[r['num']]['blocks'].add(r['block'])
        d[r['num']]['sides'].add(r['side'])
        d[r['num']]['points'] = max(d[r['num']]['points'], int(r['point']))
        d[r['num']]['dtypes'].add(r['dtype'])
    return d

issues = sorted(set(r['issue'] for r in records if r['issue'] in draws))
print(f'期数: {len(issues)}')

# ═══ 1. 点位命中率 ═══
print('\n═══ 1. 点位(用户提供) 命中率 ═══')
pt_hits = []; pt_avail = []
for iss in issues:
    if iss not in points: continue
    pt = points[iss]
    pt_hits.append(len(pt & draws[iss]))
    pt_avail.append(iss)
print(f'点位: 期均命中 {statistics.mean(pt_hits):.2f} / 20 = {statistics.mean(pt_hits)/20:.4f} (n={len(pt_hits)}期)')
print(f'随机基线(20码): 5.0')

# 点位×数据交集
print('\n═══ 2. 点位×数据 交集 ("承接") ═══')
for iss in issues:
    if iss not in points: continue
    pass
cross = defaultdict(lambda: [0,0])
for iss in [i for i in issues if i in points]:
    pt = points[iss]
    d1 = set(uniq_pool(iss, True, dtypes=(1,)))
    d2all = set(uniq_pool(iss, False, dtypes=(2,)))
    # 交集
    combos = {
        '点位∩数据1星': pt & d1,
        '点位∩数据2全': pt & d2all,
        '点位\数据(独有点位)': pt - d2all,
        '数据2全\点位': d2all - pt,
        '数据1星\点位': d1 - pt,
    }
    for k, nums in combos.items():
        cross[k][1] += len(nums)
        cross[k][0] += len(nums & draws[iss])
for k in ['点位∩数据1星','点位∩数据2全','点位\\数据(独有点位)','数据2全\\点位','数据1星\\点位']:
    if cross[k][1]:
        print(f'  {k:16s} n={cross[k][1]:5d} 命中率={cross[k][0]/cross[k][1]:.4f}')

# ═══ 3. 错位预测: 数据N → 开奖N+1 ═══
print('\n═══ 3. 错位预测 (N期数据 → N+1期开奖) ═══')
for lag in [1, 2]:
    hits = []
    for iss in issues:
        if iss+lag not in draws: continue
        pool = uniq_pool(iss, True)
        hits.append(len(set(pool) & draws[iss+lag]))
    print(f'数据星lag{lag}: 期均 {statistics.mean(hits):.2f} (池~20码, 随机期望5.0)')

# ═══ 4. 数据2全池排序选号 (无前视) ═══
print('\n═══ 4. 数据2全池(65码)排序选号 ═══')
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]

def logit(p):
    p = max(1e-6, min(1-1e-6, p))
    return math.log(p/(1-p))

def train_weights(train_issues):
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        pool = uniq_pool(iss, False, dtypes=(2,))  # 数据2全池
        for n, attrs in pool.items():
            hit = 1 if n in draws[iss] else 0
            z = (n-1)//10; t = n%10
            fc[f"Z{z}"][1]+=1; fc[f"Z{z}"][0]+=hit
            fc[f"尾{t}"][1]+=1; fc[f"尾{t}"][0]+=hit
            if iss-1 in draws:
                rep = 1 if n in draws[iss-1] else 0
                fc[f"重{rep}"][1]+=1; fc[f"重{rep}"][0]+=hit
    base = sum(v[0] for v in fc.values())/sum(v[1] for v in fc.values())
    w = {}
    for k, (h, n) in fc.items():
        if n >= 25:
            w[k] = logit(h/n) - logit(base)
    return w, base

def score(n, iss, w, base):
    z = (n-1)//10; t = n%10
    s = logit(base)
    s += w.get(f"Z{z}", 0) + w.get(f"尾{t}", 0)
    if iss-1 in draws:
        s += w.get(f"重{1 if n in draws[iss-1] else 0}", 0)
    return s

print(f"候选: 数据2全池(约65码, 覆盖16.4/20)")
for top_k in [10, 15, 20, 25, 30, 40]:
    per = []
    for iss in test_issues:
        train_issues = [i for i in issues if i < iss]
        w, base = train_weights(train_issues)
        pool = uniq_pool(iss, False, dtypes=(2,))
        scored = sorted(pool.keys(), key=lambda n: score(n, iss, w, base), reverse=True)
        picks = scored[:top_k]
        per.append(len(set(picks) & draws[iss]))
    print(f'  Top{top_k:2d}: 期均命中 {statistics.mean(per):.2f} (随机取{top_k}码≈{16.4*top_k/65:.1f})')

# 随机基线明确
import random
rng = random.Random(1)
for top_k in [15, 20, 30]:
    per = []
    for iss in test_issues:
        pool = list(uniq_pool(iss, False, dtypes=(2,)).keys())
        picks = rng.sample(pool, min(top_k, len(pool)))
        per.append(len(set(picks) & draws[iss]))
    print(f'  随机Top{top_k:2d}: 期均 {statistics.mean(per):.2f}')