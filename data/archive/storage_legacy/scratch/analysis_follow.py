# -*- coding: utf-8 -*-
"""跟随号码统计 深度挖掘 — 结构解析 + 命中分析 v2"""
import openpyxl, re, sys, statistics
from collections import defaultdict, Counter
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

# ── 解析 跟随号码统计 ──
ws = wb['跟随号码统计']
rows = list(ws.iter_rows(values_only=True))
periods = []
cur = None
for i, r in enumerate(rows):
    v = r[0]
    if v and re.search(r'(\d+)期数据(1|2)', str(v)):
        if cur: periods.append(cur)
        m = re.search(r'(\d+)期数据(1|2)', str(v))
        cur = {'issue': int(m.group(1)), 'type': int(m.group(2)), 'rows': []}
        continue
    if cur is not None:
        cur['rows'].append(r)
if cur: periods.append(cur)

by_issue = defaultdict(dict)
for p in periods:
    by_issue[p['issue']][p['type']] = p['rows']
issues = sorted(by_issue.keys())

# ── 开奖数据 ──
wsd = wb['全量开奖数据']
draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        nums = [int(x) for x in r[3:23] if x is not None]
        draws[int(r[1])] = (set(nums), nums)

def cell_num(v):
    if v is None: return None
    s = str(v).strip()
    if not s or s == 'nan': return None
    m = re.match(r'^(\d+)\*?$', s)
    return int(m.group(1)) if m else None

def parse_block(block):
    """按 4组×4行 结构解析, 返回 16 个 (cycle, window, cells) 条目
    块内布局: [行0..3]=cycle0, [行5..8]=cycle1, [行10..13]=cycle2, [行15..18]=cycle3 (有空行分隔)
    但块 rows 已去掉表头; 通用法: 非空行的顺序即 row_idx 0..15"""
    entries = []  # (row_idx, [(col, num, starred)])
    ridx = 0
    for r in block:
        cells = [(ci, n, ('*' in str(v))) for ci, v in enumerate(r) if (n := cell_num(v)) is not None]
        if cells:
            entries.append((ridx, cells))
        ridx += 1
    return entries

# ── 行结构: 每期每块应有16行, 校验 ──
rowcnt = Counter()
for iss in issues:
    for t in (1, 2):
        if t in by_issue[iss]:
            rowcnt[len(parse_block(by_issue[iss][t]))] += 1
print('有效行数分布(应为16):', dict(rowcnt))

# ── 每行(0-15) 命中率 ──
# 数据1: 全带星; 数据2: 带星部分 + 全部
d1_row = defaultdict(list); d2star_row = defaultdict(list); d2all_row = defaultdict(list)
d1_col = defaultdict(list); d2star_col = defaultdict(list); d2all_col = defaultdict(list)
total_hits_d1 = 0; total_n_d1 = 0
for iss in issues:
    if iss not in draws: continue
    dset = draws[iss][0]
    for t in (1, 2):
        if t not in by_issue[iss]: continue
        for ridx, cells in parse_block(by_issue[iss][t]):
            for ci, n, st in cells:
                hit = 1 if n in dset else 0
                if t == 1:
                    d1_row[ridx].append(hit); d1_col[ci].append(hit)
                    total_hits_d1 += hit; total_n_d1 += 1
                else:
                    d2all_row[ridx].append(hit); d2all_col[ci].append(hit)
                    if st:
                        d2star_row[ridx].append(hit); d2star_col[ci].append(hit)

print('\n=== 数据1 每行(0-15)命中率 ===')
for ri in range(16):
    if ri in d1_row:
        v = d1_row[ri]
        print(f'row{ri:2d}: {statistics.mean(v):.3f} (n={len(v)})')

print('\n=== 数据2带* 每行命中率 ===')
for ri in range(16):
    if ri in d2star_row:
        v = d2star_row[ri]
        print(f'row{ri:2d}: {statistics.mean(v):.3f} (n={len(v)})')

print('\n=== 数据1 每列命中率 ===')
for ci in range(10):
    if ci in d1_col:
        v = d1_col[ci]
        print(f'col{ci}: {statistics.mean(v):.3f} (n={len(v)})')

print('\n=== 数据2带* 每列命中率 ===')
for ci in range(10):
    if ci in d2star_col:
        v = d2star_col[ci]
        print(f'col{ci}: {statistics.mean(v):.3f} (n={len(v)})')

print('\n随机基线: 0.250')
print(f'数据1 总: {total_hits_d1}/{total_n_d1} = {total_hits_d1/total_n_d1:.4f}')