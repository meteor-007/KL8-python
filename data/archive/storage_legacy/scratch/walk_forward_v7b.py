# -*- coding: utf-8 -*-
"""跟随号码统计 — v7b: 终版组合模型 无前视验证 (缓存优化版)
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if 2026089 <= i <= 2026209)

points = {}
with open(r'D:\Dpanqianyi\Python-Project\data\daily_points.txt', encoding='utf-8') as f:
    for line in f:
        m = re.search(r'period:(\d+)', line)
        pm = re.search(r'points:([\d\s]+)', line)
        if m and pm:
            points[int(m.group(1))] = set(int(p) for p in pm.group(1).split() if p)

# ── 一次性完整解析: per issue → {dtype: [(num, star_ord, block, side)]} 及 d2all ──
all_rows = list(ws.iter_rows())
period_cache = {}
for idx, row in enumerate(all_rows):
    v = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[\s\S]*?数据(1|2)', v)
    if not m: continue
    iss = int(m.group(1)); dtype = int(m.group(2))
    if iss not in period_cache:
        period_cache[iss] = {'d1': [], 'd2': [], 'd2all': set(), 'd1star_uniq': set()}
    pc = period_cache[iss]
    for b_idx, off in enumerate([1, 6, 11, 16]):
        for ro in range(4):
            if idx+off+ro >= len(all_rows): continue
            trow = all_rows[idx+off+ro]
            cells = []
            for c in range(min(10, len(trow))):
                cv = str(trow[c].value or "").strip()
                if cv and cv != 'nan':
                    n = int(cv.replace('*',''))
                    st = '*' in cv
                    if st:
                        cells.append((c, n))
                    if dtype == 2:
                        pc['d2all'].add(n)
            cells.sort()
            for si, (c, n) in enumerate(cells):
                side = 'L' if c < 4 else 'R'
                rec = (n, si, b_idx, side)
                if dtype == 1:
                    pc['d1'].append(rec); pc['d1star_uniq'].add(n)
                else:
                    pc['d2'].append(rec)

print(f'解析完成: {len(period_cache)} 期')

def logit(p):
    p = max(1e-6, min(1-1e-6, p))
    return math.log(p/(1-p))

def build_weights(train_issues, src):
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        recs = period_cache[iss][src]
        for n, si, b, side in recs:
            hit = 1 if n in draws[iss] else 0
            fc[f"星{min(si,7)}"][1]+=1; fc[f"星{min(si,7)}"][0]+=hit
            fc[f"侧{side}"][1]+=1; fc[f"侧{side}"][0]+=hit
            fc[f"尾{n%10}"][1]+=1; fc[f"尾{n%10}"][0]+=hit
            fc[f"Z{(n-1)//10}"][1]+=1; fc[f"Z{(n-1)//10}"][0]+=hit
            if iss-1 in draws:
                k = f"重{1 if n in draws[iss-1] else 0}"
                fc[k][1]+=1; fc[k][0]+=hit
    base = sum(v[0] for v in fc.values())/sum(v[1] for v in fc.values())
    w = {}
    for k, (h, n) in fc.items():
        if n >= 30:
            w[k] = logit(h/n) - logit(base)
    return w, base

def score_rec(n, si, side, iss, w, base):
    s = logit(base)
    s += (w.get(f"星{min(si,7)}", 0) + w.get(f"侧{side}", 0) + w.get(f"尾{n%10}", 0)
          + w.get(f"Z{(n-1)//10}", 0))
    if iss-1 in draws:
        s += w.get(f"重{1 if n in draws[iss-1] else 0}", 0)
    return s

TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]
print(f'测试期: {TEST_START}~{issues[-1]} ({len(test_issues)}期)')

import random
rng = random.Random(11)
print('\n═══ 终版组合模型 (无前视, 池内去重取最高分) ═══')
for src, label in [('d1', '数据1星'), ('d2', '数据2星')]:
    print(f'── 池: {label} ──')
    for top_k in [5, 8, 12]:
        per = []; rnd = []
        for iss in test_issues:
            train_issues = [i for i in issues if i < iss]
            w, base = build_weights(train_issues, src)
            best = {}
            for n, si, b, side in period_cache[iss][src]:
                s = score_rec(n, si, side, iss, w, base)
                if n not in best or s > best[n]: best[n] = s
            sc = sorted(best.items(), key=lambda kv: -kv[1])
            picks = [n for n, _ in sc[:top_k]]
            per.append(len(set(picks) & draws[iss]))
            rp = rng.sample(list(best.keys()), min(top_k, len(best)))
            rnd.append(len(set(rp) & draws[iss]))
        print(f'  Top{top_k:2d}: 模型 {statistics.mean(per):.2f} | 随机 {statistics.mean(rnd):.2f} | 提升 {statistics.mean(per)-statistics.mean(rnd):+.2f}')

# ═══ 点位独特号 无前视 ═══
print('\n═══ 点位独特号 (点位 - 数据2全) ═══')
per_uniq = []; pool_sizes = []
for iss in test_issues:
    if iss not in points: continue
    uniq = points[iss] - period_cache[iss]['d2all']
    per_uniq.append(len(uniq & draws[iss])); pool_sizes.append(len(uniq))
print(f'点位独特号: 池均 {statistics.mean(pool_sizes):.1f} 码, 期均命中 {statistics.mean(per_uniq):.2f} = {statistics.mean(per_uniq)/statistics.mean(pool_sizes):.4f}')
# 点位非独特
per_ov = []; pool_ov = []
for iss in test_issues:
    if iss not in points: continue
    ov = points[iss] & period_cache[iss]['d2all']
    per_ov.append(len(ov & draws[iss])); pool_ov.append(len(ov))
print(f'点位∩数据2: 池均 {statistics.mean(pool_ov):.1f} 码, 期均命中 {statistics.mean(per_ov):.2f} = {statistics.mean(per_ov)/statistics.mean(pool_ov):.4f}')

# ═══ 简单规则 ═══
print('\n═══ 简单可解释规则 (无前视, 数据1星池) ═══')
rules = {
    '尾2/7/8/3+非重': lambda n, si, side, iss: (n%10 in (2,7,8,3)) and (iss-1 not in draws or n not in draws[iss-1]),
    '尾2/7/8/3': lambda n, si, side, iss: n%10 in (2,7,8,3),
    '右侧星(≥5号)': lambda n, si, side, iss: side == 'R',
    '右侧+尾2/7/8/3': lambda n, si, side, iss: side=='R' and n%10 in (2,7,8,3),
    '非重号': lambda n, si, side, iss: iss-1 not in draws or n not in draws[iss-1],
    '尾2+非重': lambda n, si, side, iss: n%10==2 and (iss-1 not in draws or n not in draws[iss-1]),
}
for name, fn in rules.items():
    per = []
    for iss in test_issues:
        cand = set()
        seen = set()
        for n, si, b, side in period_cache[iss]['d1']:
            seen.add(n)
            if fn(n, si, side, iss): cand.add(n)
        per.append(len(cand & draws[iss]) / max(1, len(cand)))
    print(f'  {name:14s}: 单码命中率 {statistics.mean(per):.4f} (基线0.25)')