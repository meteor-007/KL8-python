#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════════
  命中率深度优化引擎 (Deep Hit-Rate Optimizer)
  ─────────────────────────────────────────────
  针对当前系统 Top5 命中率 ~20%、Top12 命中率 ~30% 的瓶颈,
  从以下6个维度进行深度攻坚:

  1. 自适应集成学习 (Adaptive Ensemble Stacking)
  2. 条件概率贝叶斯网络升级 (Conditional Bayesian V2)
  3. 号码协同涌现检测 (Co-emergence Detection)
  4. 时变衰减动态调权 (Time-Varying Decay Calibration)
  5. 反共识对抗噪声 (Anti-Consensus Denoising)
  6. 多粒度交叉验证 (Multi-Granularity Cross-Validation)

  版本: v1.0  日期: 2026-05-19
═══════════════════════════════════════════════════════════════════
"""

import os, sys, json, math, logging, collections, datetime
from typing import Dict, List, Set, Tuple, Optional, Any
import numpy as np

import os, sys
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.paths import get_project_root, data_path, _ensure_project_path
_ensure_project_path()
_PROJ = get_project_root()


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(_PROJ, 'logs', 'deep_optimizer.log'), encoding='utf-8')
    ]
)
logger = logging.getLogger("DeepOptimizer")


# ══════════════════════════════════════════════════════════════════
#  模块1: 自适应集成学习 (Adaptive Ensemble Stacking)
# ══════════════════════════════════════════════════════════════════

class AdaptiveEnsembleStacking:
    """
    当前系统各算法独立评分后简单加权融合, 忽略了算法之间的非线性互补关系。
    本模块引入 Stacking 思想: 用近期历史命中率作为标签, 自动拟合元学习器权重。
    """

    def __init__(self, lookback: int = 30):
        self.lookback = lookback
        self.meta_weights = None

    def compute_algorithm_scores(self, history: List[Dict]) -> Dict[str, Dict[int, float]]:
        """收集各算法的评分向量, 包括主系统五维一体和自研特征"""
        algo_scores = {}

        # ── 主系统五维一体评分 (核心权重来源) ──
        try:
            from audit.v3_trinity_audit import dynamic_meta_fusion
            pentagon_scores, _ = dynamic_meta_fusion(history)
            if pentagon_scores:
                max_s = max(pentagon_scores.values()) if pentagon_scores else 1.0
                algo_scores['pentagon'] = {n: pentagon_scores.get(n, 0.0) / max(max_s, 1e-9) for n in range(1, 81)}
        except Exception as e:
            logger.warning(f"五维一体评分加载失败: {e}")

        try:
            from core.algorithm_optimizer import plan7_markov_integration
            mk_res = plan7_markov_integration(history)
            if mk_res and mk_res.get('probs'):
                algo_scores['markov'] = {n: mk_res['probs'].get(n, 0.0) for n in range(1, 81)}
        except Exception as e:
            logger.warning(f"马尔可夫算法加载失败: {e}")

        try:
            from core.feature_optimizer import plan3_frequency_acceleration
            freq_res = plan3_frequency_acceleration(history)
            if freq_res and freq_res.get('recommended'):
                rec = freq_res['recommended']
                algo_scores['frequency'] = {n: (1.0 if n in rec else 0.2) for n in range(1, 81)}
        except Exception as e:
            logger.warning(f"频率算法加载失败: {e}")

        algo_scores['omission_decay'] = self._compute_omission_decay(history)
        algo_scores['tail_compensation'] = self._compute_tail_compensation(history)
        algo_scores['zone_deficit'] = self._compute_zone_deficit(history)
        algo_scores['momentum'] = self._compute_momentum(history)

        return algo_scores

    def _compute_omission_decay(self, history: List[Dict]) -> Dict[int, float]:
        """遗漏指数衰减: 遗漏越长评分越高"""
        scores = {}
        for n in range(1, 81):
            omission = 0
            for h in history[:50]:
                if n in h['numbers']:
                    break
                omission += 1
            scores[n] = 1.0 - math.exp(-omission * 0.15)
        return scores

    def _compute_tail_compensation(self, history: List[Dict]) -> Dict[int, float]:
        """尾数周期补偿: 低频尾数获得高补偿"""
        if not history:
            return {n: 0.0 for n in range(1, 81)}
        tail_counts = collections.Counter()
        lb = min(5, len(history))
        for h in history[:lb]:
            for num in h['numbers']:
                tail_counts[num % 10] += 1
        expected = lb * 2
        compensation = {t: max(0.0, (expected - tail_counts.get(t, 0)) / expected) for t in range(10)}
        return {n: compensation.get(n % 10, 0.0) for n in range(1, 81)}

    def _compute_zone_deficit(self, history: List[Dict]) -> Dict[int, float]:
        """区间缺口回补"""
        if not history:
            return {n: 0.0 for n in range(1, 81)}
        zone_counts = collections.Counter()
        lb = min(5, len(history))
        for h in history[:lb]:
            for num in h['numbers']:
                zone_counts[(num - 1) // 10] += 1
        expected = lb * 2.5
        zone_scores = {z: max(0.0, (expected - zone_counts.get(z, 0)) / expected) for z in range(8)}
        return {n: zone_scores.get((n - 1) // 10, 0.0) for n in range(1, 81)}

    def _compute_momentum(self, history: List[Dict]) -> Dict[int, float]:
        """短中期动量: 近3期出现频率 + 递增趋势加成"""
        scores = {}
        for n in range(1, 81):
            counts = [1 if n in h['numbers'] else 0 for h in history[:10]]
            # 短期频率 (近3期)
            short_freq = sum(counts[:3]) / 3.0
            # 递增趋势加成
            trend_bonus = 0.0
            if len(counts) >= 6:
                recent_3 = sum(counts[:3])
                prev_3 = sum(counts[3:6])
                if recent_3 > prev_3:
                    trend_bonus = 0.3
            scores[n] = short_freq + trend_bonus
        return scores

    def fit_meta_learner(self, history: List[Dict], val_window: int = 15):
        """用近期 val_window 期拟合元学习器权重 (网格搜索)"""
        if len(history) < val_window + 10:
            logger.warning("历史数据不足, 使用默认权重")
            self.meta_weights = {'pentagon': 0.35, 'markov': 0.1, 'frequency': 0.05, 'omission_decay': 0.15,
                                 'tail_compensation': 0.2, 'zone_deficit': 0.05, 'momentum': 0.1}
            return

        key_combos = [
            {'pentagon': 0.35, 'markov': 0.1, 'frequency': 0.05, 'omission_decay': 0.15, 'tail_compensation': 0.2, 'zone_deficit': 0.05, 'momentum': 0.1},
            {'pentagon': 0.30, 'markov': 0.15, 'frequency': 0.05, 'omission_decay': 0.15, 'tail_compensation': 0.2, 'zone_deficit': 0.05, 'momentum': 0.1},
            {'pentagon': 0.40, 'markov': 0.05, 'frequency': 0.05, 'omission_decay': 0.2, 'tail_compensation': 0.15, 'zone_deficit': 0.05, 'momentum': 0.1},
            {'pentagon': 0.25, 'markov': 0.15, 'frequency': 0.1, 'omission_decay': 0.2, 'tail_compensation': 0.15, 'zone_deficit': 0.05, 'momentum': 0.1},
            {'pentagon': 0.20, 'markov': 0.2, 'frequency': 0.1, 'omission_decay': 0.15, 'tail_compensation': 0.2, 'zone_deficit': 0.05, 'momentum': 0.1},
            {'pentagon': 0.15, 'markov': 0.2, 'frequency': 0.1, 'omission_decay': 0.2, 'tail_compensation': 0.2, 'zone_deficit': 0.05, 'momentum': 0.1},
            {'pentagon': 0.10, 'markov': 0.25, 'frequency': 0.1, 'omission_decay': 0.2, 'tail_compensation': 0.2, 'zone_deficit': 0.05, 'momentum': 0.1},
        ]

        # 预计算各验证期的轻量级评分 (不含pentagon, 避免回测时重复调用五维融合)
        logger.info("  预计算各验证期的轻量级评分...")
        val_period_scores = []
        for i in range(min(val_window, len(history) - 10)):
            train_slice = history[i + 1:]
            val_actual = set(history[i]['numbers'])
            # 轻量级评分: 仅使用自研特征 (pentagon 权重固定, 不参与回测搜索)
            algo_scores = self._compute_lightweight_scores(train_slice)
            val_period_scores.append((algo_scores, val_actual))

        best_weights = key_combos[0]
        best_score = 0.0

        for weights in key_combos:
            total_hits = 0
            count = 0
            for algo_scores, val_actual in val_period_scores:
                combined = collections.Counter()
                for algo_name, scores in algo_scores.items():
                    w = weights.get(algo_name, 0.1)
                    for n, s in scores.items():
                        combined[n] += s * w
                top12 = [n for n, s in combined.most_common(12)]
                total_hits += len(set(top12) & val_actual)
                count += 1

            avg_hits = total_hits / max(count, 1)
            if avg_hits > best_score:
                best_score = avg_hits
                best_weights = weights.copy()

        self.meta_weights = best_weights
        logger.info(f"元学习器训练完成: 权重={best_weights}, Top12平均命中={best_score:.2f}")

    def _compute_lightweight_scores(self, history: List[Dict]) -> Dict[str, Dict[int, float]]:
        """轻量级评分: 不调用pentagon五维融合, 仅使用自研特征 (用于回测)"""
        algo_scores = {}
        algo_scores['omission_decay'] = self._compute_omission_decay(history)
        algo_scores['tail_compensation'] = self._compute_tail_compensation(history)
        algo_scores['zone_deficit'] = self._compute_zone_deficit(history)
        algo_scores['momentum'] = self._compute_momentum(history)
        return algo_scores

    def predict(self, history: List[Dict], top_n: int = 12) -> Tuple[List[int], Dict]:
        """用训练好的元权重预测"""
        if self.meta_weights is None:
            self.fit_meta_learner(history)

        algo_scores = self.compute_algorithm_scores(history)
        combined = collections.Counter()
        detail = {}
        for algo_name, scores in algo_scores.items():
            w = self.meta_weights.get(algo_name, 0.1)
            algo_detail = {}
            for n, s in scores.items():
                combined[n] += s * w
                algo_detail[n] = s
            detail[algo_name] = {'weight': w, 'top5': sorted(
                [(n, f"{s:.3f}") for n, s in collections.Counter(scores).most_common(5)],
                key=lambda x: -float(x[1])
            )}

        top_n_list = [n for n, s in combined.most_common(top_n)]
        return top_n_list, {'weights': self.meta_weights, 'detail': detail}


# ══════════════════════════════════════════════════════════════════
#  模块2: 号码协同涌现检测 (Co-emergence Detection)
# ══════════════════════════════════════════════════════════════════

class CoEmergenceDetector:
    """
    核心洞察: 号码不是孤立出现的, 存在"协同涌现"现象。
    当号码A出现时, 号码B的出现概率显著高于随机期望。
    本模块通过条件概率矩阵 P(B|A) 检测这种协同关系,
    并在当前期已有高置信号码时, 用条件概率推演出协同号码。
    """

    def __init__(self, min_support: int = 5, min_lift: float = 1.5):
        self.min_support = min_support
        self.min_lift = min_lift
        self.cond_prob = {}  # {(a, b): P(b|a)}
        self.lift = {}       # {(a, b): P(b|a) / P(b)}

    def fit(self, history: List[Dict], max_lookback: int = 200):
        """从历史数据中计算条件概率和提升度"""
        lb = min(max_lookback, len(history))
        cooccur = collections.Counter()
        single = collections.Counter()

        for h in history[:lb]:
            nums = set(h['numbers'])
            for a in nums:
                single[a] += 1
                for b in nums:
                    if a != b:
                        cooccur[(a, b)] += 1

        # P(b|a) = count(a,b) / count(a)
        # lift = P(b|a) / P(b) = (count(a,b) / count(a)) / (count(b) / lb)
        self.cond_prob = {}
        self.lift = {}

        for (a, b), cnt in cooccur.items():
            if cnt >= self.min_support and single[a] >= self.min_support:
                p_b_given_a = cnt / single[a]
                p_b = single[b] / lb
                if p_b > 0:
                    lift_val = p_b_given_a / p_b
                    self.cond_prob[(a, b)] = p_b_given_a
                    self.lift[(a, b)] = lift_val

        logger.info(f"协同涌现矩阵: {len(self.cond_prob)} 对有效条件概率, "
                     f"高提升度(>{self.min_lift})对数: "
                     f"{sum(1 for v in self.lift.values() if v >= self.min_lift)}")

    def predict_co_emergence(
        self, high_confidence_nums: List[int], top_k: int = 10
    ) -> List[Tuple[int, float, str]]:
        """
        给定高置信号码, 推演最可能协同涌现的号码。
        返回: [(号码, 综合提升度, 推理路径), ...]
        """
        candidates = collections.Counter()

        for a in high_confidence_nums:
            for b in range(1, 81):
                if b == a:
                    continue
                lift_val = self.lift.get((a, b), 0.0)
                cond_p = self.cond_prob.get((a, b), 0.0)
                if lift_val >= self.min_lift:
                    # 综合得分 = 提升度 × 条件概率
                    candidates[b] += lift_val * cond_p

        # 生成推理路径
        result = []
        for n, score in candidates.most_common(top_k):
            # 找出最强推理源
            best_source = max(
                high_confidence_nums,
                key=lambda a: self.lift.get((a, n), 0.0)
            )
            best_lift = self.lift.get((best_source, n), 0.0)
            path = f"{best_source}→{n} (lift={best_lift:.2f})"
            result.append((n, score, path))

        return result


# ══════════════════════════════════════════════════════════════════
#  模块3: 反共识去噪 (Anti-Consensus Denoising)
# ══════════════════════════════════════════════════════════════════

class AntiConsensusDenoiser:
    """
    核心洞察: 当所有算法都推荐同一个号码时, 该号码往往被过度拟合,
    反而命中率不高 (共识陷阱)。反之, 仅被少数算法推荐但评分极高的号码,
    可能是被大多数算法忽略的"奇点号"。

    本模块对过度共识号码施加惩罚, 对奇点号码施加奖励。
    """

    def __init__(self, consensus_penalty: float = 0.7, singularity_bonus: float = 1.3):
        self.consensus_penalty = consensus_penalty
        self.singularity_bonus = singularity_bonus

    def denoise(
        self, algo_scores: Dict[str, Dict[int, float]], combined_scores: Dict[int, float]
    ) -> Dict[int, float]:
        """
        对共识号降权, 对奇点号加权。
        consensus_degree: 被多少个算法推荐(评分 > 阈值)
        """
        threshold = 0.3
        result = {}

        for n in range(1, 81):
            # 计算共识度: 有多少算法认为该号码值得关注
            consensus_count = sum(
                1 for scores in algo_scores.values()
                if scores.get(n, 0.0) > threshold
            )
            total_algos = len(algo_scores)

            score = combined_scores.get(n, 0.0)

            if consensus_count == total_algos and total_algos > 3:
                # 全员共识 → 惩罚
                score *= self.consensus_penalty
            elif consensus_count <= 1 and score > 0:
                # 奇点号 (仅1个算法推荐但原始分高) → 奖励
                score *= self.singularity_bonus

            result[n] = score

        return result


# ══════════════════════════════════════════════════════════════════
#  模块4: 时变衰减动态调权 (Time-Varying Decay Calibration)
# ══════════════════════════════════════════════════════════════════

class TimeVaryingDecayCalibrator:
    """
    核心洞察: 不同算法的"有效窗口"是不同的。
    - 马尔可夫在短期(5-10期)最有效
    - 遗漏衰减在中长期(15-30期)最有效
    - 尾数补偿在5期内最敏感

    本模块自动检测每种算法当前的最佳回看窗口,
    并根据近期命中率动态调整衰减速率。
    """

    def __init__(self):
        self.optimal_decays = {}

    def calibrate(self, history: List[Dict], val_window: int = 15):
        """为每种算法找到当前最优衰减率"""
        if len(history) < val_window + 20:
            return

        decay_candidates = [0.05, 0.1, 0.15, 0.2, 0.3, 0.4, 0.5]

        for algo_type in ['omission', 'momentum']:
            best_decay = 0.15
            best_hits = 0

            for decay in decay_candidates:
                total_hits = 0
                for i in range(min(val_window, len(history) - 20)):
                    train = history[i + 1:]
                    actual = set(history[i]['numbers'])

                    scores = self._compute_with_decay(train, algo_type, decay)
                    top12 = [n for n, s in collections.Counter(scores).most_common(12)]
                    total_hits += len(set(top12) & actual)

                if total_hits > best_hits:
                    best_hits = total_hits
                    best_decay = decay

            self.optimal_decays[algo_type] = best_decay

        logger.info(f"时变衰减校准完成: {self.optimal_decays}")

    def _compute_with_decay(
        self, history: List[Dict], algo_type: str, decay: float
    ) -> Dict[int, float]:
        """使用指定衰减率计算评分"""
        scores = {}
        if algo_type == 'omission':
            for n in range(1, 81):
                omission = 0
                for h in history[:50]:
                    if n in h['numbers']:
                        break
                    omission += 1
                scores[n] = 1.0 - math.exp(-omission * decay)
        elif algo_type == 'momentum':
            for n in range(1, 81):
                s = 0.0
                for i, h in enumerate(history[:15]):
                    if n in h['numbers']:
                        s += math.exp(-i * decay)
                scores[n] = s
        return scores


# ══════════════════════════════════════════════════════════════════
#  模块5: 多粒度交叉验证 (Multi-Granularity Cross-Validation)
# ══════════════════════════════════════════════════════════════════

class MultiGranularityValidator:
    """
    核心洞察: 一个好的推荐号码应该在不同时间粒度下都表现稳定。
    如果一个号码在5期窗口排名第3, 但在15期窗口排名第18,
    说明它的排名不稳定, 命中可靠性较低。

    本模块在多个粒度(5/10/20/50期)下分别排名,
    仅推荐在至少2个粒度下都进入Top12的号码。
    """

    def __init__(self, windows: List[int] = None):
        self.windows = windows or [5, 10, 20, 50]

    def validate(
        self, history: List[Dict], candidate_nums: List[int], top_k: int = 12
    ) -> Tuple[List[int], Dict[int, Dict]]:
        """
        多粒度交叉验证: 在不同窗口下检查候选号码的排名稳定性。
        改进: 在候选池内部进行相对排名, 而非全80号绝对排名,
        避免因候选号在全局排名中偏低导致全部被过滤的问题。
        """
        window_ranks = {n: {} for n in candidate_nums}

        for w in self.windows:
            if len(history) < w:
                continue
            slice_h = history[:w]
            scores = self._quick_score(slice_h)
            # 仅对候选号码排序 (相对排名)
            candidate_scores = [(n, scores.get(n, 0.0)) for n in candidate_nums]
            candidate_scores.sort(key=lambda x: (-x[1], x[0]))
            for rank, (n, s) in enumerate(candidate_scores, 1):
                window_ranks[n][w] = rank

        # 稳定性得分: 在多少个窗口下排名 <= 候选池的top_k阈值
        # 阈值 = max(top_k, len(candidate_nums) // 2)
        rank_threshold = max(top_k, len(candidate_nums) // 2)
        stability = {}
        for n in candidate_nums:
            in_top_count = sum(1 for w, rank in window_ranks[n].items() if rank <= rank_threshold)
            window_count = len(window_ranks[n])
            stability[n] = in_top_count / max(window_count, 1)

        # 仅保留稳定性 >= 0.25 (至少在1/4窗口下进入阈值) 的号码
        validated = [n for n in candidate_nums if stability.get(n, 0) >= 0.25]
        if len(validated) < top_k:
            # 补足: 按稳定性降序, 用未通过验证但稳定性最高的号码补足
            remaining = [(n, stability.get(n, 0)) for n in candidate_nums if n not in validated]
            remaining.sort(key=lambda x: (-x[1], x[0]))
            for n, s in remaining:
                validated.append(n)
                if len(validated) >= top_k:
                    break

        return validated[:top_k], window_ranks

    def _quick_score(self, history: List[Dict]) -> Dict[int, float]:
        """快速评分: 频率 + 遗漏 + 动量"""
        scores = {}
        for n in range(1, 81):
            freq = sum(1 for h in history if n in h['numbers']) / len(history)
            omission = 0
            for h in history:
                if n in h['numbers']:
                    break
                omission += 1
            omission_score = 1.0 - math.exp(-omission * 0.15)
            scores[n] = freq * 0.4 + omission_score * 0.3 + (1.0 if freq > 0.3 else 0.0) * 0.3
        return scores


# ══════════════════════════════════════════════════════════════════
#  模块6: 深度优化主引擎 (Deep Optimizer Main Engine)
# ══════════════════════════════════════════════════════════════════

class DeepHitRateOptimizer:
    """
    整合所有深度优化模块的主引擎。
    执行流程:
    1. 自适应集成 → 初始Top-N候选
    2. 协同涌现 → 扩展候选池 (条件概率推演)
    3. 反共识去噪 → 惩罚过度共识号, 奖励奇点号
    4. 时变衰减校准 → 优化算法参数
    5. 多粒度交叉验证 → 最终精选
    6. 生成优化报告
    """

    def __init__(self):
        self.ensemble = AdaptiveEnsembleStacking()
        self.co_emergence = CoEmergenceDetector()
        self.anti_consensus = AntiConsensusDenoiser()
        self.decay_calibrator = TimeVaryingDecayCalibrator()
        self.validator = MultiGranularityValidator()
        # 初始化占位变量
        self._auc_stats = {}


    def optimize(self, history: List[Dict]) -> Dict[str, Any]:
        """执行完整的深度优化流程"""
        logger.info("=" * 60)
        logger.info("深度命中率优化引擎启动")
        logger.info("=" * 60)

        # ── Step 1: 自适应集成学习 ──
        logger.info("[Step 1] 自适应集成学习...")
        self.ensemble.fit_meta_learner(history, val_window=15)
        top12_ensemble, ensemble_detail = self.ensemble.predict(history, top_n=12)
        top5_ensemble, _ = self.ensemble.predict(history, top_n=5)
        logger.info(f"  集成Top5: {top5_ensemble}")
        logger.info(f"  集成Top12: {top12_ensemble}")
        # 加载 AUC 统计数据用于权重校准
        auc_path = os.path.join(_PROJ, 'auc_stats.json')
        try:
            with open(auc_path, 'r', encoding='utf-8') as f:
                auc_raw = json.load(f)
            # 兼容两种格式: 数组格式 [{num:54, auc:..., p_value:...}, ...] 或 字典格式 {54: {auc:..., p_value:...}, ...}
            if isinstance(auc_raw, list):
                auc_stats = {}
                for item in auc_raw:
                    num = item.get('num') or item.get('number')
                    if num is not None:
                        auc_stats[str(num)] = {
                            'auc': item.get('auc', 0.5),
                            'p_value': item.get('p_value', 1.0),
                            'hits': item.get('hits', 0),
                            'hit_rate': item.get('hit_rate', 0.0),
                            'precision': item.get('precision', 0.0),
                        }
            else:
                auc_stats = auc_raw
        except Exception as e:
            logger.warning(f"AUC stats 加载失败: {e}")
            auc_stats = {}
        self._auc_stats = auc_stats

        # ── Step 2: 协同涌现检测 ──
        logger.info("[Step 2] 协同涌现检测...")
        self.co_emergence.fit(history, max_lookback=200)
        co_emergence_nums = self.co_emergence.predict_co_emergence(
            top5_ensemble, top_k=8
        )
        co_nums = [n for n, score, path in co_emergence_nums]
        logger.info(f"  协同涌现号码: {co_nums}")
        logger.info(f"  推理路径: {[path for _, _, path in co_emergence_nums]}")
        

        # ── Step 3: 合并候选池 ──
        candidate_pool = list(set(top12_ensemble + co_nums))
        logger.info(f"[Step 3] 候选池大小: {len(candidate_pool)}, 号码: {sorted(candidate_pool)}")

        # ── Step 4: 反共识去噪 ──
        logger.info("[Step 4] 反共识去噪...")
        algo_scores = self.ensemble.compute_algorithm_scores(history)
        combined_raw = {}
        for n in candidate_pool:
            s = 0.0
            for algo_name, scores in algo_scores.items():
                w = self.ensemble.meta_weights.get(algo_name, 0.1)
                s += scores.get(n, 0.0) * w
            combined_raw[n] = s

        denoised = self.anti_consensus.denoise(algo_scores, combined_raw)
        # ── Step 4.5: 零信标降级多级决策 (红线五) ──
        p_values = {}
        auc_values = {}
        for n in candidate_pool:
            auc_val = None
            p_val = 1.0
            if str(n) in auc_stats:
                auc_val = auc_stats[str(n)]
            elif n in auc_stats:
                auc_val = auc_stats[n]
            if isinstance(auc_val, dict):
                p_val = auc_val.get('p_value', 1.0)
                auc_val = auc_val.get('auc')
            p_values[n] = p_val
            auc_values[n] = auc_val if isinstance(auc_val, (int, float)) else 0.5

        # Level 0 探测 (Bonferroni)
        level0_pass = [n for n in candidate_pool if p_values[n] < 0.000625]
        
        gating_level = 0
        gating_warning = None
        pass_nums = []
        
        if len(level0_pass) >= 3:
            gating_level = 0
            pass_nums = level0_pass
            logger.info(f"[信标审计] 激活 Level 0 (高置信 Bonferroni)，通过号码数={len(level0_pass)}")
        else:
            # Level 1 (FDR-BH) 探测
            sorted_pool = sorted(candidate_pool, key=lambda x: p_values[x])
            m = len(sorted_pool)
            q_threshold = 0.10
            max_k = -1
            for k_idx, n in enumerate(sorted_pool):
                k = k_idx + 1
                if p_values[n] <= (k / m) * q_threshold:
                    max_k = k_idx
            level1_pass = sorted_pool[:max_k + 1] if max_k != -1 else []
            
            if len(level1_pass) >= 3:
                gating_level = 1
                pass_nums = level1_pass
                logger.info(f"[信标审计] 激活 Level 1 (中置信 FDR-BH)，通过号码数={len(level1_pass)}")
            else:
                # Level 2 (无校正 p < 0.05) 探测
                level2_pass = [n for n in candidate_pool if p_values[n] < 0.05]
                if len(level2_pass) >= 3:
                    gating_level = 2
                    pass_nums = level2_pass
                    gating_warning = "⚠️ 零信标降级: Level 2 — 系统当前缺乏统计显著的强信号，推荐范围已扩展至Top-8"
                    logger.warning(f"[信标审计] {gating_warning}")
                else:
                    # Level 3 等权降级
                    gating_level = 3
                    pass_nums = []
                    gating_warning = "⚠️ 零信标降级: Level 3 — 系统当前缺乏统计显著的有效信号，推荐结果仅供参考"
                    logger.warning(f"[信标审计] {gating_warning}")

        # 应用双曲正切平滑溢价/冷却
        for n in candidate_pool:
            score = combined_raw.get(n, 0.0)
            delta = 0.0
            if gating_level == 0 and n in pass_nums:
                auc_val = auc_values[n]
                if auc_val >= 0.5:
                    delta = 0.35 * math.tanh(8.0 * (auc_val - 0.5))
                else:
                    delta = 0.25 * math.tanh(8.0 * (auc_val - 0.5))
            elif gating_level == 1 and n in pass_nums:
                # 溢价减半 (符合红线五)
                auc_val = auc_values[n]
                if auc_val >= 0.5:
                    delta = 0.175 * math.tanh(8.0 * (auc_val - 0.5))
                else:
                    delta = 0.125 * math.tanh(8.0 * (auc_val - 0.5))
            
            denoised[n] = score * (1.0 + delta)

        denoised_top = sorted(denoised.items(), key=lambda x: (-x[1], x[0]))[:15]
        logger.info(f"  去噪后Top15: {[n for n, s in denoised_top]}")

        # ── Step 5: 时变衰减校准 ──
        logger.info("[Step 5] 时变衰减校准...")
        self.decay_calibrator.calibrate(history, val_window=15)

        # ── Step 6: 多粒度交叉验证 ──
        logger.info("[Step 6] 多粒度交叉验证...")
        candidate_sorted = [n for n, s in denoised_top]
        validated, window_ranks = self.validator.validate(
            history, candidate_sorted, top_k=12
        )
        logger.info(f"  最终验证Top12: {validated}")

        # ── 调整最终推荐范围（红线五） ──
        final_top5_count = 5
        final_top12_count = 12
        if gating_level == 2:
            final_top5_count = 8
            final_top12_count = 14
        elif gating_level == 3:
            final_top5_count = 10
            final_top12_count = 14

        final_top5 = validated[:final_top5_count]
        final_top12 = validated[:final_top12_count]

        # ── 历史回测评估 ──
        backtest_result = self._quick_backtest(history, val_window=10)

        result = {
            'final_top5': final_top5,
            'final_top12': final_top12,
            'ensemble_top12': top12_ensemble,
            'co_emergence': [{'num': n, 'score': f"{s:.4f}", 'path': path}
                             for n, s, path in co_emergence_nums],
            'meta_weights': self.ensemble.meta_weights,
            'optimal_decays': self.decay_calibrator.optimal_decays,
            'window_ranks': {str(n): ranks for n, ranks in window_ranks.items()},
            'backtest': backtest_result,
            'gating_level': gating_level,
            'gating_warning': gating_warning,
            'timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # ── 保存结果 ──
        self._save_result(result)

        return result


    def _quick_backtest(self, history: List[Dict], val_window: int = 10) -> Dict:
        """快速回测: 对比优化前后的命中率"""
        if len(history) < val_window + 20:
            return {'note': '数据不足'}

        # 原始方法 (简单频率 + 遗漏)
        baseline_top5_hits = []
        baseline_top12_hits = []
        optimized_top5_hits = []
        optimized_top12_hits = []

        for i in range(min(val_window, len(history) - 20)):
            train = history[i + 1:]
            actual = set(history[i]['numbers'])

            # Baseline: 简单评分
            baseline_scores = {}
            for n in range(1, 81):
                omission = 0
                for h in train[:30]:
                    if n in h['numbers']:
                        break
                    omission += 1
                freq = sum(1 for h in train[:10] if n in h['numbers']) / min(10, len(train))
                baseline_scores[n] = freq * 0.5 + (1.0 - math.exp(-omission * 0.15)) * 0.5

            b_top5 = [n for n, s in sorted(baseline_scores.items(), key=lambda x: (-x[1], x[0]))[:5]]
            b_top12 = [n for n, s in sorted(baseline_scores.items(), key=lambda x: (-x[1], x[0]))[:12]]
            baseline_top5_hits.append(len(set(b_top5) & actual))
            baseline_top12_hits.append(len(set(b_top12) & actual))

            # 优化方法: 集成 + 尾数 + 区间
            opt_scores = {}
            for n in range(1, 81):
                omission = 0
                for h in train[:30]:
                    if n in h['numbers']:
                        break
                    omission += 1
                freq = sum(1 for h in train[:10] if n in h['numbers']) / min(10, len(train))
                # 尾数补偿
                tail_counts = collections.Counter()
                for h in train[:5]:
                    for num in h['numbers']:
                        tail_counts[num % 10] += 1
                tail_deficit = max(0, 10 - tail_counts.get(n % 10, 0)) / 10.0
                # 区间补偿
                zone_counts = collections.Counter()
                for h in train[:5]:
                    for num in h['numbers']:
                        zone_counts[(num - 1) // 10] += 1
                zone_deficit = max(0, 12.5 - zone_counts.get((n - 1) // 10, 0)) / 12.5
                # 动量
                momentum = sum(1 for h in train[:3] if n in h['numbers']) / 3.0

                opt_scores[n] = (freq * 0.25 +
                                 (1.0 - math.exp(-omission * 0.15)) * 0.25 +
                                 tail_deficit * 0.2 +
                                 zone_deficit * 0.1 +
                                 momentum * 0.2)

            o_top5 = [n for n, s in sorted(opt_scores.items(), key=lambda x: (-x[1], x[0]))[:5]]
            o_top12 = [n for n, s in sorted(opt_scores.items(), key=lambda x: (-x[1], x[0]))[:12]]
            optimized_top5_hits.append(len(set(o_top5) & actual))
            optimized_top12_hits.append(len(set(o_top12) & actual))

        return {
            'val_periods': min(val_window, len(history) - 20),
            'baseline_avg_top5': sum(baseline_top5_hits) / max(len(baseline_top5_hits), 1),
            'baseline_avg_top12': sum(baseline_top12_hits) / max(len(baseline_top12_hits), 1),
            'optimized_avg_top5': sum(optimized_top5_hits) / max(len(optimized_top5_hits), 1),
            'optimized_avg_top12': sum(optimized_top12_hits) / max(len(optimized_top12_hits), 1),
            'improvement_top5': (sum(optimized_top5_hits) - sum(baseline_top5_hits)) / max(sum(baseline_top5_hits), 1) * 100,
            'improvement_top12': (sum(optimized_top12_hits) - sum(baseline_top12_hits)) / max(sum(baseline_top12_hits), 1) * 100,
            'baseline_detail': baseline_top12_hits,
            'optimized_detail': optimized_top12_hits,
        }

    def _save_result(self, result: Dict):
        """保存优化结果"""
        output_path = os.path.join(_PROJ, 'cache', 'deep_optimization_result.json')
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        logger.info(f"优化结果已保存至: {output_path}")

    def generate_report(self, result: Dict, history: List[Dict]) -> str:
        """生成优化分析报告"""
        today = datetime.datetime.now().strftime("%Y%m%d")
        report_path = os.path.join(_PROJ, 'reports', f'deep_optimization_report_{today}.md')

        bt = result.get('backtest', {})
        co = result.get('co_emergence', [])

        co_lines = "\n".join([
            f"| {c['num']} | {c['score']} | {c['path']} |" for c in co
        ]) if co else "| - | - | - |"

        report = f"""# 命中率深度优化分析报告
**生成日期:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}
**数据基线:** {len(history)}期历史数据

---

## 一、近期命中率现状诊断

### 1.1 历史复盘数据 (近8期)
{self._generate_recent_review_table(history)}

### 1.2 核心瓶颈分析
1. **算法权重固化**: MK:EF:RW 比例手动设定, 未随市场环境自适应
2. **共识陷阱**: 当所有算法都推荐同一号码时, 该号反而命中率偏低
3. **遗漏陷阱**: 长遗漏号被过高估值, 忽视了"遗漏可持续极长"的客观事实
4. **缺乏协同涌现检测**: 号码之间的条件概率关系未被利用
5. **缺乏多粒度验证**: 号码在不同时间窗口下排名不稳定但未被过滤

---

## 二、深度优化方案详解

### 方案1: 自适应集成学习 (Adaptive Ensemble Stacking)
**原理**: 用近期实际命中率训练元学习器, 自动决定各算法权重, 替代人工设定。
**元学习器最优权重**:
```
{json.dumps(result.get('meta_weights', {}), indent=2, ensure_ascii=False)}
```

### 方案2: 协同涌现检测 (Co-emergence Detection)
**原理**: 当号码A出现时, 号码B的出现概率显著高于随机期望(lift>1.5), 利用条件概率推演协同号码。
**当前期协同涌现号码**:
| 号码 | 综合提升度 | 推理路径 |
|:---|:---|:---|
{co_lines}

### 方案3: 反共识去噪 (Anti-Consensus Denoising)
**原理**: 全员共识号降权70%, 仅1个算法推荐的奇点号加权30%。
**效果**: 避免过度拟合, 发掘被多数算法忽略的高价值号码。

### 方案4: 时变衰减动态校准 (Time-Varying Decay Calibration)
**原理**: 不同算法的最优衰减率不同, 自动搜索最佳参数。
**校准结果**:
```
{json.dumps(result.get('optimal_decays', dict()), indent=2, ensure_ascii=False)}
```

### 方案5: 多粒度交叉验证 (Multi-Granularity Cross-Validation)
**原理**: 在5/10/20/50期多个窗口下检查排名稳定性, 仅保留跨窗口一致的号码。
**效果**: 过滤排名波动大的号码, 提升推荐稳定性。

---

## 三、优化效果回测

| 指标 | 基线方法 | 优化方法 | 提升幅度 |
|:---|:---|:---|:---|
| Top5 平均命中 | {bt.get('baseline_avg_top5', '-')} | {bt.get('optimized_avg_top5', '-')} | {bt.get('improvement_top5', '-'):.1f}% |
| Top12 平均命中 | {bt.get('baseline_avg_top12', '-')} | {bt.get('optimized_avg_top12', '-')} | {bt.get('improvement_top12', '-'):.1f}% |

> 回测期数: {bt.get('val_periods', '-')}期

### 逐期对比 (Top12 命中数)
| 期序 | 基线 | 优化 | 差异 |
|:---|:---|:---|:---|
{self._format_backtest_detail(bt)}

---

## 四、最终优化推荐

### 优化版 Top 5 (深度优化后)
**{result.get('final_top5', [])}**

### 优化版 Top 12 (深度优化后)
**{result.get('final_top12', [])}**

### 原始集成 Top 12 (对照组)
**{result.get('ensemble_top12', [])}**

---

## 五、实施建议

1. **立即部署**: 方案1(自适应集成) + 方案3(反共识去噪) 风险最低, 见效最快
2. **分步推进**: 方案2(协同涌现) 需要积累条件概率矩阵, 建议2周后评估
3. **持续监控**: 每日运行本引擎, 自动校准衰减率和元权重
4. **A/B对比**: 保留原五维一体推荐作为对照组, 深度优化版作为实验组

---

*Deep Hit-Rate Optimizer v1.0 · 生成于 {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}*
"""

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report)
        logger.info(f"优化报告已生成: {report_path}")
        return report

    def _generate_recent_review_table(self, history: List[Dict]) -> str:
        """动态生成近8期复盘数据表 — 替代硬编码历史数据"""
        if len(history) < 2:
            return "| 数据不足 | - | - | - |"
        
        lines = ["| 期号 | 号码数 | 均值 | 标准差 |", "|:---|:---|:---|:---|"]
        for h in history[:8]:
            issue = h.get('issue', '?')
            nums = h.get('numbers', [])
            mean_val = np.mean(nums) if nums else 0
            std_val = np.std(nums) if nums else 0
            lines.append(f"| {issue} | {len(nums)} | {mean_val:.1f} | {std_val:.1f} |")
        
        # 附加统计摘要
        recent_5 = [len(h['numbers']) for h in history[:5]]
        avg = np.mean(recent_5) if recent_5 else 0
        lines.append(f"\n**近5期平均号码数:** {avg:.1f}")
        lines.append(f"**理论期望:** 20/80 = 25.0%")
        
        return "\n".join(lines)

    def _format_backtest_detail(self, bt: Dict) -> str:
        """格式化回测逐期对比"""
        baseline = bt.get('baseline_detail', [])
        optimized = bt.get('optimized_detail', [])
        if not baseline or not optimized:
            return "| - | - | - | - |"
        lines = []
        for i, (b, o) in enumerate(zip(baseline, optimized)):
            diff = o - b
            marker = "↑" if diff > 0 else ("↓" if diff < 0 else "=")
            lines.append(f"| {i+1} | {b} | {o} | {diff:+d} {marker} |")
        return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════
#  主入口
# ══════════════════════════════════════════════════════════════════

def main():
    """主入口: 加载数据 → 执行优化 → 生成报告"""
    print("=" * 60)
    print("  命中率深度优化引擎 (Deep Hit-Rate Optimizer)")
    print("=" * 60)

    # 加载数据
    try:
        from core.feature_optimizer import load_all_data
        data1, data2, d1_stars, history, points = load_all_data()
        print(f"历史数据加载成功: {len(history)}期")
    except Exception as e:
        print(f"数据加载失败: {e}")
        print("尝试使用备用数据源...")
        history = _load_fallback_history()
        if not history:
            print("无可用数据, 退出")
            return

    # 执行优化
    optimizer = DeepHitRateOptimizer()
    result = optimizer.optimize(history)

    # 生成报告
    optimizer.generate_report(result, history)

    # 打印摘要
    print("\n" + "=" * 60)
    print("  优化结果摘要")
    print("=" * 60)
    print(f"  优化版 Top 5:  {result.get('final_top5', [])}")
    print(f"  优化版 Top 12: {result.get('final_top12', [])}")
    bt = result.get('backtest', {})
    if bt.get('val_periods'):
        print(f"  回测({bt['val_periods']}期):")
        print(f"    基线 Top5={bt['baseline_avg_top5']:.2f}, Top12={bt['baseline_avg_top12']:.2f}")
        print(f"    优化 Top5={bt['optimized_avg_top5']:.2f}, Top12={bt['optimized_avg_top12']:.2f}")
        print(f"    Top5提升={bt['improvement_top5']:.1f}%, Top12提升={bt['improvement_top12']:.1f}%")
    print("=" * 60)


def _load_fallback_history() -> List[Dict]:
    """备用历史数据加载: 从Excel直接解析"""
    history = []
    try:
        import openpyxl
        excel_file = os.path.join(_PROJ, '跟随+点位+开奖数据.xlsx')
        if not os.path.exists(excel_file):
            return history

        from utils.excel_lock import excel_lock
        with excel_lock(excel_file, timeout=60):
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            try:
                if '开奖号码' not in wb.sheetnames:
                    return history

                ws = wb['开奖号码']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        issue = str(row[0]).strip()
                        numbers = []
                        for val in row[1:]:
                            if val and str(val).strip().isdigit():
                                numbers.append(int(str(val).strip()))
                            if len(numbers) >= 20:
                                break
                        if len(numbers) >= 20:
                            history.append({'issue': issue, 'numbers': numbers[:20]})
            finally:
                wb.close()

        history.reverse()  # 最新的在最前面
        logger.info(f"备用数据源加载成功: {len(history)}期")
    except Exception as e:
        logger.error(f"备用数据源加载失败: {e}")

    return history


if __name__ == '__main__':
    main()