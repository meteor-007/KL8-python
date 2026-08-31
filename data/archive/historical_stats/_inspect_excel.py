#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""检查跟随+点位+开奖数据.xlsx的结构和现有数据"""
import openpyxl
import re
import os

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '跟随+点位+开奖数据.xlsx')

if not os.path.exists(EXCEL_FILE):
    print(f"FILE NOT FOUND: {EXCEL_FILE}")
    exit(1)

print(f"File: {EXCEL_FILE}")
print(f"Size: {os.path.getsize(EXCEL_FILE) / 1024:.1f} KB")

wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
print(f"\nSheets: {wb.sheetnames}")

# 检查跟随号码统计 Sheet
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n--- Sheet: {sname} ---")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    if sname == '跟随号码统计':
        # 收集所有期号标记
        markers = []
        for r in range(1, ws.max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "")
            m = re.search(r'(\d+)期数据(1|2)', v)
            if m:
                markers.append((r, m.group(1), m.group(2)))
        
        print(f"  期号标记数: {len(markers)}")
        issues = sorted(set(iss for _, iss, _ in markers))
        print(f"  期号数: {len(issues)}")
        print(f"  期号范围: {issues[0]} ~ {issues[-1]}" if issues else "  无期号数据")
        
        # 显示前5个期号
        print(f"\n  前5个期号:")
        for iss in issues[:5]:
            data1_row = next((r for r, i, d in markers if i == iss and d == '1'), None)
            data2_row = next((r for r, i, d in markers if i == iss and d == '2'), None)
            # 读几行数据1
            print(f"    {iss}期: data1@row{data1_row}, data2@row{data2_row}")
            if data1_row:
                for dr in range(data1_row+1, min(data1_row+5, data2_row or data1_row+20)):
                    vals = [str(ws.cell(row=dr, column=c).value or '') for c in range(1, 10)]
                    if any(v for v in vals):
                        print(f"      Row{dr}: {vals}")
        
        # 显示最后3个期号
        print(f"\n  最后3个期号:")
        for iss in issues[-3:]:
            data1_row = next((r for r, i, d in markers if i == iss and d == '1'), None)
            data2_row = next((r for r, i, d in markers if i == iss and d == '2'), None)
            print(f"    {iss}期: data1@row{data1_row}, data2@row{data2_row}")
            if data1_row:
                for dr in range(data1_row+1, min(data1_row+5, data2_row or data1_row+20)):
                    vals = [str(ws.cell(row=dr, column=c).value or '') for c in range(1, 10)]
                    if any(v for v in vals):
                        print(f"      Row{dr}: {vals}")
    else:
        # 其他Sheet - 只显示基本信息
        for r in range(1, min(5, ws.max_row + 1)):
            vals = [str(ws.cell(row=r, column=c).value or '')[:15] for c in range(1, min(6, ws.max_column + 1))]
            print(f"  Row{r}: {vals}")

wb.close()
