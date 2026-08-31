# -*- coding: utf-8 -*-
"""跟随号码统计 — v4: 唯一号码层面分析 + 特征独立性 + 改进模型
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)

records = []
all_rows = list(ws.iter_rows())
BLOCK_OFFSETS = [1, 6, 11, 16]
for r_idx, row in enumerate(all_rows):
    first_val = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
    if not m: continue
    issue = int(m.group(1)); dtype = int(m.group(2))
    for b_idx, offset in enumerate(BLOCK_OFFSETS):
        for row_off in range(4):
            ri = r_idx + offset + row_off
            if ri >= len(all_rows): continue
            trow = all_rows[ri]
            for col_idx in list(range(4)) + list(range(5, 9)):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                side = 'L' if col_idx < 4 else 'R'
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    block=b_idx, side=side, row=row_off,
                    col=col_idx if side=='L' else col_idx-5))

# ── 唯一号码级 记录: 每期每号码聚合 (星号池) ──
def unique_pool(iss):
    """每期唯一星号号码 -> 属性集"""
    d = {}
    for r in records:
        if r['issue'] != iss or not r['star']: continue
        if r['num'] not in d:
            d[r['num']] = dict(blocks=set(), sides=set(), points=0, dtypes=set())
        d[r['num']]['blocks'].add(r['block'])
        d[r['num']]['sides'].add(r['side'])
        d[r['num']]['points'] = max(d[r['num']]['points'], int(r['point']))
        d[r['num']]['dtypes'].add(r['dtype'])
    return d

issues = sorted(set(r['issue'] for r in records if r['issue'] in draws))
print(f'期数: {len(issues)}')

# 每期唯一池大小 & 命中
sizes = []; hits = []
for iss in issues:
    pool = unique_pool(iss)
    sizes.append(len(pool))
    hits.append(len(set(pool) & draws[iss]))
print(f'唯一星池: 期均 {statistics.mean(sizes):.1f} 码, 期均命中 {statistics.mean(hits):.2f}')

# ═══ 唯一号级 特征命中率 (整体, 无前视暂用全样本看结构) ═══
print('\n═══ 唯一号级 特征命中率 (全样本) ═══')
feat = defaultdict(lambda: [0,0])
for iss in issues:
    pool = unique_pool(iss)
    draw = draws[iss]
    for n, attrs in pool.items():
        hit = 1 if n in draw else 0
        z = (n-1)//10; t = n % 10
        feat[f"Z{z*10+1:02d}-{z*10+10:02d}"][1]+=1; feat[f"Z{z*10+1:02d}-{z*10+10:02d}"][0]+=hit
        feat[f"尾{t}"][1]+=1; feat[f"尾{t}"][0]+=hit
        feat[f"点{attrs['points']}"][1]+=1; feat[f"点{attrs['points']}"][0]+=hit
        side = 'R' if 'R' in attrs['sides'] else 'L'
        feat[f"侧{side}"][1]+=1; feat[f"侧{side}"][0]+=hit
        if iss-1 in draws:
            rep = 1 if n in draws[iss-1] else 0
            feat[f"重{rep}"][1]+=1; feat[f"重{rep}"][0]+=hit
        # 是否同时出现在数据1和数据2
        both = 1 if len(attrs['dtypes'])==2 else 0
        feat[f"双源{both}"][1]+=1; feat[f"双源{both}"][0]+=hit

overall = sum(v[0] for v in feat.values())/sum(v[1] for v in feat.values())
print(f'唯一池整体命中率: {overall:.4f}')
for key in ['侧L','侧R','点0','点1','双源0','双源1','重0','重1']:
    h, n = feat[key]
    print(f'  {key:6s} n={n:5d} 率={h/n:.4f} z={((h/n-overall)/math.sqrt(overall*(1-overall)/n)):+.2f}')
print('  分区:')
for z in range(8):
    h, n = feat[f"Z{z*10+1:02d}-{z*10+10:02d}"]
    print(f'    {z*10+1:02d}-{z*10+10:02d}: n={n:4d} 率={h/n:.4f} z={((h/n-overall)/math.sqrt(overall*(1-overall)/n)):+.2f}')
print('  尾数:')
for t in range(10):
    h, n = feat[f"尾{t}"]
    print(f'    {t}尾: n={n:4d} 率={h/n:.4f} z={((h/n-overall)/math.sqrt(overall*(1-overall)/n)):+.2f}')

# ═══ 特征独立性: 尾2 与 重号 是否独立 ═══
print('\n═══ 特征独立性交叉 ═══')
cross = defaultdict(lambda: [0,0])
for iss in issues:
    pool = unique_pool(iss)
    draw = draws[iss]
    for n, attrs in pool.items():
        if iss-1 not in draws: continue
        hit = 1 if n in draw else 0
        t2 = (n%10==2); rep = n in draws[iss-1]
        key = f"尾2={int(t2)},重={int(rep)}"
        cross[key][1]+=1; cross[key][0]+=hit
for k in sorted(cross):
    h, n = cross[k]
    print(f'  {k:12s} n={n:4d} 率={h/n:.4f}')

# ═══ 改进模型: 唯一号级 无前视 logistic 打分 ═══
print('\n═══ 唯一号级 无前视打分选择 ═══')
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]

def logit(p):
    p = max(1e-6, min(1-1e-6, p))
    return math.log(p/(1-p))

def train_model(train_issues):
    """返回特征权重 (logit) 与基线"""
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        pool = unique_pool(iss)
        for n, attrs in pool.items():
            hit = 1 if n in draws[iss] else 0
            z = (n-1)//10; t = n%10
            side = 'R' if 'R' in attrs['sides'] else 'L'
            fc[f"Z{z}"][1]+=1; fc[f"Z{z}"][0]+=hit
            fc[f"尾{t}"][1]+=1; fc[f"尾{t}"][0]+=hit
            fc[f"侧{side}"][1]+=1; fc[f"侧{side}"][0]+=hit
            fc[f"点{attrs['points']}"][1]+=1; fc[f"点{attrs['points']}"][0]+=hit
            if iss-1 in draws:
                rep = 1 if n in draws[iss-1] else 0
                fc[f"重{rep}"][1]+=1; fc[f"重{rep}"][0]+=hit
    base = sum(v[0] for v in fc.values())/sum(v[1] for v in fc.values())
    weights = {}
    for k, (h, n) in fc.items():
        if n >= 25:
            weights[k] = logit(h/n) - logit(base)
    return weights, base

def score_num(n, attrs, iss, weights, base):
    z = (n-1)//10; t = n%10
    side = 'R' if 'R' in attrs['sides'] else 'L'
    feats = [f"Z{z}", f"尾{t}", f"侧{side}", f"点{attrs['points']}"]
    if iss-1 in draws:
        feats.append(f"重{1 if n in draws[iss-1] else 0}")
    s = logit(base)
    for f in feats:
        s += weights.get(f, 0)
    return s

def run(top_k):
    per = []; rankq = defaultdict(list)
    for iss in test_issues:
        train_issues = [i for i in issues if i < iss]
        weights, base = train_model(train_issues)
        pool = unique_pool(iss)
        scored = sorted(pool.items(), key=lambda kv: score_num(kv[0], kv[1], iss, weights, base), reverse=True)
        picks = [n for n, _ in scored[:top_k]]
        per.append(len(set(picks) & draws[iss]))
        n = len(scored)
        for qi, (n2, _) in enumerate(scored):
            rankq[int(qi/max(1,n-1)*5)].append(1 if n2 in draws[iss] else 0)
    return statistics.mean(per), rankq

print(f"{'TopK':>5s} {'期均命中':>8s} {'单码率':>8s} {'随机基线':>8s}")
for k in [5, 8, 12, 16]:
    per_m, rankq = run(k)
    import random
    rng = random.Random(7)
    rnd = []
    for iss in test_issues:
        pool = unique_pool(iss)
        picks = rng.sample(list(pool.keys()), min(k, len(pool)))
        rnd.append(len(set(picks) & draws[iss]))
    print(f'{k:5d} {per_m:8.2f} {per_m/k:8.4f} {statistics.mean(rnd):8.2f}')

# 排名分位
_, rankq = run(8)
print('\n排名分位命中率 (无前视, Top8模型):')
for b in range(5):
    v = rankq[b]
    if v: print(f'  分位{b}: n={len(v)} 率={statistics.mean(v):.4f}')