# -*- coding: utf-8 -*-
"""跟随号码统计 — 无前视行走验证 v1
用 ≤N-1 期的数据训练, 预测 N 期, 评估各信号的真实预测力
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"; BORDER_CLR = "FFD966B3"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

# ── 加载记录 ──
records = []
all_rows = list(ws.iter_rows())
BLOCK_OFFSETS = [1, 6, 11, 16]
for r_idx, row in enumerate(all_rows):
    first_val = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
    if not m: continue
    issue = int(m.group(1)); dtype = int(m.group(2))
    for b_idx, offset in enumerate(BLOCK_OFFSETS):
        for row_off in range(4):
            ri = r_idx + offset + row_off
            if ri >= len(all_rows): continue
            trow = all_rows[ri]
            for col_idx in list(range(4)) + list(range(5, 9)):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                side = 'L' if col_idx < 4 else 'R'
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    block=b_idx, side=side, row=row_off, col=col_idx if side=='L' else col_idx-5))

# ── 开奖 ──
draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(set(r['issue'] for r in records if r['issue'] in draws))
print(f'可用期数(有开奖): {len(issues)}')

# 每期的推荐池
def pool(issue, src='star'):
    """period 推荐池: 数据1星号 ∪ 数据2星号"""
    s = set()
    for r in records:
        if r['issue'] == issue and r['star']:
            s.add(r['num'])
    return s

# 每期各大信号的号码集
def signal_numbers(issue):
    d1 = set(r['num'] for r in records if r['issue']==issue and r['dtype']==1 and r['star'])
    d2s = set(r['num'] for r in records if r['issue']==issue and r['dtype']==2 and r['star'])
    d2all = set(r['num'] for r in records if r['issue']==issue and r['dtype']==2)
    b1r = set(r['num'] for r in records if r['issue']==issue and r['dtype']==1 and r['block']==1 and r['side']=='R')
    b2r = set(r['num'] for r in records if r['issue']==issue and r['dtype']==1 and r['block']==2 and r['side']=='R')
    pt = set(r['num'] for r in records if r['issue']==issue and r['point'])
    return d1, d2s, d2all, b1r, b2r, pt

def hit_rate(nums, draw):
    return len(nums & draw) / max(1, len(nums))

# ═══ 行走验证: 各信号池的期均命中 ═══
TEST_START = issues[30]  # 前30期训练
print(f'\n══════ 行走验证 (测试期 {TEST_START}~{issues[-1]}, 共{len(issues)-30}期) ══════')
test_issues = [i for i in issues if i >= TEST_START]
summary = defaultdict(list)
for iss in test_issues:
    d1, d2s, d2all, b1r, b2r, pt = signal_numbers(iss)
    draw = draws[iss]
    summary['数据1星'].append(len(d1 & draw))
    summary['数据2星'].append(len(d2s & draw))
    summary['数据1∪数据2星'].append(len((d1|d2s) & draw))
    summary['数据2全池'].append(len(d2all & draw))
    summary['B1右侧'].append(len(b1r & draw))
    summary['B2右侧'].append(len(b2r & draw))
    summary['点位'].append(len(pt & draw))

print(f"{'信号':14s} {'期均命中':>8s} {'池均大小':>8s} {'命中率':>8s}")
for k in ['数据1星','数据2星','数据1∪数据2星','数据2全池','B1右侧','B2右侧','点位']:
    v = summary[k]
    pools = []
    for iss in test_issues:
        d1, d2s, d2all, b1r, b2r, pt = signal_numbers(iss)
        pools.append(len({'数据1星':d1,'数据2星':d2s,'数据1∪数据2星':d1|d2s,'数据2全池':d2all,
                         'B1右侧':b1r,'B2右侧':b2r,'点位':pt}[k]))
    print(f"{k:14s} {statistics.mean(v):8.2f} {statistics.mean(pools):8.1f} "
          f"{statistics.mean(v)/statistics.mean(pools):8.4f}")

# ═══ 核心: 单码历史命中率打分 (无前视) 排行选择 ═══
print('\n══════ 核心: 单码历史命中率打分选择 ══════')
# 对每个测试期: 用历史(≤N-1)算每个号码"被推荐时命中率", 从当前推荐池挑Top-K
def run_selector(top_k, pool_src, min_hist=20, add_prior=True):
    """pool_src: 'd1star' | 'd1d2star'"""
    total_hits = 0; total_pick = 0; per_period = []
    for iss in test_issues:
        # 训练数据: 所有 < iss 的记录
        hist = defaultdict(lambda: [0, 0])  # num -> [hits, recs]
        for r in records:
            if r['issue'] < iss and r['star']:
                s = hist[r['num']]
                s[1] += 1
                s[0] += 1 if r['num'] in draws[r['issue']] else 0
        # 当前推荐池
        if pool_src == 'd1star':
            cand = set(r['num'] for r in records if r['issue']==iss and r['dtype']==1 and r['star'])
        else:
            cand = set(r['num'] for r in records if r['issue']==iss and r['star'])
        # 打分: 经验命中率 + 先验(全局0.25)平滑
        def score(n):
            h, t = hist.get(n, (0, 0))
            if t < min_hist: return 0.25  # 样本不足, 用先验
            if add_prior:
                return (h + 5*0.25) / (t + 5)
            return h / t
        scored = sorted(cand, key=score, reverse=True)
        picks = scored[:top_k]
        draw = draws[iss]
        hits = len(set(picks) & draw)
        total_hits += hits; total_pick += len(picks); per_period.append(hits)
    return total_hits/total_pick, statistics.mean(per_period), per_period

print(f"{'配置':40s} {'单码命中率':>10s} {'期均命中':>8s}")
for k in [5, 8, 12, 16]:
    for src in ['d1star', 'd1d2star']:
        if src == 'd1star' and k > 16: continue
        rate, per_p, _ = run_selector(k, src)
        print(f"Top{k:2d} | {src:10s} 历史命中率打分: {rate:10.4f} {per_p:8.2f}")

# 随机基线: 从池中随机选 Top-k
print('\n随机基线 (从数据1星池随机选):')
for k in [5, 8, 12]:
    import random
    rng = random.Random(42)
    total = 0; per = []
    for iss in test_issues:
        cand = list(set(r['num'] for r in records if r['issue']==iss and r['dtype']==1 and r['star']))
        picks = rng.sample(cand, min(k, len(cand)))
        per.append(len(set(picks) & draws[iss]))
    print(f'  Top{k}: 期均 {statistics.mean(per):.2f}')

# ═══ 杀号信号验证: 历史低命中号码是否应回避 ═══
print('\n══════ 杀号验证: 历史低命中号码 vs 全池 ══════')
for iss in test_issues:
    pass
low_hits = []; high_hits = []
for iss in test_issues:
    hist = defaultdict(lambda: [0, 0])
    for r in records:
        if r['issue'] < iss and r['star']:
            s = hist[r['num']]; s[1]+=1; s[0]+= 1 if r['num'] in draws[r['issue']] else 0
    cand = set(r['num'] for r in records if r['issue']==iss and r['star'])
    draw = draws[iss]
    low = [n for n in cand if hist[n][1] >= 20 and hist[n][0]/hist[n][1] < 0.20]
    high = [n for n in cand if hist[n][1] >= 20 and hist[n][0]/hist[n][1] > 0.30]
    low_hits.append(len(set(low) & draw))
    high_hits.append(len(set(high) & draw))
print(f'回避池(历史<0.20, ≥20次): 期均命中 {statistics.mean(low_hits):.2f}')
print(f'精选池(历史>0.30, ≥20次): 期均命中 {statistics.mean(high_hits):.2f}')