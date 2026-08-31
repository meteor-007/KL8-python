# -*- coding: utf-8 -*-
"""跟随号码统计 — 无前视行走验证 v2: 特征级预测力 + 组合模型
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

records = []
all_rows = list(ws.iter_rows())
BLOCK_OFFSETS = [1, 6, 11, 16]
for r_idx, row in enumerate(all_rows):
    first_val = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
    if not m: continue
    issue = int(m.group(1)); dtype = int(m.group(2))
    for b_idx, offset in enumerate(BLOCK_OFFSETS):
        for row_off in range(4):
            ri = r_idx + offset + row_off
            if ri >= len(all_rows): continue
            trow = all_rows[ri]
            for col_idx in list(range(4)) + list(range(5, 9)):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                side = 'L' if col_idx < 4 else 'R'
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    block=b_idx, side=side, row=row_off,
                    col=col_idx if side=='L' else col_idx-5))

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(set(r['issue'] for r in records if r['issue'] in draws))
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]

def star_recs(iss):
    return [r for r in records if r['issue']==iss and r['star']]

# ═══ 特征级无前视验证: 每特征分箱 → 命中率 ═══
print(f'═══ 特征级无前视验证 (训练 < 测试期, 测试 {TEST_START}~{issues[-1]}) ═══')
# 收集所有 (issue, feature, hit) 三元组
feat_log = defaultdict(list)  # feat_key -> [0/1 hits]
base_log = []
for iss in test_issues:
    for r in star_recs(iss):
        hit = 1 if r['num'] in draws[iss] else 0
        base_log.append(hit)
        feat_log[f"B{r['block']}X{r['side']}"].append(hit)
        z = (r['num']-1)//10
        feat_log[f"Z{z*10+1:02d}-{z*10+10:02d}"].append(hit)
        feat_log[f"尾{r['num']%10}"].append(hit)
        feat_log[f"点{int(r['point'])}"].append(hit)
        # 上期是否开出 (重号效应)
        if iss-1 in draws:
            feat_log[f"上期出{int(r['num'] in draws[iss-1])}"].append(hit)

overall = statistics.mean(base_log)
print(f'整体星号命中率: {overall:.4f} (n={len(base_log)})')
def show(feats, title):
    print(f'\n── {title} ──')
    for f in feats:
        v = feat_log[f]
        if not v: continue
        n = len(v); rate = statistics.mean(v)
        diff = rate - overall
        # 二项检验
        se = math.sqrt(overall*(1-overall)/n)
        z = diff / se if se else 0
        flag = '***' if abs(z) > 3 else ('**' if abs(z) > 2 else ('*' if abs(z) > 1.5 else ''))
        print(f'  {f:8s} n={n:5d} 率={rate:.4f} Δ={diff:+.4f} z={z:+.2f}{flag}')

show([f"B{b}X{s}" for b in range(4) for s in 'LR'], 'Block×Side')
show([f"Z{z*10+1:02d}-{z*10+10:02d}" for z in range(8)], '分区')
show([f"尾{t}" for t in range(10)], '尾数')
show(['点0','点1'], '点位标记')
show(['上期出0','上期出1'], '上期开出效应')

# ═══ 组合打分模型: 无前视 ═══
print('\n══════ 组合打分模型 (无前视) ══════')
# 特征权重: 用训练期的经验命中率 (每个特征独立)
def train_feature_prior(train_issues):
    """返回 feature->rate 映射 (训练期)"""
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        for r in star_recs(iss):
            s = fc[f"B{r['block']}X{r['side']}"]
            s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            z = (r['num']-1)//10
            s = fc[f"Z{z*10+1:02d}-{z*10+10:02d}"]
            s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            s = fc[f"尾{r['num']%10}"]
            s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            s = fc[f"点{int(r['point'])}"]
            s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
    return {k: v[0]/v[1] for k, v in fc.items() if v[1] >= 20}

def score_num(r, priors, overall_rate):
    """号码x在当前期的组合分 = 各特征席位率的均值(偏离基线的加权)"""
    feats = [f"B{r['block']}X{r['side']}", f"Z{(r['num']-1)//10*10+1:02d}-{(r['num']-1)//10*10+10:02d}",
             f"尾{r['num']%10}", f"点{int(r['point'])}"]
    vals = [priors.get(f, overall_rate) for f in feats]
    return statistics.mean(vals)

def run_combo(top_k, mode):
    per = []
    for iss in test_issues:
        train_issues = [i for i in issues if i < iss]
        priors = train_feature_prior(train_issues)
        cand = star_recs(iss)
        if mode == 'combo':
            scored = sorted(cand, key=lambda r: score_num(r, priors, 0.25), reverse=True)
            picks = [r['num'] for r in scored[:top_k]]
        elif mode == 'pos_only':
            scored = sorted(cand, key=lambda r: priors.get(f"B{r['block']}X{r['side']}", 0.25), reverse=True)
            picks = [r['num'] for r in scored[:top_k]]
        elif mode == 'anti':
            # 反推荐: 选最不可能中的 (验证杀号)
            scored = sorted(cand, key=lambda r: score_num(r, priors, 0.25))
            picks = [r['num'] for r in scored[:top_k]]
        per.append(len(set(picks) & draws[iss]))
    return statistics.mean(per)

print(f"{'模式':10s} {'Top5':>6s} {'Top8':>6s} {'Top12':>6s} {'Top16':>6s}")
for mode in ['combo', 'pos_only', 'random_est', 'anti']:
    if mode == 'random_est':
        # 随机基线估算: 池均5.13命中/20.3码
        row = []
        for k in [5,8,12,16]:
            row.append(f"{5.13*k/20.3:.2f}")
        print(f"{'随机基线':10s} {'  '.join(row)}")
        continue
    res = []
    for k in [5,8,12,16]:
        res.append(f"{run_combo(k, mode):.2f}")
    print(f"{mode:10s} {'  '.join(res)}")

# ═══ 杀号: 回避池 vs 精选池 归一期均 ═══
print('\n══════ 杀号池归一比较 (每期) ══════')
def pool_stats(threshold, direction):
    per = []
    for iss in test_issues:
        hist = defaultdict(lambda: [0,0])
        for r in records:
            if r['issue'] < iss and r['star']:
                s = hist[r['num']]; s[1]+=1; s[0]+= 1 if r['num'] in draws[r['issue']] else 0
        cand = star_recs(iss)
        if direction == 'low':
            sel = [r for r in cand if hist[r['num']][1] >= 20 and hist[r['num']][0]/hist[r['num']][1] <= threshold]
        else:
            sel = [r for r in cand if hist[r['num']][1] >= 20 and hist[r['num']][0]/hist[r['num']][1] >= threshold]
        if sel:
            hits = len(set(r['num'] for r in sel) & draws[iss])
            per.append(hits/len(sel))
    return statistics.mean(per)
print(f'回避池(历史率≤0.15): 单码命中率 {pool_stats(0.15,"low"):.4f}')
print(f'精选池(历史率≥0.35): 单码命中率 {pool_stats(0.35,"high"):.4f}')
print(f'全池基线: {0.25:.4f}')