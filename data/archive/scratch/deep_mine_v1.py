# -*- coding: utf-8 -*-
"""跟随号码统计 — 全方位深度挖掘 v1
维度: 标记力 / 位置 / 单码 / 时间 / 分区尾数 / 数据1vs2 / 开奖位置 / 无前视验证
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict, Counter
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"; BORDER_CLR = "FFD966B3"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']

# ── 加载全部记录 (带标记) ──
records = []  # dict(issue, dtype, num, star, point, win, block, side, row, col)
all_rows = list(ws.iter_rows())
BLOCK_OFFSETS = [1, 6, 11, 16]
for r_idx, row in enumerate(all_rows):
    first_val = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
    if not m: continue
    issue = int(m.group(1)); dtype = int(m.group(2))
    for b_idx, offset in enumerate(BLOCK_OFFSETS):
        for row_off in range(4):
            ri = r_idx + offset + row_off
            if ri >= len(all_rows): continue
            trow = all_rows[ri]
            for col_idx in range(4):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    win=any(getattr(cell.border, s).color.rgb == BORDER_CLR for s in ('left','right','top','bottom')
                            if getattr(cell.border, s) and getattr(cell.border, s).color),
                    block=b_idx, side='L', row=row_off, col=col_idx))
            for col_idx in range(5, 9):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    win=any(getattr(cell.border, s).color.rgb == BORDER_CLR for s in ('left','right','top','bottom')
                            if getattr(cell.border, s) and getattr(cell.border, s).color),
                    block=b_idx, side='R', row=row_off, col=col_idx-5))

print('records:', len(records))
issues_sorted = sorted(set(r['issue'] for r in records))
print('issues:', len(issues_sorted), issues_sorted[0], '-', issues_sorted[-1])

# 校验: 中奖标记 vs 全量开奖数据
wsd = wb['全量开奖数据']
draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
agree = 0; total = 0; mismatches = 0
for iss in issues_sorted:
    if iss not in draws: continue
    for rec in [r for r in records if r['issue'] == iss]:
        total += 1
        in_draw = rec['num'] in draws[iss]
        if rec['win'] == in_draw: agree += 1
        else: mismatches += 1
print(f'中奖标记 vs 实际开奖: 一致 {agree}/{total} = {agree/total:.4f}, 不一致 {mismatches}')

# 用实际开奖作为 ground truth (更可靠)
def hit(rec): return rec['num'] in draws.get(rec['issue'], set())

# ═══ 维度1: 标记力 ═══
print('\n══════ 维度1: 标记组合命中率 ══════')
grp = defaultdict(list)
for r in records:
    if r['issue'] in draws:
        key = f"d{r['dtype']}|星{int(r['star'])}|点{int(r['point'])}"
        grp[key].append(r)
for key in sorted(grp, key=lambda k: -statistics.mean(hit(x) for x in grp[k])):
    v = grp[key]
    n = len(v); h = sum(hit(x) for x in v)
    print(f'{key:12s} n={n:6d} 命中={h:5d} 率={h/n:.4f}  vs随机0.25')

# 数据2 非星号 (背景码) 的命中率
d2b = [r for r in records if r['dtype']==2 and not r['star'] and r['issue'] in draws]
print(f'\n数据2非星号(背景): n={len(d2b)} 命中率={sum(hit(x) for x in d2b)/len(d2b):.4f}')

# ═══ 维度2: 位置 (block×side×row×col) ═══
print('\n══════ 维度2: Block×Side 命中率 (据1) ══════')
pos = defaultdict(list)
for r in records:
    if r['dtype']==1 and r['issue'] in draws:
        pos[(r['block'], r['side'])].append(r)
for k in sorted(pos):
    v = pos[k]; h = sum(hit(x) for x in v)
    print(f'B{k[0]} {k[1]} n={len(v):4d} 率={h/len(v):.4f}')

print('\n── Block×Row (据1) ──')
pos2 = defaultdict(list)
for r in records:
    if r['dtype']==1 and r['issue'] in draws:
        pos2[(r['block'], r['row'])].append(r)
for k in sorted(pos2):
    v = pos2[k]; h = sum(hit(x) for x in v)
    print(f'B{k[0]}R{k[1]} n={len(v):4d} 率={h/len(v):.4f}')

# ═══ 维度3: 单码质量 ═══
print('\n══════ 维度3: 单码命中率 (据1星号 ≥60次推荐) ══════')
num_stats = defaultdict(lambda: [0,0])  # num -> [hits, total]
for r in records:
    if r['dtype']==1 and r['issue'] in draws:
        s = num_stats[r['num']]
        s[1] += 1; s[0] += 1 if hit(r) else 0
ranked = sorted(num_stats.items(), key=lambda kv: -kv[1][0]/kv[1][1])
print('Top 15:')
for num, (h, n) in ranked[:15]:
    print(f'  {num:02d}: {h}/{n} = {h/n:.4f}')
print('Bottom 15:')
for num, (h, n) in ranked[-15:]:
    print(f'  {num:02d}: {h}/{n} = {h/n:.4f}')

# ═══ 维度4: 时间效应 (滚动窗口) ═══
print('\n══════ 维度4: 星号命中率随窗口变化 ══════')
for dd in [1,2]:
    recs = [r for r in records if r['dtype']==dd and r['star'] and r['issue'] in draws]
    rates = []
    for iss in issues_sorted:
        if iss not in draws: continue
        sub = [r for r in recs if r['issue']==iss]
        if sub:
            rates.append((iss, sum(hit(x) for x in sub)/len(sub)))
    # 每10期滚动
    win = 10
    for i in range(0, len(rates)-win+1, win):
        chunk = rates[i:i+win]
        avg = sum(x[1] for x in chunk)/len(chunk)
        print(f'  数据{dd} 期{chunk[0][0]}~{chunk[-1][0]}: 命中率 {avg:.3f}')

# ═══ 维度5: 分区/尾数 ═══
print('\n══════ 维度5: 分区命中率 (据1星) ══════')
zone = defaultdict(list)
for r in records:
    if r['dtype']==1 and r['issue'] in draws:
        z = (r['num']-1)//10
        zone[z].append(r)
for z in sorted(zone):
    v = zone[z]; h = sum(hit(x) for x in v)
    print(f'  {z*10+1:02d}-{z*10+10:02d}: n={len(v):4d} 率={h/len(v):.4f}')

print('\n── 尾数命中率 (据1星) ──')
tail = defaultdict(list)
for r in records:
    if r['dtype']==1 and r['issue'] in draws:
        t = r['num'] % 10
        tail[t].append(r)
for t in sorted(tail):
    v = tail[t]; h = sum(hit(x) for x in v)
    print(f'  {t}尾: n={len(v):4d} 率={h/len(v):.4f}')

# ═══ 维度6: 数据1 vs 数据2 ═══
print('\n══════ 维度6: 数据1 vs 数据2 ══════')
for dd in [1,2]:
    recs = [r for r in records if r['dtype']==dd and r['star'] and r['issue'] in draws]
    uniq_per_period = defaultdict(set)
    for r in recs: uniq_per_period[r['issue']].add(r['num'])
    avg_pool = statistics.mean(len(v) for v in uniq_per_period.values())
    per_period_hits = []
    for iss, nums in uniq_per_period.items():
        per_period_hits.append(len(nums & draws[iss]))
    print(f'数据{dd}: 平均推荐池 {avg_pool:.1f}码, 期均命中 {statistics.mean(per_period_hits):.2f}, '
          f'命中率 {statistics.mean(per_period_hits)/avg_pool:.4f}')

# 数据2 全池覆盖
for dd in [2]:
    recs = [r for r in records if r['dtype']==dd and r['issue'] in draws]
    uniq_per_period = defaultdict(set)
    for r in recs: uniq_per_period[r['issue']].add(r['num'])
    per_period_hits = []
    for iss, nums in uniq_per_period.items():
        per_period_hits.append(len(nums & draws[iss]))
    avg_pool = statistics.mean(len(v) for v in uniq_per_period.values())
    print(f'数据2全池: 平均覆盖 {avg_pool:.1f}码, 期均覆盖开奖 {statistics.mean(per_period_hits):.2f}/20')

# ═══ 维度7: 点位 ═══
print('\n══════ 维度7: 点位（粉色填充）命中率 ══════')
pt = [r for r in records if r['point'] and r['issue'] in draws]
npt = [r for r in records if not r['point'] and r['issue'] in draws]
print(f'点位: n={len(pt)} 率={sum(hit(x) for x in pt)/len(pt):.4f}')
print(f'非点位: n={len(npt)} 率={sum(hit(x) for x in npt)/len(npt):.4f}')
# 点位×星号
pt_star = [r for r in pt if r['star']]
pt_nostar = [r for r in pt if not r['star']]
if pt_star: print(f'点位+星: n={len(pt_star)} 率={sum(hit(x) for x in pt_star)/len(pt_star):.4f}')
if pt_nostar: print(f'点位+非星: n={len(pt_nostar)} 率={sum(hit(x) for x in pt_nostar)/len(pt_nostar):.4f}')

# ═══ 维度8: 开奖位置分析 ═══
print('\n══════ 维度8: 星号号码在开奖序列中的位置 ══════')
wsd_rows = list(wsd.iter_rows(values_only=True))[1:]
ord_map = {}
for rr in wsd_rows:
    if rr[1]:
        ord_map[int(rr[1])] = [int(x) for x in rr[3:23] if x is not None]
cpos = Counter()
for r in records:
    if r['dtype']==1 and r['star'] and r['issue'] in ord_map:
        if r['num'] in ord_map[r['issue']]:
            cpos[ord_map[r['issue']].index(r['num'])] += 1
print('星号命中号码在开奖序列中的位置分布 (位置0-19):')
for i in range(20):
    print(f'  位置{i:2d}: {cpos[i]}')

print('\n完成')