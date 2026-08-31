# -*- coding: utf-8 -*-
"""跟随号码统计 — v3: 开奖均匀性 + 特征关联 + 周期内排名单调性
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

# ── 开奖均匀性 ──
draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if i >= 2026089 and i <= 2026209)
print(f'═══ 开奖均匀性检查 ({len(issues)}期 × 20码) ═══')
zone_cnt = defaultdict(int); tail_cnt = defaultdict(int); pos_cnt = defaultdict(int)
repeats = []
for iss in issues:
    nums = list(draws[iss])
    for n in nums:
        zone_cnt[(n-1)//10] += 1
        tail_cnt[n%10] += 1
    prev = draws.get(iss-1, set())
    repeats.append(len(set(nums) & prev))
print('分区开奖数:', {f'{z*10+1:02d}-{z*10+10:02d}': zone_cnt[z] for z in sorted(zone_cnt)})
print('期望(均匀):', len(issues)*20/8)
print('尾数开奖数:', {t: tail_cnt[t] for t in sorted(tail_cnt)})
print('期望(均匀):', len(issues)*20/10)
print('重复开奖(上期同号): 均值 {:.2f} /20 (随机期望 20*20/80=5.0)'.format(statistics.mean(repeats)))

# ── 记录加载 ──
records = []
all_rows = list(ws.iter_rows())
BLOCK_OFFSETS = [1, 6, 11, 16]
for r_idx, row in enumerate(all_rows):
    first_val = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[^\d]*(\d)', first_val)
    if not m: continue
    issue = int(m.group(1)); dtype = int(m.group(2))
    for b_idx, offset in enumerate(BLOCK_OFFSETS):
        for row_off in range(4):
            ri = r_idx + offset + row_off
            if ri >= len(all_rows): continue
            trow = all_rows[ri]
            for col_idx in list(range(4)) + list(range(5, 9)):
                cell = trow[col_idx]
                v = str(cell.value or "").strip().replace('*', '')
                if not v.isdigit(): continue
                num = int(v)
                if not (1 <= num <= 80): continue
                side = 'L' if col_idx < 4 else 'R'
                records.append(dict(issue=issue, dtype=dtype, num=num,
                    star='*' in str(cell.value or ""),
                    point=(cell.fill.fgColor.rgb == POINT_FILL if cell.fill and cell.fill.fgColor else False),
                    block=b_idx, side=side, row=row_off,
                    col=col_idx if side=='L' else col_idx-5))

# ── 特征关联: B1XR 号码的 分区/尾数 分布 vs 全池 ──
print('\n═══ 特征关联: B1XR 的组成 ═══')
def zone_tail_dist(recs, label):
    zc = defaultdict(int); tc = defaultdict(int); n = len(recs)
    for r in recs:
        zc[(r['num']-1)//10] += 1
        tc[r['num']%10] += 1
    zdist = {z*10+1: round(zc[z]/n*100,1) for z in sorted(zc)}
    tdist = {t: round(tc[t]/n*100,1) for t in sorted(tc)}
    print(f'{label} (n={n}):')
    print(f'  分区%: {zdist}')
    print(f'  尾数%: {tdist}')

all_star = [r for r in records if r['star']]
b1xr = [r for r in all_star if r['block']==1 and r['side']=='R']
zone_tail_dist(all_star, '全星池')
zone_tail_dist(b1xr, 'B1XR池')

# ── 周期内排名单调性: 用无前视打分, 看各排名段命中率 ──
print('\n═══ 周期内排名单调性 (无前视打分) ═══')
issues_all = sorted(set(r['issue'] for r in records if r['issue'] in draws))
TEST_START = issues_all[30]
test_issues = [i for i in issues_all if i >= TEST_START]

def star_recs(iss): return [r for r in records if r['issue']==iss and r['star']]

# 训练特征先验
def train_priors(train_issues):
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        for r in star_recs(iss):
            fc[f"B{int(r['point'])}"][1] += 1  # 占位
            s = fc[f"B{r['block']}X{r['side']}"]; s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            z = (r['num']-1)//10
            s = fc[f"Z{z}"]; s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            s = fc[f"尾{r['num']%10}"]; s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            s = fc[f"点{int(r['point'])}"]; s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
            if iss-1 in draws:
                s = fc[f"重{int(r['num'] in draws[iss-1])}"]; s[1]+=1; s[0]+= 1 if r['num'] in draws[iss] else 0
    return {k: v[0]/v[1] for k, v in fc.items() if v[1] >= 15}

def score_num(r, priors, base):
    feats = [f"B{r['block']}X{r['side']}", f"Z{(r['num']-1)//10}", f"尾{r['num']%10}", f"点{int(r['point'])}"]
    if r['issue']-1 in draws:
        feats.append(f"重{int(r['num'] in draws[r['issue']-1])}")
    vals = [priors.get(f, base) for f in feats]
    return statistics.mean(vals)

# 收集所有被评分号码的 hit 按排名分位
rank_buckets = defaultdict(list)
for iss in test_issues:
    train_issues = [i for i in issues_all if i < iss]
    priors = train_priors(train_issues)
    cand = star_recs(iss)
    if len(cand) < 10: continue
    scored = sorted(cand, key=lambda r: score_num(r, priors, 0.25), reverse=True)
    n = len(scored)
    for rank, r in enumerate(scored):
        q = rank / max(1, n-1)  # 0=最高, 1=最低
        bucket = int(q * 5)  # 5分位
        rank_buckets[bucket].append(1 if r['num'] in draws[iss] else 0)

print(f"{'打分分位(0=最高)':18s} {'n':>6s} {'命中率':>10s}")
for b in range(5):
    v = rank_buckets[b]
    if v:
        print(f'  分位{b} ({b*20}-{b*20+19}%)     n={len(v):5d}  {statistics.mean(v):.4f}')

# ── 关键: 当期推荐独立性 → 用"训练集=t-1期"单特征排序对比 ──
print('\n═══ 上期开出挑号 (直接从池中剔除上期已出号码) ═══')
no_repeat = []; with_repeat = []
for iss in test_issues:
    if iss-1 not in draws: continue
    cand = star_recs(iss)
    prev = draws[iss-1]
    nr = [r['num'] for r in cand if r['num'] not in prev]
    wr = [r['num'] for r in cand if r['num'] in prev]
    no_repeat.append(len(set(nr) & draws[iss]))
    with_repeat.append(len(set(wr) & draws[iss]))
print(f'池1(剔除上期已出, 每期均{statistics.mean([len([r for r in star_recs(i) if r["num"] not in draws[i-1]]) for i in test_issues if i-1 in draws]):.1f}码): 期均命中 {statistics.mean(no_repeat):.2f}')
print(f'池2(仅上期已出, 每期均{statistics.mean([len([r for r in star_recs(i) if r["num"] in draws[i-1]]) for i in test_issues if i-1 in draws]):.1f}码): 期均命中 {statistics.mean(with_repeat):.2f}')
print(f'全池: 期均命中 {statistics.mean([len(set(r["num"] for r in star_recs(i)) & draws[i]) for i in test_issues]):.2f}')