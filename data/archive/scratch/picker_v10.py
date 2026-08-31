# -*- coding: utf-8 -*-
"""跟随号码统计 — v10: 终版规则选号器 (无前视回溯验证 + 当期推荐输出)
规则分层: R1尾2区0-3非重 > R2尾2非重 > R3右侧尾2/7/8/3 > R4尾2/7/8/3非重 > 基础池
负信号排除: 52/62/72(尾2区5-7), 重号, 41-50区, 61-70区, 尾0/1/4/6
"""
import openpyxl, re, sys, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if 2026089 <= i <= 2026209)

all_rows = list(ws.iter_rows())
period_cache = {}
for idx, row in enumerate(all_rows):
    v = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[\s\S]*?数据(1|2)', v)
    if not m: continue
    iss = int(m.group(1)); dtype = int(m.group(2))
    if iss not in period_cache:
        period_cache[iss] = {'d1': set(), 'd2': set()}
    pc = period_cache[iss]
    for b_idx, off in enumerate([1, 6, 11, 16]):
        for ro in range(4):
            if idx+off+ro >= len(all_rows): continue
            trow = all_rows[idx+off+ro]
            for c in range(min(10, len(trow))):
                cv = str(trow[c].value or "").strip()
                if cv and cv != 'nan' and '*' in cv:
                    n = int(cv.replace('*',''))
                    pc['d1' if dtype == 1 else 'd2'].add(n)

BASE_TAILS = (2, 7, 8, 3)

def tier(n, iss):
    """返回分层分数: 越高越优先"""
    repeat = iss-1 in draws and n in draws[iss-1]
    z = (n-1) // 10; t = n % 10
    if repeat: return 0                    # 重号: 直接垫底
    if 41 <= n <= 50 or 61 <= n <= 70: return 1   # 弱区: 最低非0层
    if t == 2 and z in (0, 1, 2, 3): return 6     # R1 尾2区0-3+非重
    if t == 2: return 5                           # R2 尾2+非重 (含52/62/72? 不, 见下)
    if t in (7, 8, 3): return 4                   # R4 尾7/8/3+非重
    return 2                                       # 基础池

def tier_side(n, iss, side):
    """考虑右侧加权的分层"""
    repeat = iss-1 in draws and n in draws[iss-1]
    z = (n-1) // 10; t = n % 10
    if repeat: return 0
    if 41 <= n <= 50 or 61 <= n <= 70: return 1
    if t == 2 and z in (0, 1, 2, 3): return 6
    if t == 2: return 3 if z in (5, 6, 7) else 5  # 尾2区5-7降权但保留
    if t in (7, 8, 3): return 4 + (1 if side == 'R' else 0)  # 右侧+尾7/8/3
    if side == 'R': return 3
    return 2

# ═══ 无前视回溯: 规则分层选号 Top-K ═══
print('═══ 规则分层选号器 无前视回溯 (数据1∪数据2星去重, 右侧重加权) ═══')
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]
best_side = {}
for iss in test_issues:
    # 每号取最高星序位置的侧
    for src in ['d1', 'd2']:
        pass
# 重新解析带side
period_side = {}
for idx, row in enumerate(all_rows):
    v = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[\s\S]*?数据(1|2)', v)
    if not m: continue
    iss = int(m.group(1)); dtype = int(m.group(2))
    if iss not in period_side:
        period_side[iss] = defaultdict(list)
    for b_idx, off in enumerate([1, 6, 11, 16]):
        for ro in range(4):
            if idx+off+ro >= len(all_rows): continue
            trow = all_rows[idx+off+ro]
            cells = []
            for c in range(min(10, len(trow))):
                cv = str(trow[c].value or "").strip()
                if cv and cv != 'nan' and '*' in cv:
                    cells.append((c, int(cv.replace('*',''))))
            cells.sort()
            for si, (c, n) in enumerate(cells):
                side = 'R' if c >= 5 else 'L'
                period_side[iss][n].append(side)

for use_side in [False, True]:
    for top_k in [8, 12]:
        per = []; pools = []
        for iss in test_issues:
            cand = {}
            alln = period_cache[iss]['d1'] | period_cache[iss]['d2']
            for n in alln:
                if use_side:
                    sides = period_side[iss][n]
                    sc = tier_side(n, iss, 'R' if 'R' in sides else 'L')
                else:
                    sc = tier(n, iss)
                cand[n] = sc
            sc = sorted(cand.items(), key=lambda kv: (-kv[1], kv[0]))
            picks = [n for n, _ in sc[:top_k]]
            per.append(len(set(picks) & draws[iss])); pools.append(len(alln))
        print(f'  {"右侧重" if use_side else "纯分层":4s} Top{top_k:2d}: 期均命中 {statistics.mean(per):.2f} (池均{statistics.mean(pools):.1f}码)')

# 随机对照
import random
rng = random.Random(7)
for top_k in [8, 12]:
    rnd = []
    for iss in test_issues:
        pool = period_cache[iss]['d1'] | period_cache[iss]['d2']
        picks = rng.sample(list(pool), min(top_k, len(pool)))
        rnd.append(len(set(picks) & draws[iss]))
    print(f'  随机对照 Top{top_k:2d}: 期均命中 {statistics.mean(rnd):.2f}')

# ═══ 当期推荐 (最后一期) ═══
print('\n═══ 最新期推荐池 ═══')
latest = issues[-1]
pool = sorted(period_cache[latest]['d1'] | period_cache[latest]['d2'])
print(f'期号: {latest}  池内{len(pool)}码 (去重)')
for n in pool:
    sides = period_side[latest][n]
    s = tier_side(n, latest, 'R' if 'R' in sides else 'L')
    tags = []
    t = n % 10; z = (n-1)//10
    if t == 2 and z in (0,1,2,3): tags.append('R1尾2区0-3')
    elif t == 2: tags.append('R2尾2')
    if t in (7,8,3): tags.append('尾7/8/3')
    if 'R' in sides: tags.append('右侧')
    if latest-1 in draws and n in draws[latest-1]: tags.append('重号!')
    if 41<=n<=50 or 61<=n<=70: tags.append('弱区!')
    print(f'  {n:02d} [层{s}] {"+".join(tags) if tags else "-"}')
print(f'\n建议: 优先选层6/5/4, 避开重号与弱区, 层0/1直接排除')