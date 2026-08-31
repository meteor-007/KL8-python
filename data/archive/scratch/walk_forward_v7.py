# -*- coding: utf-8 -*-
"""跟随号码统计 — v7: 终版组合模型 无前视验证
特征: 星号序号(反热) / 尾数 / 分区 / 重号 / B1XR / 点位唯一
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
POINT_FILL = "FFFCE4EC"
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if 2026089 <= i <= 2026209)

# 点位
points = {}
with open(r'D:\Dpanqianyi\Python-Project\data\daily_points.txt', encoding='utf-8') as f:
    for line in f:
        m = re.search(r'period:(\d+)', line)
        pm = re.search(r'points:([\d\s]+)', line)
        if m and pm:
            points[int(m.group(1))] = set(int(p) for p in pm.group(1).split() if p)

# 点位独特号 无前视验证
all_rows = list(ws.iter_rows())
print('═══ 点位独特号 无前视验证 ═══')
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]
d2all_cache = {}
def d2_all(iss):
    if iss not in d2all_cache:
        s = set()
        for idx, row in enumerate(all_rows):
            v = str(row[0].value or "").strip()
            m = re.search(r'(\d{7})期[\s\S]*?数据2', v)
            if not m or int(m.group(1)) != iss: continue
            for off in [1, 6, 11, 16]:
                for ro in range(4):
                    trow = all_rows[idx+off+ro]
                    for c in range(min(10, len(trow))):
                        cv = str(trow[c].value or "").strip()
                        if cv and cv != 'nan':
                            s.add(int(cv.replace('*','')))
            break
        d2all_cache[iss] = s
    return d2all_cache[iss]

uniq_pt = []; all_pt = []
for iss in test_issues:
    if iss not in points: continue
    pt = points[iss]
    d2 = d2_all(iss)
    uniq = pt - d2
    uniq_pt.append(len(uniq))
    # 命中
per_uniq = []
for iss in test_issues:
    if iss not in points: continue
    pt = points[iss]; d2 = d2_all(iss)
    uniq = pt - d2
    per_uniq.append(len(uniq & draws[iss]))
print(f'点位独特号: 期均 {statistics.mean(per_uniq):.2f} 命中 / 池均 {statistics.mean(uniq_pt):.1f} 码 = {statistics.mean(per_uniq)/statistics.mean(uniq_pt):.4f}')

# ═══ 终版组合模型: 从当期推荐池(数据1∪数据2星) 打分选Top-K ═══
print('\n═══ 终版组合模型 (无前视) ═══')
all_rows = list(ws.iter_rows())

def parse_period_records(iss, pool_src):
    """返回 [(num, star_ordinal, block, side, tail, zone, repeat_flags)]"""
    out = []
    for idx, row in enumerate(all_rows):
        v = str(row[0].value or "").strip()
        m = re.search(r'(\d{7})期[\s\S]*?数据(1|2)', v)
        if not m or int(m.group(1)) != iss: continue
        dtype = int(m.group(2))
        if pool_src == 'd1' and dtype != 1: continue
        if pool_src == 'd2star' and dtype != 2: continue
        for b_idx, off in enumerate([1, 6, 11, 16]):
            for ro in range(4):
                trow = all_rows[idx+off+ro]
                cells = []
                for c in range(min(10, len(trow))):
                    cv = str(trow[c].value or "").strip()
                    if cv and cv != 'nan' and '*' in cv:
                        cells.append((c, int(cv.replace('*',''))))
                cells.sort()
                for si, (c, n) in enumerate(cells):
                    side = 'L' if c < 4 else 'R'
                    out.append((n, si, b_idx, side))
    return out

def logit(p):
    p = max(1e-6, min(1-1e-6, p))
    return math.log(p/(1-p))

def build_weights(train_issues, pool_src):
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        recs = parse_period_records(iss, pool_src)
        for n, si, b, side in recs:
            hit = 1 if n in draws[iss] else 0
            fc[f"星{min(si,7)}"][1]+=1; fc[f"星{min(si,7)}"][0]+=hit
            fc[f"侧{side}"][1]+=1; fc[f"侧{side}"][0]+=hit
            fc[f"尾{n%10}"][1]+=1; fc[f"尾{n%10}"][0]+=hit
            fc[f"Z{(n-1)//10}"][1]+=1; fc[f"Z{(n-1)//10}"][0]+=hit
            if iss-1 in draws:
                fc[f"重{1 if n in draws[iss-1] else 0}"][1]+=1; fc[f"重{1 if n in draws[iss-1] else 0}"][0]+=hit
    base = sum(v[0] for v in fc.values())/sum(v[1] for v in fc.values())
    w = {}
    for k, (h, n) in fc.items():
        if n >= 30:
            w[k] = logit(h/n) - logit(base)
    return w, base

def score_rec(n, si, b, side, iss, w, base):
    s = logit(base)
    s += w.get(f"星{min(si,7)}", 0) + w.get(f"侧{side}", 0) + w.get(f"尾{n%10}", 0) + w.get(f"Z{(n-1)//10}", 0)
    if iss-1 in draws:
        s += w.get(f"重{1 if n in draws[iss-1] else 0}", 0)
    return s

import random
rng = random.Random(11)
for pool_src in ['d1', 'd2star', 'both']:
    print(f'\n── 池: {pool_src} ──')
    for top_k in [5, 8, 12]:
        per = []; rnd = []
        for iss in test_issues:
            train_issues = [i for i in issues if i < iss]
            w, base = build_weights(train_issues, pool_src)
            recs = parse_period_records(iss, pool_src)
            # 去重: 同号取最高分
            best = {}
            for n, si, b, side in recs:
                s = score_rec(n, si, b, side, iss, w, base)
                if n not in best or s > best[n]: best[n] = s
            sc = sorted(best.items(), key=lambda kv: -kv[1])
            picks = [n for n, _ in sc[:top_k]]
            per.append(len(set(picks) & draws[iss]))
            # 随机
            rp = rng.sample(list(best.keys()), min(top_k, len(best)))
            rnd.append(len(set(rp) & draws[iss]))
        print(f'  Top{top_k:2d}: 模型 {statistics.mean(per):.2f} | 随机 {statistics.mean(rnd):.2f}')

# ═══ 分区+尾数 直接组合规则 (简单可解释) ═══
print('\n═══ 简单可解释规则 (无前视) ═══')
# 规则1: 只选 尾2/7/8/3 且 非重号 lemurs
for rule_name, rule_fn in [
    ('尾2/7/8/3+非重', lambda n, si, b, side, iss: (n%10 in (2,7,8,3)) and (iss-1 not in draws or n not in draws[iss-1])),
    ('尾2/7/8/3', lambda n, si, b, side, iss: n%10 in (2,7,8,3)),
    ('非重号', lambda n, si, b, side, iss: (iss-1 not in draws or n not in draws[iss-1])),
    ('星≥5(右侧)', lambda n, si, b, side, iss: side == 'R'),
    ('星≥5+尾2/7/8/3', lambda n, si, b, side, iss: side=='R' and n%10 in (2,7,8,3)),
]:
    per = []
    for iss in test_issues:
        recs = parse_period_records(iss, 'd1')
        cand = set()
        for n, si, b, side in recs:
            if rule_fn(n, si, b, side, iss):
                cand.add(n)
        per.append(len(cand & draws[iss]) / max(1, len(cand)))
    print(f'  {rule_name}: 单码命中率 {statistics.mean(per):.4f} (基线0.25)')