# -*- coding: utf-8 -*-
"""跟随号码统计 — v8: 信号时效稳定性 + 尾2构成
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict, Counter
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if 2026089 <= i <= 2026209)

all_rows = list(ws.iter_rows())
period_cache = {}
for idx, row in enumerate(all_rows):
    v = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[\s\S]*?数据(1|2)', v)
    if not m: continue
    iss = int(m.group(1)); dtype = int(m.group(2))
    if iss not in period_cache:
        period_cache[iss] = {'d1': [], 'd2': []}
    pc = period_cache[iss]
    for b_idx, off in enumerate([1, 6, 11, 16]):
        for ro in range(4):
            if idx+off+ro >= len(all_rows): continue
            trow = all_rows[idx+off+ro]
            cells = []
            for c in range(min(10, len(trow))):
                cv = str(trow[c].value or "").strip()
                if cv and cv != 'nan' and '*' in cv:
                    cells.append((c, int(cv.replace('*',''))))
            cells.sort()
            for si, (c, n) in enumerate(cells):
                side = 'L' if c < 4 else 'R'
                rec = (n, si, b_idx, side)
                if dtype == 1: pc['d1'].append(rec)
                else: pc['d2'].append(rec)

# ═══ 1. 信号时效: 分3段看 尾2/尾0/右侧/重号 命中率 ═══
print('═══ 信号时效稳定性 (数据1星, 3段时间窗) ═══')
N = len(issues)
seg = [issues[:N//3], issues[N//3:2*N//3], issues[2*N//3:]]
for seg_name, seg_issues in zip(['前1/3', '中1/3', '后1/3'], seg):
    out = defaultdict(lambda: [0,0])
    for iss in seg_issues:
        for n, si, b, side in period_cache[iss]['d1']:
            hit = 1 if n in draws[iss] else 0
            out['尾2'][1]+=1; out['尾2'][0]+= hit if n%10==2 else 0
            out['尾0'][1]+=1; out['尾0'][0]+= hit if n%10==0 else 0
            out['右侧'][1]+=1; out['右侧'][0]+= hit if side=='R' else 0
            out['重号'][1]+=1; out['重号'][0]+= hit if (iss-1 in draws and n in draws[iss-1]) else 0
            out['非重'][1]+=1; out['非重'][0]+= hit if (iss-1 not in draws or n not in draws[iss-1]) else 0
    print(f'  {seg_name}: 尾2={out["尾2"][0]}/{out["尾2"][1]}={out["尾2"][0]/out["尾2"][1]:.3f} '
          f'| 尾0={out["尾0"][0]/out["尾0"][1]:.3f} | 右侧={out["右侧"][0]/out["右侧"][1]:.3f} '
          f'| 非重={out["非重"][0]/out["非重"][1]:.3f} | 重={out["重号"][0]/out["重号"][1]:.3f}')

# ═══ 2. 尾2号码构成 ═══
print('\n═══ 尾2号码 命中率明细 ═══')
t2 = defaultdict(lambda: [0,0])
for iss in issues:
    for n, si, b, side in period_cache[iss]['d1']:
        if n % 10 == 2:
            t2[n][1] += 1
            t2[n][0] += 1 if n in draws[iss] else 0
for n in sorted(t2):
    h, t = t2[n]
    print(f'  {n:02d}: {h}/{t} = {h/t:.3f}')

# ═══ 3. 61-70 区与其他区 时间稳定性 ═══
print('\n═══ 61-70区 命中率 (数据1星) 分段 ═══')
for seg_name, seg_issues in zip(['前1/3', '中1/3', '后1/3'], seg):
    h = t = 0
    for iss in seg_issues:
        for n, si, b, side in period_cache[iss]['d1']:
            if 61 <= n <= 70:
                t += 1
                h += 1 if n in draws[iss] else 0
    print(f'  {seg_name}: {h}/{t} = {h/t:.3f}')

# ═══ 4. 组合规则 在最近30期的表现 ═══
print('\n═══ 关键规则 最近30期单独验证 ═══')
recent = issues[-30:]
for name, fn in [
    ('尾2+非重', lambda n, si, side, iss: n%10==2 and (iss-1 not in draws or n not in draws[iss-1])),
    ('右侧+尾2/7/8/3', lambda n, si, side, iss: side=='R' and n%10 in (2,7,8,3)),
    ('尾2/7/8/3+非重', lambda n, si, side, iss: n%10 in (2,7,8,3) and (iss-1 not in draws or n not in draws[iss-1])),
]:
    hits = 0; tot = 0
    for iss in recent:
        cand = set()
        for n, si, b, side in period_cache[iss]['d1']:
            if fn(n, si, side, iss): cand.add(n)
        hits += len(cand & draws[iss]); tot += len(cand)
    print(f'  {name:14s}: {hits}/{tot} = {hits/tot:.4f} (基线0.25)')

# 全池基线
hits = 0; tot = 0
for iss in recent:
    cand = set(n for n, si, b, side in period_cache[iss]['d1'])
    hits += len(cand & draws[iss]); tot += len(cand)
print(f'  {"数据1全池":14s}: {hits}/{tot} = {hits/tot:.4f}')