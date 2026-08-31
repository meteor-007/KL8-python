# -*- coding: utf-8 -*-
"""跟随号码统计 — v9: 尾2×区交互 + 终版规则 + 选号器
"""
import openpyxl, re, sys, math, statistics
from collections import defaultdict
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

DATA_FILE = r'D:\Dpanqianyi\Python-Project\data\跟随+点位+开奖数据.xlsx'
wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
ws = wb['跟随号码统计']
wsd = wb['全量开奖数据']

draws = {}
for r in list(wsd.iter_rows(values_only=True))[1:]:
    if r[1]:
        draws[int(r[1])] = set(int(x) for x in r[3:23] if x is not None)
issues = sorted(i for i in draws if 2026089 <= i <= 2026209)

all_rows = list(ws.iter_rows())
period_cache = {}
for idx, row in enumerate(all_rows):
    v = str(row[0].value or "").strip()
    m = re.search(r'(\d{7})期[\s\S]*?数据(1|2)', v)
    if not m: continue
    iss = int(m.group(1)); dtype = int(m.group(2))
    if iss not in period_cache:
        period_cache[iss] = {'d1': [], 'd2': []}
    pc = period_cache[iss]
    for b_idx, off in enumerate([1, 6, 11, 16]):
        for ro in range(4):
            if idx+off+ro >= len(all_rows): continue
            trow = all_rows[idx+off+ro]
            cells = []
            for c in range(min(10, len(trow))):
                cv = str(trow[c].value or "").strip()
                if cv and cv != 'nan' and '*' in cv:
                    cells.append((c, int(cv.replace('*',''))))
            cells.sort()
            for si, (c, n) in enumerate(cells):
                side = 'L' if c < 4 else 'R'
                rec = (n, si, b_idx, side)
                if dtype == 1: pc['d1'].append(rec)
                else: pc['d2'].append(rec)

# ═══ 尾2×区 交互 (无前视化: 用全样本看结构, 再用最近30期验证) ═══
print('═══ 尾2×区 交互命中率 (数据1星) ═══')
cross = defaultdict(lambda: [0,0])
for iss in issues:
    for n, si, b, side in period_cache[iss]['d1']:
        hit = 1 if n in draws[iss] else 0
        z = (n-1)//10
        t = n % 10
        if t == 2:
            cross[f"尾2-区{z}"][1]+=1; cross[f"尾2-区{z}"][0]+=hit
        else:
            cross[f"非尾2-区{z}"][1]+=1; cross[f"非尾2-区{z}"][0]+=hit
print(' 尾2 明细:')
for k in sorted(cross):
    if k.startswith('尾2'):
        h, n = cross[k]
        print(f'   {k}: n={n:4d} 率={h/n:.3f}')
print(' 非尾2 明细:')
for k in sorted(cross):
    if k.startswith('非尾2') and cross[k][1] > 0:
        h, n = cross[k]
        print(f'   {k}: n={n:4d} 率={h/n:.3f}')

# ═══ 终版规则集 (全部无前视验证) ═══
print('\n═══ 终版规则 最近30期 + 全期验证 ═══')
TEST = issues[30:]
RULES = {
    'R1 尾2区0-3+非重': lambda n, iss: (n%10==2 and (n-1)//10 in (0,1,2,3) and (iss-1 not in draws or n not in draws[iss-1])),
    'R2 尾2+非重': lambda n, iss: (n%10==2 and (iss-1 not in draws or n not in draws[iss-1])),
    'R3 右侧+尾2/7/8/3': lambda n, iss, side: (side=='R' and n%10 in (2,7,8,3)),
    'R4 尾2/7/8/3+非重': lambda n, iss: (n%10 in (2,7,8,3) and (iss-1 not in draws or n not in draws[iss-1])),
    'R5 非重+非61-70+非41-50': lambda n, iss: ((iss-1 not in draws or n not in draws[iss-1]) and not (61<=n<=70) and not (41<=n<=50)),
    'R6 尾2': lambda n, iss: n%10==2,
}
for name, fn in RULES.items():
    hits = tot = 0
    hits_recent = tot_recent = 0
    for iss in issues:
        cand = set()
        for n, si, b, side in period_cache[iss]['d1']:
            if name.startswith('R3'):
                if fn(n, iss, side): cand.add(n)
            elif fn(n, iss): cand.add(n)
        h = len(cand & draws[iss]); t = len(cand)
        hits += h; tot += t
        if iss >= issues[-30]:
            hits_recent += h; tot_recent += t
    print(f'  {name:18s}: 全期 {hits}/{tot}={hits/tot:.4f} | 近30期 {hits_recent}/{tot_recent}={hits_recent/tot_recent:.4f}')

# ═══ 终版选号器: 对当期推荐池排序 ═══
print('\n═══ 终版选号器 (无前视, 数据1∪数据2星去重) ═══')
TEST_START = issues[30]
test_issues = [i for i in issues if i >= TEST_START]

def logit(p):
    p = max(1e-6, min(1-1e-6, p))
    return math.log(p/(1-p))

def build_w(train_issues):
    fc = defaultdict(lambda: [0,0])
    for iss in train_issues:
        for n, si, b, side in period_cache[iss]['d1']:
            hit = 1 if n in draws[iss] else 0
            z = (n-1)//10; t = n%10
            fc[f"Z{z}尾{t}"][1]+=1; fc[f"Z{z}尾{t}"][0]+=hit
            fc[f"星{min(si,7)}"][1]+=1; fc[f"星{min(si,7)}"][0]+=hit
            if iss-1 in draws:
                fc[f"重{int(n in draws[iss-1])}"][1]+=1; fc[f"重{int(n in draws[iss-1])}"][0]+=hit
    base = sum(v[0] for v in fc.values())/sum(v[1] for v in fc.values())
    w = {}
    for k, (h, n) in fc.items():
        if n >= 20: w[k] = logit(h/n) - logit(base)
    return w, base

per = []
for iss in test_issues:
    train = [i for i in issues if i < iss]
    w, base = build_w(train)
    best = {}
    for src in ['d1','d2']:
        for n, si, b, side in period_cache[iss][src]:
            s = logit(base) + w.get(f"Z{(n-1)//10}尾{n%10}", 0) + w.get(f"星{min(si,7)}", 0)
            if iss-1 in draws:
                s += w.get(f"重{int(n in draws[iss-1])}", 0)
            if n not in best or s > best[n]: best[n] = s
    sc = sorted(best.items(), key=lambda kv: -kv[1])
    picks = [n for n, _ in sc[:12]]
    per.append(len(set(picks) & draws[iss]))
print(f'去重全星池 Top12: 期均 {statistics.mean(per):.2f} (池约{statistics.mean([len({n for n,si,b,side in period_cache[i]["d1"]} | {n for n,si,b,side in period_cache[i]["d2"]}) for i in test_issues]):.1f}码, 期均命中约5.1)')

# 随机对照
import random
rng = random.Random(5)
rnd = []
for iss in test_issues:
    pool = {n for n,si,b,side in period_cache[iss]['d1']} | {n for n,si,b,side in period_cache[iss]['d2']}
    picks = rng.sample(list(pool), 12)
    rnd.append(len(set(picks) & draws[iss]))
print(f'随机12码: 期均 {statistics.mean(rnd):.2f}')