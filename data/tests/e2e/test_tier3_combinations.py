# -*- coding: utf-8 -*-
"""
Tier 3: Cross-Subsystem Combinations & Feedback Loops (跨子系统协同与闭环自学习)
=============================================================================
验证 7 路子系统共识聚合、KillSeeker 杀号碰撞与共振打标、以及自主学习闭环状态机。
"""
import os
import sys
import unittest
import openpyxl

PROJ_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJ_DIR not in sys.path:
    sys.path.insert(0, PROJ_DIR)
BACKEND_DIR = os.path.join(PROJ_DIR, "backend")
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from backend.utils.paths import data_path
from backend.core.aggregation.consensus_engine import ConsensusEngine
from core.spatial_points import (
    load_draws_from_file as load_spatial_draws,
    calculate_spatial_point_features,
    rank_spatial_picks,
    cross_validate_spatial_picks,
)
from core.gold_pick2 import (
    load_draws_from_file as load_gold_draws,
    calculate_gold_pick2_features,
    cross_validate_pick2_picks,
    GoldPick2Learner,
    batch_update,
)
from core.follow_analysis import (
    load_draws_from_history as load_follow_draws,
    daily_follow_picks,
    cross_validate_follow_picks,
)
from core.formula_jingle.jingle_cross_validator import cross_validate_jingle
from backend.learning.autonomous_learner import AutonomousLearner


class TestTier3ConsensusAggregation(unittest.TestCase):
    """终审大团长 7 路共识汇总与 8 区空间平衡测试"""

    def setUp(self):
        self.engine = ConsensusEngine(PROJ_DIR)
        self.draws = self.engine.load_draws()

    def test_01_consensus_engine_execution(self):
        """验证 ConsensusEngine 聚合推演、8区覆盖评估与产物生成"""
        self.assertGreater(len(self.draws), 100)
        res = self.engine.run_aggregation(n_review=5)
        self.assertIsInstance(res, dict)
        self.assertIn("target_period", res)
        self.assertIn("consensus_dan_pool", res)
        self.assertIn("eight_zones_status", res)
        self.assertIn("stable_top10", res)
        self.assertIn("walk_forward", res)

        # 验证 8 区覆盖评估结构
        zone_status = res["eight_zones_status"]
        self.assertIn("full_coverage", zone_status)
        self.assertIn("miss_zones", zone_status)

        # 验证 Stable Top10 必须为 10 码且在 1..80 范围
        top10 = res["stable_top10"]
        self.assertEqual(len(top10), 10)
        self.assertTrue(all(1 <= x <= 80 for x in top10))

        # 验证输出文件已生成
        report_file = os.path.join(PROJ_DIR, "outputs", "aggregation", res["report_file"])
        self.assertTrue(os.path.exists(report_file))


class TestTier3CrossValidationAndKillCollisions(unittest.TestCase):
    """跨系统交叉风控、KillSeeker 撞车预警与共振打标测试"""

    def test_01_pick2_cross_validation_collision_tagging(self):
        """验证定金选2交叉风控检测与杀号避雷标签"""
        flags = cross_validate_pick2_picks(PROJ_DIR, golden=12, hot=28, top5_pairs=[])
        self.assertIn("safety_audit", flags)
        self.assertIn("golden_killed_by_killseeker", flags)

    def test_02_spatial_cross_validation_tagging(self):
        """验证空间点位系统跨子系统风控与共振打标"""
        draws = load_spatial_draws(data_path("kl8_history_final.txt"))
        pts_data = calculate_spatial_point_features(draws)
        picks = rank_spatial_picks(pts_data)
        cross_res = cross_validate_spatial_picks(PROJ_DIR, picks)
        self.assertIn("kill_conflicts", cross_res)
        self.assertIn("resonance_numbers", cross_res)
        self.assertIn("number_tags", cross_res)

    def test_03_follow_cross_validation_tagging(self):
        """验证跟随分析跨系统多维共振与冲突检测"""
        draws = load_follow_draws(data_path("kl8_history_final.txt"))
        picks = daily_follow_picks(draws)
        cross_res = cross_validate_follow_picks(PROJ_DIR, picks)
        self.assertIn("kill_conflicts", cross_res)
        self.assertIn("resonance_numbers", cross_res)
        self.assertIn("detailed_tags", cross_res)

    def test_04_jingle_cross_validation_with_custom_kill(self):
        """验证顺口溜口诀对杀号池号码的精准拦截打标"""
        rec_nums = [5, 12, 33, 44, 78]
        custom_kill = {12, 78}
        cross = cross_validate_jingle(rec_nums, custom_kill_set=custom_kill)
        self.assertIn(12, cross["clash_numbers"])
        self.assertIn(78, cross["clash_numbers"])
        self.assertNotIn(5, cross["clash_numbers"])
        self.assertEqual(len(cross["detailed_tags"]), 5)


class TestTier3AutonomousLearningFeedback(unittest.TestCase):
    """自主闭环学习状态机与防过拟合约束测试"""

    def test_01_gold_pick2_batch_update_threshold(self):
        """验证自学习更新必须满足样本容量阈值 (>=50 允许更新，<50 阻断更新)"""
        under_samples = [{"hit": True, "score": 0.8}] * 25
        self.assertFalse(batch_update(under_samples, min_batch=50))

        valid_samples = [{"hit": (i % 2 == 0), "score": 0.7} for i in range(60)]
        self.assertTrue(batch_update(valid_samples, min_batch=50))

    def test_02_learner_level3_freeze_protection(self):
        """验证在 Level 3 弱信号风险状态下参数强制冻结 (SKIPPED)"""
        learner = GoldPick2Learner()
        res = learner.update_weights_from_review(level=3, failure_modes=["GOLDEN_MISS"])
        self.assertEqual(res["status"], "SKIPPED")

    def test_03_autonomous_learner_state_machine(self):
        """验证主系统 AutonomousLearner 状态机初始化与当前状态读取"""
        learner = AutonomousLearner()
        state = learner.get_current_state()
        self.assertIsInstance(state, dict)
        self.assertIn("strategy_mode", state)
        self.assertIn("pentagon_weights", state)
        self.assertIn("environment_state", state)


class TestTier3ExcelDataETL(unittest.TestCase):
    """Excel 核心报表与格式化资产数据一致性校验"""

    def test_01_excel_master_sheet_integrity(self):
        """验证主数据表跟随+点位+开奖数据.xlsx与跟随号码统计工作表完整性"""
        excel_path = data_path("跟随+点位+开奖数据.xlsx")
        self.assertTrue(os.path.exists(excel_path), f"Master Excel not found: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        self.assertIn("跟随号码统计", sheet_names)
        ws = wb["跟随号码统计"]
        self.assertGreater(ws.max_row, 100)
        wb.close()


if __name__ == "__main__":
    unittest.main()
